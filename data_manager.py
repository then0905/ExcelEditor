import pandas as pd
import json
import os
import warnings
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border
import gc

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


class DataManager:
    def __init__(self, config_path="config.json"):
        self.config_path = config_path
        self._full_config = self._load_config(config_path)  # 完整配置（以 Excel 路徑為 key）
        self.config = {}  # 當前 Excel 的配置（指向 _full_config 的子 dict）
        self.excel_path = None
        self.master_dfs = {}  # 存放母表 DataFrame
        self.sub_dfs = {}  # 存放子表 DataFrame
        self.need_config_alert = False  # 標記是否需要彈出配置視窗
        self.dirty = False  # 標記資料是否有未儲存的變更
        self.sheet_styles = {}  # 存放各工作表的格式資訊

        # --- 外部文字表相關變數 ---
        self.text_df = None  # 存放文字表的完整 DataFrame
        self.text_dict = {}  # 快速查找用字典 {Key: Value}
        self.text_file_path = ""  # 文字表路徑
        self.text_key_col = "Key"  # 假設文字表的 Key 欄位名
        self.text_val_col = "Text"  # 假設文字表的 Value 欄位名
        self.text_modified = False  # 標記是否修改過
        self.text_sheetnames = []  # 文字表的工作表
        self.text_modifications = {}  # 記錄修改: {sheet_name: {key: new_value}}

        self._excel_file_handle = None  # 保存 Excel 文件句柄
        self._text_file_handle = None  # 保存文字表文件句柄

    def _load_config(self, path):
        if not os.path.exists(path):
            return {}
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}

    @staticmethod
    def _ws_to_dataframe(ws):
        """
        將 openpyxl Worksheet 轉為全字串 pandas DataFrame，
        模擬 pd.read_excel(..., dtype=str).fillna("") 的行為，
        但只需開一次檔案（避免雙重 I/O）。
        """
        import math
        from datetime import datetime, date, time as dtime

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return pd.DataFrame()

        # 建立 header（仿 pandas Unnamed: N 規則）
        header_row = rows[0]
        headers = []
        for i, h in enumerate(header_row):
            h_str = str(h).strip() if h is not None else ""
            headers.append(h_str if h_str else f"Unnamed: {i}")

        def _to_str(v):
            if v is None:
                return ""
            if isinstance(v, bool):
                return str(v)
            if isinstance(v, float):
                if math.isnan(v):
                    return ""
                # 整數值的 float（如 1.0）轉為 "1"，與 pd.read_excel dtype=str 一致
                if v == int(v):
                    return str(int(v))
                return str(v)
            if isinstance(v, int):
                return str(v)
            if isinstance(v, (datetime, date)):
                return v.strftime("%Y-%m-%d")
            if isinstance(v, dtime):
                return v.strftime("%H:%M:%S")
            return str(v)

        num_cols = len(headers)
        data = []
        for row in rows[1:]:
            # 補齊或截斷，確保與 header 長度一致
            padded = list(row) + [None] * (num_cols - len(row)) if len(row) < num_cols else list(row[:num_cols])
            data.append([_to_str(v) for v in padded])

        return pd.DataFrame(data, columns=headers)

    @staticmethod
    def _drop_empty_rows(df):
        """移除整列都是空白的行（向量化，比 apply per-row 快）。
        回傳 (filtered_df, non_empty_mask)，mask 可供 _capture_sheet_styles 重用。"""
        stripped = df.apply(lambda s: s.str.strip())
        mask = ~stripped.eq("").all(axis=1)
        return df[mask].reset_index(drop=True), mask

    def _get_col_type_map(self, sheet_name):
        """取得工作表各欄位的資料型別對應"""
        col_types = {}
        if "#" in sheet_name:
            parts = sheet_name.split("#", 1)
            master_name, sub_name = parts[0], parts[1]
            if master_name in self.config:
                cols = self.config[master_name].get("sub_sheets", {}).get(sub_name, {}).get("columns", {})
                for col_name, col_conf in cols.items():
                    col_types[col_name] = col_conf.get("type", "string")
        else:
            if sheet_name in self.config:
                cols = self.config[sheet_name].get("columns", {})
                for col_name, col_conf in cols.items():
                    col_types[col_name] = col_conf.get("type", "string")
        return col_types

    @staticmethod
    def _convert_value_for_excel(value, col_type):
        """根據欄位型別轉換值供 Excel 寫入"""
        if value == "" or value is None:
            return value
        try:
            if col_type == "int":
                return int(float(str(value)))
            elif col_type == "float":
                return float(str(value))
            elif col_type == "bool":
                return str(value).lower() in ('true', '1', 'yes')
        except (ValueError, TypeError):
            pass
        return value

    def _prepare_df_for_save(self, sheet_name, df):
        """儲存前根據 config 轉換 DataFrame 的數值欄位為正確型別"""
        col_types = self._get_col_type_map(sheet_name)
        if not col_types:
            return df
        df = df.copy()
        for col_name in df.columns:
            col_type = col_types.get(col_name, "string")
            if col_type in ("int", "float", "bool"):
                df[col_name] = df[col_name].apply(
                    lambda v, ct=col_type: self._convert_value_for_excel(v, ct)
                )
        return df

    @staticmethod
    def _copy_cell_style(cell):
        """複製儲存格的格式資訊"""
        return {
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "alignment": copy(cell.alignment),
            "border": copy(cell.border),
            "number_format": cell.number_format,
        }

    @staticmethod
    def _apply_cell_style(cell, style):
        """套用格式到儲存格"""
        cell.font = style["font"]
        cell.fill = style["fill"]
        cell.alignment = style["alignment"]
        cell.border = style["border"]
        cell.number_format = style["number_format"]

    def _capture_sheet_styles(self, sheet_name, df, ws, mask=None):
        """擷取工作表的格式資訊（背景色、字體、欄寬、列高等）。
        mask: 來自 _drop_empty_rows 的 non-empty mask，避免重複計算。"""
        styles = {
            "col_widths": {},
            "row_heights": {},
            "header_styles": {},
            "cell_styles": {},
        }
        num_cols = len(df.columns)

        # 欄寬
        for col_idx in range(1, num_cols + 1):
            col_letter = get_column_letter(col_idx)
            dim = ws.column_dimensions.get(col_letter)
            if dim and dim.width:
                styles["col_widths"][col_idx] = dim.width

        # 標題行格式 (Excel row 1)
        header_dim = ws.row_dimensions.get(1)
        if header_dim and header_dim.height:
            styles["row_heights"][1] = header_dim.height
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            styles["header_styles"][col_idx] = self._copy_cell_style(cell)

        # 找出存活的行（重用已計算的 mask，避免重複 strip+eq）
        if mask is None:
            stripped = df.apply(lambda s: s.str.strip())
            mask = ~stripped.eq("").all(axis=1)
        surviving_indices = df[mask].index.tolist()

        # 資料行格式
        for new_idx, orig_idx in enumerate(surviving_indices):
            orig_excel_row = orig_idx + 2   # DataFrame 0-based → Excel row (header=1)
            new_excel_row = new_idx + 2     # 重新排列後的 Excel row

            # 列高
            dim = ws.row_dimensions.get(orig_excel_row)
            if dim and dim.height:
                styles["row_heights"][new_excel_row] = dim.height

            # 各儲存格格式
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=orig_excel_row, column=col_idx)
                styles["cell_styles"][(new_excel_row, col_idx)] = self._copy_cell_style(cell)

        self.sheet_styles[sheet_name] = styles

    def _apply_sheet_styles(self, ws, sheet_name, df):
        """將儲存的格式套用到工作表"""
        if sheet_name not in self.sheet_styles:
            return

        styles = self.sheet_styles[sheet_name]
        num_cols = len(df.columns)
        max_data_row = len(df) + 1  # +1 因為 header 佔 row 1

        # 欄寬
        for col_idx, width in styles["col_widths"].items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        # 列高
        for excel_row, height in styles["row_heights"].items():
            if excel_row <= max_data_row:
                ws.row_dimensions[excel_row].height = height

        # 標題行格式
        for col_idx, style in styles["header_styles"].items():
            if col_idx <= num_cols:
                self._apply_cell_style(ws.cell(row=1, column=col_idx), style)

        # 資料儲存格格式
        for (excel_row, col_idx), style in styles["cell_styles"].items():
            if excel_row <= max_data_row and col_idx <= num_cols:
                self._apply_cell_style(ws.cell(row=excel_row, column=col_idx), style)

    def load_excel(self, file_path):
        # 先關閉之前的文件
        self.close_excel()

        self.excel_path = file_path
        self.need_config_alert = False
        self.master_dfs = {}
        self.sub_dfs = {}
        self.sheet_styles = {}

        # 從 _full_config 取出該 Excel 的獨立配置區段
        excel_key = os.path.normpath(file_path)
        if excel_key not in self._full_config:
            self._full_config[excel_key] = {}
        self.config = self._full_config[excel_key]

        # 清空文字表狀態
        self.text_dict = {}
        self.text_df = None
        self.text_file_path = ""
        self.text_modified = False
        self.text_sheetnames = []
        self.text_modifications = {}
        self._text_file_handle = None

        # 若該 Excel 的配置有文字表路徑，嘗試載入
        if "global_text_path" in self.config:
            self.load_external_text(self.config["global_text_path"])

        # 只開一次檔案：用 openpyxl 同時讀取資料與格式，避免雙重 I/O
        try:
            wb = load_workbook(file_path, data_only=True)
            try:
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    df = self._ws_to_dataframe(ws)

                    if sheet.endswith(".json"):
                        # 計算空行 mask 一次，同時供格式擷取和過濾使用
                        filtered_df, mask = self._drop_empty_rows(df)
                        self._capture_sheet_styles(sheet, df, ws, mask=mask)
                        self.master_dfs[sheet] = filtered_df

                        if sheet not in self.config:
                            self.config[sheet] = {
                                "use_icon": False,
                                "image_path": "",
                                "classification_key": self.master_dfs[sheet].columns[0],
                                "primary_key": self.master_dfs[sheet].columns[0],
                                "columns": {col: {"type": "string"} for col in self.master_dfs[sheet].columns},
                                "sub_sheets": {}
                            }
                            self.need_config_alert = True

                    elif "#" in sheet:
                        filtered_df, mask = self._drop_empty_rows(df)
                        self._capture_sheet_styles(sheet, df, ws, mask=mask)
                        self.sub_dfs[sheet] = filtered_df

            finally:
                wb.close()

            if self.need_config_alert:
                self.save_config()

        except Exception as e:
            print(f"載入 Excel 失敗: {e}")
            raise

    def close_excel(self):
        """關閉 Excel 文件並清理資源"""
        if self._excel_file_handle is not None:
            try:
                self._excel_file_handle.close()
            except Exception as e:
                # 即使失敗也要繼續，避免阻斷清理流程
                print(f"關閉 Excel 失敗: {e}")
            finally:
                # 無論成功與否，都清空引用（防止記憶體洩漏）
                self._excel_file_handle = None

        gc.collect()

    def load_external_text(self, path, key_col="TextID", val_col="TextContent"):
        # 先關閉之前的文字表
        self.close_text_file()

        if not os.path.exists(path):
            print(f"文字表路徑不存在: {path}")
            return False

        try:
            self.text_file_path = path
            self.text_sheetnames = []
            self.text_dict = {}

            # 使用 openpyxl read_only 模式，比 pandas 快 2-5x
            wb = load_workbook(path, data_only=True, read_only=True)
            try:
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    self.text_sheetnames.append(sheet_name)

                    rows = ws.iter_rows(values_only=True)
                    # 讀取 header
                    header = next(rows, None)
                    if header is None:
                        continue

                    header = [str(h).strip() if h is not None else "" for h in header]

                    # 判斷 key/val 欄位索引
                    if key_col in header and val_col in header:
                        k_idx = header.index(key_col)
                        v_idx = header.index(val_col)
                        k_name = key_col
                        v_name = val_col
                    elif len(header) >= 2:
                        print(f"[{sheet_name}] 缺少欄位，改用前兩欄")
                        k_idx = 0
                        v_idx = 1
                        k_name = header[0]
                        v_name = header[1]
                    else:
                        continue

                    for row in rows:
                        if k_idx >= len(row) or v_idx >= len(row):
                            continue
                        k_val = row[k_idx]
                        v_val = row[v_idx]
                        key_str = str(k_val) if k_val is not None else ""
                        val_str = str(v_val) if v_val is not None else ""
                        if key_str:
                            self.text_dict[key_str] = {
                                "value": val_str,
                                "sheet": sheet_name,
                                "key_col": k_name,
                                "val_col": v_name
                            }
            finally:
                wb.close()

            return True

        except Exception as e:
            print(f"載入文字表失敗: {e}")
            self.close_text_file()
            return False

    def close_text_file(self):
        """關閉文字表文件並清理資源"""
        if self._text_file_handle is not None:
            try:
                self._text_file_handle.close()
            except:
                pass
            self._text_file_handle = None

        gc.collect()

    def save_config(self):
        with open(self.config_path, 'w', encoding='utf-8') as f:
            json.dump(self._full_config, f, indent=4, ensure_ascii=False)

    def save_excel(self):
        """
        儲存正在編輯的Excel
        """
        if not self.excel_path:
            return

        try:
            # 儲存前清除空白行
            for sheet in self.master_dfs:
                self.master_dfs[sheet], _ = self._drop_empty_rows(self.master_dfs[sheet])
            for sheet in self.sub_dfs:
                self.sub_dfs[sheet], _ = self._drop_empty_rows(self.sub_dfs[sheet])

            if not os.path.exists(self.excel_path):
                with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='w') as writer:
                    for sheet, df in self.master_dfs.items():
                        self._prepare_df_for_save(sheet, df).to_excel(writer, sheet_name=sheet, index=False)
                    for sheet, df in self.sub_dfs.items():
                        self._prepare_df_for_save(sheet, df).to_excel(writer, sheet_name=sheet, index=False)
                    # 套用儲存的格式
                    for sheet_name in list(self.master_dfs) + list(self.sub_dfs):
                        if sheet_name in writer.sheets:
                            df = self.master_dfs.get(sheet_name) or self.sub_dfs.get(sheet_name)
                            self._apply_sheet_styles(writer.sheets[sheet_name], sheet_name, df)
            else:
                # 先關閉現有句柄
                self.close_excel()

                wb = load_workbook(self.excel_path)
                for sheet_name, df in self.master_dfs.items():
                    self._update_sheet_content(wb, sheet_name, df)
                for sheet_name, df in self.sub_dfs.items():
                    self._update_sheet_content(wb, sheet_name, df)
                wb.save(self.excel_path)
                wb.close()  # 確保關閉

            # 存外部文字表
            if getattr(self, "text_modified", False) and getattr(self, "text_file_path", None):
                try:
                    print(f"正在同步儲存外部文字表至: {self.text_file_path}")
                    self._save_external_text()
                    self.text_modified = False
                    self.text_modifications = {}
                    print("外部文字表儲存成功")
                except Exception as e:
                    print(f"外部文字表儲存失敗: {e}")

            # 強制垃圾回收
            gc.collect()
            self.dirty = False

        except Exception as e:
            print(f"儲存失敗: {e}")
            raise e

    def _save_external_text(self):
        """
        儲存外部文字表：只修改有變動的儲存格，保留格式。
        使用按 sheet 分組 + 行索引 O(1) 查找，取代逐 key 線性掃描。
        """
        if not self.text_file_path or not os.path.exists(self.text_file_path):
            return

        # 先關閉文字表句柄
        self.close_text_file()

        # 按 sheet 分組 modifications
        mods_by_sheet = {}
        for key, new_val in self.text_modifications.items():
            info = self.text_dict.get(key)
            if info:
                mods_by_sheet.setdefault(info["sheet"], []).append((key, new_val, info))

        if not mods_by_sheet:
            return

        wb = load_workbook(self.text_file_path)

        try:
            for sheet_name, entries in mods_by_sheet.items():
                if sheet_name not in wb.sheetnames:
                    print(f"警告: 工作表 {sheet_name} 不存在於文字表中")
                    continue

                ws = wb[sheet_name]

                # 從 header 找出 key/val 欄位索引
                key_col_name = entries[0][2]["key_col"]
                val_col_name = entries[0][2]["val_col"]
                key_col_idx = None
                val_col_idx = None

                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value == key_col_name:
                        key_col_idx = col_idx
                    if cell.value == val_col_name:
                        val_col_idx = col_idx

                if key_col_idx is None or val_col_idx is None:
                    print(f"警告: 在 {sheet_name} 中找不到欄位 {key_col_name} 或 {val_col_name}")
                    continue

                # 建立行索引（一次掃描 O(n)）
                row_index = {}
                for r in range(2, ws.max_row + 1):
                    cell_key = ws.cell(r, key_col_idx).value
                    if cell_key is not None:
                        row_index[str(cell_key)] = r

                # O(1) 更新每個 modification
                for key, new_val, info in entries:
                    if key in row_index:
                        ws.cell(row_index[key], val_col_idx).value = new_val
                    else:
                        print(f"警告: 在 {sheet_name} 中找不到 key={key}")

            wb.save(self.text_file_path)
        finally:
            wb.close()
            gc.collect()

    def _update_sheet_content(self, wb, sheet_name, df):
        """
        核心邏輯：
        1. 找到工作表 (如果沒有就建立)
        2. 填入標題 (Header)
        3. 填入數據 (Body)
        4. 清除多餘的舊資料 (值與格式)
        5. 套用儲存的格式資訊
        """
        # 1. 取得或建立 Worksheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name

        col_types = self._get_col_type_map(sheet_name)

        current_row_idx = 2
        for row in df.itertuples(index=False):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row_idx, column=col_idx)
                col_name = df.columns[col_idx - 1]
                cell.value = self._convert_value_for_excel(value, col_types.get(col_name, "string"))
            current_row_idx += 1

        # 清除多餘的舊資料（值與格式）
        max_row_in_excel = ws.max_row
        if max_row_in_excel >= current_row_idx:
            for r in range(current_row_idx, max_row_in_excel + 1):
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.value = None
                    cell.font = Font()
                    cell.fill = PatternFill()
                    cell.alignment = Alignment()
                    cell.border = Border()
                    cell.number_format = 'General'

        # 套用儲存的格式
        self._apply_sheet_styles(ws, sheet_name, df)

    def update_cell(self, is_sub, sheet_name, row_idx, col_name, value):
        """
        當 UI 修改時呼叫此方法更新 DataFrame
        修正：依照 Config 的設定來轉型，而不是依賴 DataFrame 原本的 type
        """
        target_dict = self.sub_dfs if is_sub else self.master_dfs

        if sheet_name in target_dict:
            df = target_dict[sheet_name]

            try:
                col_type = "string"

                if not is_sub:
                    if sheet_name in self.config:
                        col_type = self.config[sheet_name]["columns"].get(col_name, {}).get("type", "string")
                else:
                    master_name = sheet_name.split("#")[0]
                    sub_name = sheet_name.split("#")[1]
                    if master_name in self.config:
                        col_type = self.config[master_name]["sub_sheets"].get(sub_name, {}) \
                            .get("columns", {}).get(col_name, {}).get("type", "string")

                if col_type == "int":
                    value = int(value)
                elif col_type == "float":
                    value = float(value)
                elif col_type == "bool":
                    if isinstance(value, str):
                        value = value.lower() in ['true', '1', 'yes']
                    else:
                        value = bool(value)

            except Exception as e:
                pass

            col_conf = self.config.get(sheet_name, {}).get("columns", {}).get(col_name, {})
            if col_conf.get("link_to_text"):
                raw_key = self.master_dfs[sheet_name].at[row_idx, col_name]
                self._update_external_text(raw_key, value)
            else:
                df.at[row_idx, col_name] = value

            self.dirty = True

    def _update_external_text(self, key, new_value):
        """
        內部方法：記錄文字表的修改
        """
        if not hasattr(self, "text_modifications"):
            self.text_modifications = {}

        self.text_modifications[str(key)] = str(new_value)

        if str(key) in self.text_dict:
            self.text_dict[str(key)]["value"] = str(new_value)

        self.text_modified = True

    def get_text_value(self, key):
        if not self.text_dict:
            return key

        text_info = self.text_dict.get(str(key))
        if text_info:
            return text_info["value"]
        else:
            return f"<{key} Missing>"

    def update_linked_text(self, key, new_text):
        """ 給 UI 呼叫：更新文字內容 (不改 Key) """
        if not hasattr(self, "text_dict") or self.text_dict is None:
            return

        self._update_external_text(key, new_text)
        self.dirty = True

    def cleanup(self):
        """清理所有資源"""
        self.close_excel()
        self.close_text_file()

        # 清空所有 DataFrame 與格式資料
        self.master_dfs.clear()
        self.sub_dfs.clear()
        self.text_dict.clear()
        self.sheet_styles.clear()

        # 強制垃圾回收
        gc.collect()

    def __del__(self):
        """析構函數：確保資源釋放"""
        self.cleanup()