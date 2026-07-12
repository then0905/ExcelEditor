#!/usr/bin/env python3
"""JsonEditor Pro — PySide6 · App-quality dark UI (Spec v2)"""

import sys, os, json, re, uuid, traceback, datetime
import pandas as pd

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QSplitter, QTabWidget, QTabBar,
    QListWidget, QListWidgetItem, QLineEdit, QPushButton, QLabel,
    QVBoxLayout, QHBoxLayout, QScrollArea, QCheckBox, QComboBox,
    QTextEdit, QTableView, QHeaderView, QAbstractItemView,
    QStackedWidget, QFileDialog, QMenu, QSizePolicy, QFrame,
    QInputDialog, QMessageBox, QStyledItemDelegate, QStyle,
    QAbstractItemDelegate,
    QDialog, QDialogButtonBox, QFormLayout, QLayout, QCompleter,
    QTableWidget, QTableWidgetItem, QPlainTextEdit,
    QColorDialog, QSpinBox, QRadioButton, QButtonGroup, QToolButton,
    QListView, QGridLayout,
)
from PySide6.QtCore import (
    Qt, Signal, QAbstractTableModel, QModelIndex,
    QTimer, QSize, QThread, QRect, QRectF, QPoint, QStringListModel,
)
from PySide6.QtGui import (
    QAction, QColor, QKeySequence, QFont, QBrush,
    QPainter, QPen, QLinearGradient, QPainterPath,
    QIcon, QPixmap, QFontDatabase, QCursor,
)

from json_data_manager import JsonDataManager
from validation import (new_rule as _v_new_rule, normalize_rule as _v_norm,
                        OPS as _V_OPS, OP_LABELS as _V_OP_LABELS,
                        COUNT_OP_LABELS as _V_COUNT_OPS,
                        DEFAULT_COLOR as _V_DEFAULT_COLOR)


# ═══════════════════════════════════════════════════════════════
#  DESIGN SYSTEM
# ═══════════════════════════════════════════════════════════════

_C = {
    "bg":      "#0F0F1A",
    "sidebar": "#14141F",
    "panel":   "#17172A",
    "card":    "#1E1E32",
    "cardH":   "#252540",
    "input":   "#1C1C30",
    "code":    "#12121F",
    "border":  "#2A2A40",
    "borderH": "#3A3A54",
    "accent":  "#6366F1",
    "txt":     "#F0F0F5",
    "txt2":    "#8B8BA3",
    "txt3":    "#5A5A72",
    "txtAcc":  "#A5B4FC",
    "green":   "#10B981",
    "yellow":  "#EAB308",
    "red":     "#EF4444",
    "cyan":    "#0891B2",
}

_CAT = [
    {"color": "#F59E0B", "r": 245, "g": 158, "b": 11,  "text": "#FDE68A"},
    {"color": "#8B5CF6", "r": 139, "g": 92,  "b": 246, "text": "#C4B5FD"},
    {"color": "#10B981", "r": 16,  "g": 185, "b": 129, "text": "#6EE7B7"},
    {"color": "#EC4899", "r": 236, "g": 72,  "b": 153, "text": "#F9A8D4"},
    {"color": "#6366F1", "r": 99,  "g": 102, "b": 241, "text": "#A5B4FC"},
    {"color": "#EF4444", "r": 239, "g": 68,  "b": 68,  "text": "#FCA5A5"},
    {"color": "#0891B2", "r": 8,   "g": 145, "b": 178, "text": "#67E8F9"},
    {"color": "#EAB308", "r": 234, "g": 179, "b": 8,   "text": "#FDE68A"},
]

_cat_assign: dict[str, int] = {}


def _cat_for(val: str) -> dict:
    s = str(val)
    if s not in _cat_assign:
        _cat_assign[s] = len(_cat_assign) % len(_CAT)
    return _CAT[_cat_assign[s]]


def _cat_qcolor(val: str, alpha: int = 255) -> QColor:
    c = _cat_for(val)
    return QColor(c["r"], c["g"], c["b"], alpha)


# ── Global QSS ────────────────────────────────────────────────────────────────

APP_QSS = f"""
* {{
    font-family: "Segoe UI", "Noto Sans TC", sans-serif;
    font-size: 12px;
    color: {_C['txt']};
}}
QMainWindow, QDialog {{ background: {_C['bg']}; }}
QWidget {{ background: transparent; }}

QSplitter::handle:horizontal {{ background: {_C['border']}; width: 1px; }}
QSplitter::handle:vertical   {{ background: {_C['border']}; height: 1px; }}

/* ── Outer tab bar (table switcher) ── */
QTabWidget#main-tabs::pane {{
    border: none;
    background: {_C['bg']};
}}
QTabWidget#main-tabs > QTabBar {{
    background: {_C['sidebar']};
    border-bottom: 1px solid {_C['border']};
}}
QTabWidget#main-tabs > QTabBar::tab {{
    background: transparent;
    color: {_C['txt2']};
    padding: 9px 20px;
    border: none;
    border-bottom: 2px solid transparent;
    font-size: 12px;
    font-weight: 500;
    min-width: 60px;
}}
QTabWidget#main-tabs > QTabBar::tab:selected {{
    color: {_C['txt']};
    border-bottom: 2px solid {_C['accent']};
    font-weight: 600;
}}
QTabWidget#main-tabs > QTabBar::tab:hover:!selected {{
    color: {_C['txt']};
    background: rgba(255,255,255,0.03);
}}

/* ── Sub-table tab bar ── */
QTabWidget#sub-tabs::pane {{
    border: none;
    background: {_C['panel']};
}}
QTabWidget#sub-tabs > QTabBar {{
    background: {_C['panel']};
    border-bottom: 1px solid {_C['border']};
}}
QTabWidget#sub-tabs > QTabBar::tab {{
    background: transparent;
    color: {_C['txt2']};
    padding: 6px 14px;
    border: none;
    border-bottom: 2px solid transparent;
    font-size: 11px;
}}
QTabWidget#sub-tabs > QTabBar::tab:selected {{
    color: {_C['accent']};
    border-bottom: 2px solid {_C['accent']};
}}
QTabWidget#sub-tabs > QTabBar::tab:hover:!selected {{
    color: {_C['txt']};
}}

/* ── Classification list ── */
QListWidget#cls-list {{
    background: transparent;
    border: none;
    outline: none;
    padding: 4px;
}}
QListWidget#cls-list::item {{
    padding: 8px 10px;
    border-radius: 7px;
    margin: 1px 0;
}}
QListWidget#cls-list::item:selected {{
    background: rgba(99,102,241,0.18);
}}
QListWidget#cls-list::item:hover:!selected {{
    background: rgba(255,255,255,0.04);
}}

/* ── Card list ── */
QListWidget#card-list {{
    background: transparent;
    border: none;
    outline: none;
}}
QListWidget#card-list::item {{
    background: transparent;
    border: none;
    padding: 0;
    margin: 0;
}}

/* ── Inputs ── */
QLineEdit {{
    background: {_C['input']};
    border: 1px solid {_C['border']};
    border-radius: 7px;
    padding: 7px 10px;
    color: {_C['txt']};
    selection-background-color: {_C['accent']};
}}
QLineEdit:focus {{ border-color: {_C['accent']}; }}
QLineEdit[invalid="true"] {{
    border-color: {_C['red']};
    background: rgba(239,68,68,0.08);
}}
QTextEdit {{
    background: {_C['input']};
    border: 1px solid {_C['border']};
    border-radius: 7px;
    padding: 6px 10px;
    color: {_C['txt']};
    selection-background-color: {_C['accent']};
}}
QTextEdit:focus {{ border-color: {_C['accent']}; }}
QPlainTextEdit {{
    background: {_C['input']};
    border: 1px solid {_C['border']};
    border-radius: 7px;
    padding: 6px 10px;
    color: {_C['txt']};
    selection-background-color: {_C['accent']};
}}
QPlainTextEdit:focus {{ border-color: {_C['accent']}; }}
QSpinBox {{
    background: {_C['input']};
    border: 1px solid {_C['border']};
    border-radius: 7px;
    padding: 4px 6px;
    color: {_C['txt']};
    selection-background-color: {_C['accent']};
}}
QSpinBox:focus {{ border-color: {_C['accent']}; }}
QSpinBox::up-button, QSpinBox::down-button {{
    background: transparent;
    border: none;
    width: 16px;
}}
QSpinBox::up-arrow {{
    image: none;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-bottom: 5px solid {_C['txt2']};
}}
QSpinBox::down-arrow {{
    image: none;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 5px solid {_C['txt2']};
}}
QRadioButton {{ color: {_C['txt']}; background: transparent; spacing: 6px; }}
QRadioButton::indicator {{
    width: 14px; height: 14px;
    border-radius: 8px;
    border: 1px solid {_C['border']};
    background: {_C['input']};
}}
QRadioButton::indicator:hover {{ border-color: {_C['borderH']}; }}
QRadioButton::indicator:checked {{
    border: 4px solid {_C['accent']};
    background: #FFFFFF;
}}
QComboBox QLineEdit {{
    background: transparent;
    border: none;
    border-radius: 0;
    padding: 0 2px;
    color: {_C['txt']};
}}
QTextEdit#code-view {{
    background: {_C['code']};
    border: 1px solid {_C['border']};
    font-family: "Consolas", monospace;
    font-size: 11px;
    color: {_C['txt2']};
    border-radius: 8px;
    padding: 10px 12px;
}}
QComboBox {{
    background: {_C['input']};
    border: 1px solid {_C['border']};
    border-radius: 7px;
    padding: 6px 10px;
    color: {_C['txt']};
}}
QComboBox::drop-down {{ border: none; width: 20px; }}
QComboBox::down-arrow {{
    image: none;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 5px solid {_C['txt2']};
    margin-right: 6px;
}}
QComboBox:focus {{ border-color: {_C['accent']}; }}
QComboBox:disabled {{ color: {_C['txt3']}; }}
QComboBox QAbstractItemView {{
    background-color: {_C['card']};
    border: 1px solid {_C['borderH']};
    color: {_C['txt']};
    selection-background-color: {_C['accent']};
    selection-color: #FFFFFF;
    outline: none;
    padding: 2px;
}}
QComboBox QAbstractItemView::item {{
    color: {_C['txt']};
    background-color: transparent;
    padding: 5px 10px;
    min-height: 22px;
}}
QComboBox QAbstractItemView::item:hover {{
    background-color: {_C['cardH']};
    color: {_C['txt']};
}}
QComboBox QAbstractItemView::item:selected {{
    background-color: {_C['accent']};
    color: #FFFFFF;
}}
QCheckBox::indicator {{
    width: 15px; height: 15px;
    border: 1px solid {_C['border']};
    background: {_C['input']};
    border-radius: 4px;
}}
QCheckBox::indicator:checked {{
    background: {_C['accent']};
    border-color: {_C['accent']};
}}

/* ── Buttons ── */
QPushButton {{
    background: {_C['card']};
    border: 1px solid {_C['border']};
    border-radius: 7px;
    padding: 6px 14px;
    color: {_C['txt2']};
    font-weight: 500;
}}
QPushButton:hover {{
    background: {_C['cardH']};
    border-color: {_C['borderH']};
    color: {_C['txt']};
}}
QPushButton:pressed {{ background: rgba(99,102,241,0.15); }}
QPushButton[role="primary"] {{
    background: {_C['accent']};
    border-color: {_C['accent']};
    color: white;
    font-weight: 600;
}}
QPushButton[role="primary"]:hover {{ background: #5558E0; border-color: #5558E0; }}
QPushButton[role="danger"] {{
    color: {_C['red']};
    border-color: rgba(239,68,68,0.35);
    background: transparent;
}}
QPushButton[role="danger"]:hover {{
    background: rgba(239,68,68,0.12);
    border-color: {_C['red']};
    color: {_C['red']};
}}
QPushButton[role="success"] {{
    background: {_C['green']};
    border-color: {_C['green']};
    color: white;
    font-weight: 600;
}}
QPushButton[role="success"]:hover {{ background: #0ea371; }}
QPushButton[role="ghost"] {{
    background: transparent;
    border: none;
    color: {_C['txt2']};
    padding: 4px 10px;
    border-radius: 6px;
}}
QPushButton[role="ghost"]:hover {{
    background: rgba(255,255,255,0.06);
    color: {_C['txt']};
}}

/* ── Table ── */
QTableView {{
    background: {_C['panel']};
    gridline-color: {_C['border']};
    border: none; outline: none;
    selection-background-color: rgba(99,102,241,0.2);
    color: {_C['txt']};
}}
QTableView::item:hover:!selected {{ background: rgba(255,255,255,0.04); }}
QHeaderView::section {{
    background: {_C['sidebar']};
    color: {_C['txt2']};
    padding: 5px 8px;
    border: none;
    border-right: 1px solid {_C['border']};
    border-bottom: 1px solid {_C['border']};
    font-size: 11px;
    font-weight: 600;
}}
QHeaderView::section:hover {{ color: {_C['txt']}; }}
QHeaderView::section:pressed {{ background: rgba(99,102,241,0.15); color: {_C['accent']}; }}

/* ── Scrollbars ── */
QScrollBar:vertical   {{ background: transparent; width: 8px; border: none; margin: 0; }}
QScrollBar:horizontal {{ background: transparent; height: 8px; border: none; margin: 0; }}
QScrollBar::handle:vertical   {{ background: rgba(255,255,255,0.12); min-height: 24px; border-radius: 4px; margin: 2px; }}
QScrollBar::handle:horizontal {{ background: rgba(255,255,255,0.12); min-width: 24px;  border-radius: 4px; margin: 2px; }}
QScrollBar::handle:vertical:hover, QScrollBar::handle:horizontal:hover {{ background: rgba(255,255,255,0.25); }}
QScrollBar::add-line, QScrollBar::sub-line {{ width: 0; height: 0; }}
QScrollBar::add-page, QScrollBar::sub-page {{ background: transparent; }}

/* ── Menus ── */
QMenu {{
    background: {_C['panel']};
    border: 1px solid {_C['border']};
    border-radius: 8px;
    padding: 4px;
}}
QMenu::item {{ padding: 6px 20px; border-radius: 5px; color: {_C['txt2']}; }}
QMenu::item:selected {{ background: rgba(99,102,241,0.15); color: {_C['txt']}; }}
QMenu::separator {{ height: 1px; background: {_C['border']}; margin: 3px 8px; }}

/* ── Dialogs ── */
QDialog {{ background: {_C['panel']}; }}
QScrollArea {{ border: none; background: transparent; }}
QScrollArea > QWidget > QWidget {{ background: transparent; }}
QPushButton::menu-indicator {{ image: none; width: 0; }}
"""


# ── Helpers ───────────────────────────────────────────────────────────────────

# ── Tabler icon font (bundled) ──────────────────────────────────────────────
_TI_HEX = {
    "plus": "eb0b", "circle-plus": "ea69", "copy": "ea7a", "trash": "eb41",
    "table": "eba1", "columns": "eb83", "arrow-up": "ea25",
    "arrow-down": "ea16", "chevron-down": "ea5f", "folder-plus": "eaab",
    "pencil": "eb04", "note": "eb6d", "clipboard-plus": "efb2",
}
_TI_CODES = {k: chr(int(v, 16)) for k, v in _TI_HEX.items()}
_ti_family = None
_ti_cache = {}

def _resource_path(rel: str) -> str:
    bases = []
    try:
        bases.append(os.path.dirname(os.path.abspath(__file__)))
    except NameError:
        pass
    bases.append(os.path.dirname(os.path.abspath(sys.argv[0])))
    bases.append(os.getcwd())
    for b in bases:
        p = os.path.join(b, rel)
        if os.path.exists(p):
            return p
    return os.path.join(bases[0], rel)

def _ti_icon(name: str, color: str = None, px: int = 18) -> QIcon:
    """Render a Tabler glyph into a QIcon so the label keeps the normal font."""
    global _ti_family
    if _ti_family is None:
        fid = QFontDatabase.addApplicationFont(_resource_path("assets/tabler-icons.ttf"))
        fams = QFontDatabase.applicationFontFamilies(fid) if fid != -1 else []
        _ti_family = fams[0] if fams else ""
    color = color or _C["txt2"]
    ch = _TI_CODES.get(name)
    if not ch or not _ti_family:
        return QIcon()
    key = (name, color, px)
    if key in _ti_cache:
        return _ti_cache[key]
    pm = QPixmap(px, px)
    pm.fill(Qt.transparent)
    p = QPainter(pm)
    f = QFont(_ti_family)
    f.setPixelSize(px)
    p.setFont(f)
    p.setPen(QColor(color))
    p.drawText(pm.rect(), Qt.AlignCenter, ch)
    p.end()
    ic = QIcon(pm)
    _ti_cache[key] = ic
    return ic

_ROLE_ICON_COLOR = {"success": "#10B981", "danger": "#EF4444", "primary": "white"}

def _mk_btn(text: str, role: str = "", icon: str = None,
            icon_color: str = None) -> QPushButton:
    b = QPushButton(text)
    if role:
        b.setProperty("role", role)
    if icon:
        b.setIcon(_ti_icon(icon, icon_color or _ROLE_ICON_COLOR.get(role)))
    return b


def _hsep() -> QFrame:
    f = QFrame()
    f.setFrameShape(QFrame.HLine)
    f.setFixedHeight(1)
    f.setStyleSheet(f"background: {_C['border']}; border: none;")
    return f


def _vsep() -> QFrame:
    f = QFrame()
    f.setFrameShape(QFrame.VLine)
    f.setFixedWidth(1)
    f.setStyleSheet(f"background: {_C['border']}; border: none;")
    return f


def _sec_lbl(text: str) -> QLabel:
    lbl = QLabel(text.upper())
    lbl.setStyleSheet(
        f"color: {_C['txt3']}; font-size: 10px; font-weight: 600; "
        f"letter-spacing: 1px; padding: 10px 12px 5px; background: transparent;"
    )
    return lbl


def _json_highlight(raw: str) -> str:
    """JSON string → HTML with syntax highlighting."""
    import html as _html
    t = _html.escape(raw)
    t = re.sub(r'"([^"]+)":', f'<span style="color:#93C5FD">"\\1"</span>:', t)
    t = re.sub(r': "([^"]*)"', f': <span style="color:#86EFAC">"\\1"</span>', t)
    t = re.sub(r': (-?\d+\.?\d*)', f': <span style="color:#FBBF24">\\1</span>', t)
    t = re.sub(r'\b(true|false|null)\b', f'<span style="color:#F9A8D4">\\1</span>', t)
    return t


# ── Background workers ────────────────────────────────────────────────────────

class _LoadWorker(QThread):
    done  = Signal()
    error = Signal(str)

    def __init__(self, manager, path):
        super().__init__()
        self._manager = manager
        self._path    = path

    def run(self):
        try:
            self._manager.load_json(self._path)
            self.done.emit()
        except Exception as e:
            self.error.emit(str(e))


class _SaveWorker(QThread):
    done  = Signal()
    error = Signal(str)

    def __init__(self, manager):
        super().__init__()
        self._manager = manager

    def run(self):
        try:
            self._manager.save_json()
            self.done.emit()
        except Exception as e:
            self.error.emit(str(e))


# ── Image thumbnail helper ────────────────────────────────────────────────────

def _update_img_thumb(path_str: str, label: "QLabel", base_dir: "str | None") -> None:
    """Load an image into a QLabel thumbnail, resolving relative paths."""
    from PySide6.QtGui import QPixmap
    if not path_str:
        label.setPixmap(QPixmap())
        label.setText("No Image")
        return
    p = path_str if os.path.isabs(path_str) else (
        os.path.join(base_dir, path_str) if base_dir else path_str
    )
    p = os.path.normpath(p)
    px = QPixmap(p)
    if px.isNull():
        label.setPixmap(QPixmap())
        label.setText(f"找不到圖片:\n{p}")
        label.setWordWrap(True)
    else:
        label.setText("")
        w = label.width() or 220
        h = label.height() or 90
        label.setPixmap(px.scaled(w, h, Qt.KeepAspectRatio, Qt.SmoothTransformation))


# ── Non-scroll ComboBox (used in config dialog) ───────────────────────────────

# Dropdown-popup stylesheet applied DIRECTLY to the popup views: app-level QSS
# does not reliably cascade into floating popups on Windows (native white list
# with light text = unreadable), so every combo must get this explicitly.
_POPUP_QSS = (
    f"QListView {{"
    f"  background-color: {_C['card']};"
    f"  color: {_C['txt']};"
    f"  border: 1px solid {_C['borderH']};"
    f"  outline: none;"
    f"}}"
    f"QListView::item {{"
    f"  color: {_C['txt']};"
    f"  background-color: transparent;"
    f"  padding: 4px 10px;"
    f"  min-height: 22px;"
    f"}}"
    f"QListView::item:hover {{"
    f"  background-color: {_C['cardH']};"
    f"  color: {_C['txt']};"
    f"}}"
    f"QListView::item:selected {{"
    f"  background-color: {_C['accent']};"
    f"  color: #FFFFFF;"
    f"}}"
)


class _NoscrollCombo(QComboBox):
    """ComboBox that ignores scroll wheel unless the user explicitly clicked into it."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setFocusPolicy(Qt.StrongFocus)   # only gain focus via click/tab, NOT wheel
        self.view().setStyleSheet(_POPUP_QSS)

    def setEditable(self, editable):
        super().setEditable(editable)
        if editable:
            # editable combos pop a SEPARATE completer list (top-level, native
            # white) — its view is lazily created, so hand it a pre-styled one
            # or typing shows white-on-white suggestions
            comp = self.completer()
            if comp is not None:
                lv = QListView(self)
                lv.setStyleSheet(_POPUP_QSS)
                comp.setPopup(lv)

    def wheelEvent(self, e):
        if self.hasFocus():
            super().wheelEvent(e)
        else:
            e.ignore()


# ── ItemCardDelegate ──────────────────────────────────────────────────────────

class ItemCardDelegate(QStyledItemDelegate):
    """Draws each list item as a card with left category-color strip."""

    CARD_H  = 66   # card height
    PAD_V   = 4    # vertical outer padding (space between cards)
    PAD_H   = 10   # horizontal outer padding
    STRIP_W = 4    # left color strip width
    PAD_IN  = 12   # internal padding after strip

    R_PK  = Qt.UserRole + 1   # primary key string
    R_SUB = Qt.UserRole + 2   # subtitle string
    R_CAT = Qt.UserRole + 3   # category value (for color)
    R_VIO = Qt.UserRole + 4   # validation severity: "error" / "warn" / None

    def sizeHint(self, option, index):
        return QSize(option.rect.width(), self.CARD_H + self.PAD_V * 2)

    def paint(self, painter, option, index):
        painter.save()
        painter.setRenderHint(QPainter.Antialiasing)

        pk_val   = index.data(self.R_PK)  or ""
        subtitle = index.data(self.R_SUB) or ""
        cat_val  = index.data(self.R_CAT) or ""
        cat      = _cat_for(cat_val) if cat_val else _CAT[0]
        cat_qc   = QColor(cat["color"])
        cat_bg   = QColor(cat["r"], cat["g"], cat["b"], 38)   # ~15% opacity

        selected = bool(option.state & QStyle.State_Selected)
        hovered  = bool(option.state & QStyle.State_MouseOver)

        # Card rect (with outer padding)
        full = option.rect
        card = QRectF(
            full.left()   + self.PAD_H,
            full.top()    + self.PAD_V,
            full.width()  - self.PAD_H * 2,
            full.height() - self.PAD_V * 2,
        )

        # ── Card background ──
        if selected:
            grad = QLinearGradient(card.left(), 0, card.right(), 0)
            grad.setColorAt(0.0, cat_bg)
            grad.setColorAt(0.4, QColor(_C["card"]))
            painter.setBrush(QBrush(grad))
            border_c = QColor(cat["r"], cat["g"], cat["b"], 90)
        elif hovered:
            painter.setBrush(QBrush(QColor(_C["cardH"])))
            border_c = QColor(_C["borderH"])
        else:
            painter.setBrush(QBrush(QColor(_C["card"])))
            border_c = QColor(_C["border"])

        painter.setPen(QPen(border_c, 1))
        painter.drawRoundedRect(card, 10, 10)

        # ── Left color strip (clipped to card shape) ──
        clip = QPainterPath()
        clip.addRoundedRect(card, 10, 10)
        painter.setClipPath(clip)
        painter.fillRect(
            QRectF(card.left(), card.top(), self.STRIP_W, card.height()),
            cat_qc
        )
        painter.setClipping(False)

        # ── Text ──
        tx = int(card.left()) + self.STRIP_W + self.PAD_IN
        tw = int(card.width()) - self.STRIP_W - self.PAD_IN * 2 - 22
        ct = int(card.top())
        ch = int(card.height())

        # ID — monospace, bold, category text color
        id_font = QFont("Consolas", 10, QFont.Bold)
        painter.setFont(id_font)
        painter.setPen(QColor(cat["text"]))
        painter.drawText(QRect(tx, ct + 9, tw, 22), Qt.AlignLeft | Qt.AlignVCenter, pk_val)

        # Subtitle — regular, muted
        sub_font = QFont("Segoe UI", 9)
        painter.setFont(sub_font)
        painter.setPen(QColor(_C["txt2"]))
        fm  = painter.fontMetrics()
        sub = fm.elidedText(str(subtitle), Qt.ElideRight, tw)
        painter.drawText(QRect(tx, ct + 33, tw, 20), Qt.AlignLeft | Qt.AlignVCenter, sub)

        # Right arrow
        arr_font = QFont("Segoe UI", 13)
        painter.setFont(arr_font)
        painter.setPen(QColor(cat["color"]) if selected else QColor(_C["txt3"]))
        painter.drawText(
            QRect(int(card.right()) - 22, ct, 20, ch),
            Qt.AlignCenter, "›"
        )

        # Validation dot (top-right): red = error, yellow = warn
        vio = index.data(self.R_VIO)
        if vio:
            dot_c = QColor(_C["red"] if vio == "error" else _C["yellow"])
            painter.setPen(Qt.NoPen)
            painter.setBrush(QBrush(dot_c))
            painter.drawEllipse(QRectF(card.right() - 16, card.top() + 8, 7, 7))

        painter.restore()


# ── SubTableModel ─────────────────────────────────────────────────────────────

_DIRTY_BG = QColor(234, 179, 8, 30)
_ROW_EVEN = QColor(0x17, 0x17, 0x2A)
_ROW_ODD  = QColor(0x1E, 0x1E, 0x32)


class SubTableModel(QAbstractTableModel):
    def __init__(self, df, cols_cfg, manager, sheet_full_name):
        super().__init__()
        self._df           = df if df is not None else pd.DataFrame()
        self._cols_cfg     = cols_cfg or {}
        self._manager      = manager
        self._sheet        = sheet_full_name
        self._validation_cb = None   # editor hook: master-record visuals refresh

    def rowCount(self, parent=QModelIndex()):
        return 0 if parent.isValid() else len(self._df)

    def columnCount(self, parent=QModelIndex()):
        return 0 if parent.isValid() else len(self._df.columns)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        r, c    = index.row(), index.column()
        col     = self._df.columns[c]
        col_type = self._cols_cfg.get(col, {}).get("type", "string")
        val     = self._df.iat[r, c]
        row_idx = self._df.index[r]

        if role in (Qt.DisplayRole, Qt.EditRole):
            return None if col_type == "bool" else (str(val) if val is not None else "")
        irrelevant = not self._manager.binding.is_relevant(self._sheet, row_idx, col)
        if role == Qt.CheckStateRole and col_type == "bool":
            if irrelevant:
                return None               # 不相關的 bool 格不畫 checkbox
            v = val
            if isinstance(v, str):
                v = v.lower() in ("true", "1", "yes")
            return Qt.Checked if v else Qt.Unchecked
        if role == Qt.BackgroundRole:
            if irrelevant:
                return QBrush(QColor(_C["code"]))   # 更暗＝與此列無關
            vcol = self._manager.validator.cell_color(self._sheet, row_idx, col)
            if vcol:
                qc = QColor(vcol)
                qc.setAlpha(120)          # tint keeps text readable on dark rows
                return QBrush(qc)
            if (self._sheet, row_idx, col) in self._manager.dirty_cells:
                return QBrush(_DIRTY_BG)
            return QBrush(_ROW_EVEN if r % 2 == 0 else _ROW_ODD)
        if role == Qt.ToolTipRole:
            if irrelevant:
                drv = self._manager.binding.driver_of(*self._scope_pair())
                dval = ""
                if drv and drv in self._df.columns:
                    dval = str(self._df.at[row_idx, drv])
                return f"與 {dval or '此列'} 無關的欄位（保留舊值，不可編輯）"
            vrules = self._manager.validator.cell_rules(self._sheet, row_idx, col)
            if vrules:
                return "\n".join(
                    f"⚠ [{'錯誤' if ru['severity'] == 'error' else '警告'}] {ru['name']}"
                    for ru in vrules)
            return None
        if role == Qt.ForegroundRole:
            return QBrush(QColor(_C["txt3"] if irrelevant else _C["txt"]))
        return None

    def _scope_pair(self):
        if "." in self._sheet:
            m, s = self._sheet.split(".", 1)
            return m, s
        return self._sheet, ""

    def setData(self, index, value, role=Qt.EditRole):
        if not index.isValid():
            return False
        r, c    = index.row(), index.column()
        col     = self._df.columns[c]
        row_idx = self._df.index[r]
        if role == Qt.CheckStateRole:
            value = (value == Qt.Checked)
        self._manager.update_cell(self._sheet, row_idx, col, value)
        # Reflect the change in this filtered view copy at the SAME index label.
        # Do NOT swap in the full sub_table — that desyncs view rows from df rows
        # and routes later edits to whichever skill sits at low global indices.
        full = self._manager.sub_tables.get(self._sheet)
        if full is not None and row_idx in full.index and col in full.columns \
                and row_idx in self._df.index:
            self._df.at[row_idx, col] = full.at[row_idx, col]
        # validation can (un)mark OTHER cells of this row → repaint the whole row
        self.dataChanged.emit(self.index(r, 0),
                              self.index(r, self.columnCount() - 1))
        if self._validation_cb is not None:
            self._validation_cb()
        return True

    def flags(self, index):
        if not index.isValid():
            return Qt.NoItemFlags
        col      = self._df.columns[index.column()]
        col_type = self._cols_cfg.get(col, {}).get("type", "string")
        base     = Qt.ItemIsEnabled | Qt.ItemIsSelectable
        row_idx  = self._df.index[index.row()]
        if not self._manager.binding.is_relevant(self._sheet, row_idx, col):
            return base   # 欄位綁定判定與此列無關 → 鎖定不可編輯
        return base | (Qt.ItemIsUserCheckable if col_type == "bool" else Qt.ItemIsEditable)

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if orientation == Qt.Horizontal:
            if section < 0 or section >= len(self._df.columns):
                return None
            col = str(self._df.columns[section])
            if role == Qt.DisplayRole:
                return col
            if role == Qt.ToolTipRole:
                return self._cols_cfg.get(col, {}).get("note", "") or None
            return None
        if role == Qt.DisplayRole:
            return str(section + 1)
        return None

    def reload(self, df, cols_cfg=None):
        self.beginResetModel()
        self._df = df if df is not None else pd.DataFrame()
        if cols_cfg is not None:
            self._cols_cfg = cols_cfg
        self.endResetModel()

    def sort(self, column, order=Qt.AscendingOrder):
        if column < 0 or column >= len(self._df.columns):
            return
        col_name = self._df.columns[column]
        self.layoutAboutToBeChanged.emit()
        try:
            self._df = self._df.sort_values(
                col_name, ascending=(order == Qt.AscendingOrder), kind="mergesort"
            )
        except Exception:
            pass
        self.layoutChanged.emit()

    def df_index(self, view_row):
        if 0 <= view_row < len(self._df):
            return self._df.index[view_row]
        return None

    @property
    def df(self):
        return self._df


# ── EnumDelegate ──────────────────────────────────────────────────────────────

class EnumDelegate(QStyledItemDelegate):
    def __init__(self, options, parent=None):
        super().__init__(parent)
        self._options = [str(o) for o in options if str(o) != ""]

    def createEditor(self, parent, option, index):
        cb = QComboBox(parent)
        cb.addItem("")              # 允許清空 / 不選 enum 值
        cb.addItems(self._options)
        cb.view().setStyleSheet(_POPUP_QSS)   # Windows popup 不吃全域 QSS
        return cb

    def setEditorData(self, editor, index):
        val = index.data(Qt.DisplayRole) or ""
        idx = editor.findText(val)
        if idx >= 0:
            editor.setCurrentIndex(idx)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText(), Qt.EditRole)

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)


# ── Suggest field: context-filtered autocomplete ──────────────────────────────

def _get_suggestions(df, this_col, context_col, context_value):
    """Sorted, deduped previously-used values of `this_col` from `df`,
       filtered by `df[context_col] == context_value` when context_col is set."""
    if df is None or this_col not in df.columns:
        return []
    if context_col and context_col in df.columns:
        mask = df[context_col].astype(str) == str(context_value)
        vals = df.loc[mask, this_col]
    else:
        vals = df[this_col]
    seen, out = set(), []
    for v in vals.astype(str):
        v = v.strip()
        if v and v not in seen:
            seen.add(v); out.append(v)
    return sorted(out)


def _get_array_suggestions(df, this_col, context_col, context_value, limit=60):
    """Array 欄位的建議「元素」：把既有資料的逗號字串拆成單一 token 去重，
    有 context_col（建議來源）時只看同 context 值的列。"""
    seen, out = set(), []
    for cell in _get_suggestions(df, this_col, context_col, context_value):
        for tok in str(cell).split(","):
            tok = tok.strip()
            if tok and tok not in seen:
                seen.add(tok)
                out.append(tok)
    return sorted(out)[:limit]


def _suggest_chip_btn(tok, on_click):
    """建議值的可點擊小按鈕（陣列彈窗與母表欄位編輯器共用）。"""
    b = QPushButton(f"＋ {tok}")
    b.setAutoDefault(False); b.setDefault(False)
    b.setCursor(Qt.PointingHandCursor)
    b.setStyleSheet(
        f"QPushButton {{ background:{_C['card']}; color:{_C['txtAcc']}; "
        f"border:1px solid {_C['border']}; border-radius:9px; "
        f"padding:2px 8px; font-size:12px; }}"
        f"QPushButton:hover {{ background:{_C['cardH']}; "
        f"border-color:{_C['borderH']}; }}")
    b.clicked.connect(lambda _=False, t=tok: on_click(t))
    return b


class _SuggestLineEdit(QLineEdit):
    """QLineEdit that auto-pops its completer once on first focus-in.
    One-shot so that after the user picks a value the popup won't keep
    re-opening on the focus bounce (previously it 'wouldn't close')."""
    _did_autopop = False

    def focusInEvent(self, e):
        super().focusInEvent(e)
        c = self.completer()
        if c and not self._did_autopop:
            self._did_autopop = True
            QTimer.singleShot(0, c.complete)


class SuggestDelegate(QStyledItemDelegate):
    """Sub-table cell editor: line edit + popup of previously-used values,
       filtered by a sibling 'context' column on the same row."""
    def __init__(self, df_provider, this_col, context_col, parent=None):
        super().__init__(parent)
        self._df_provider  = df_provider
        self._this_col     = this_col
        self._context_col  = context_col

    def createEditor(self, parent, option, index):
        df = self._df_provider()
        ctx_val = ""
        try:
            ctx_val = str(index.model()._df.iloc[index.row()][self._context_col])
        except Exception:
            ctx_val = ""
        items = _get_suggestions(df, self._this_col, self._context_col, ctx_val)
        editor = _SuggestLineEdit(parent)
        m = QStringListModel(items, editor)
        compl = QCompleter(m, editor)
        compl.setCompletionMode(QCompleter.PopupCompletion)
        compl.setCaseSensitivity(Qt.CaseInsensitive)
        editor.setCompleter(compl)
        return editor

    def setEditorData(self, editor, index):
        editor.setText(index.data(Qt.DisplayRole) or "")

    def setModelData(self, editor, model, index):
        model.setData(index, editor.text(), Qt.EditRole)


# ── Array editor: chips widget ────────────────────────────────────────────────

class FlowLayout(QLayout):
    """Left-to-right wrapping layout used to lay out array chips."""

    def __init__(self, parent=None, spacing=4):
        super().__init__(parent)
        self._items = []
        self.setSpacing(spacing)
        self.setContentsMargins(3, 3, 3, 3)

    def addItem(self, item):       self._items.append(item)
    def count(self):               return len(self._items)
    def itemAt(self, i):           return self._items[i] if 0 <= i < len(self._items) else None
    def takeAt(self, i):           return self._items.pop(i) if 0 <= i < len(self._items) else None
    def expandingDirections(self): return Qt.Orientation(0)
    def hasHeightForWidth(self):   return True
    def heightForWidth(self, w):   return self._arrange(QRect(0, 0, w, 0), apply=False)
    def sizeHint(self):            return self.minimumSize()

    def setGeometry(self, rect):
        super().setGeometry(rect)
        self._arrange(rect, apply=True)

    def minimumSize(self):
        s = QSize()
        for it in self._items:
            s = s.expandedTo(it.minimumSize())
        m = self.contentsMargins()
        return s + QSize(m.left() + m.right(), m.top() + m.bottom())

    def _arrange(self, rect, apply):
        m = self.contentsMargins()
        x0 = rect.x() + m.left()
        x, y = x0, rect.y() + m.top()
        right = rect.right() - m.right()
        line_h = 0
        sp = self.spacing()
        for it in self._items:
            sz = it.sizeHint()
            if x > x0 and x + sz.width() > right:
                x = x0
                y += line_h + sp
                line_h = 0
            if apply:
                it.setGeometry(QRect(QPoint(x, y), sz))
            x += sz.width() + sp
            line_h = max(line_h, sz.height())
        return y + line_h + m.bottom() - rect.y()


class _Chip(QFrame):
    """One array element — double-click to edit, ✕ to remove."""
    removed = Signal(object)
    edited  = Signal()

    def __init__(self, text, parent=None):
        super().__init__(parent)
        self._text = text
        self._editing = False
        self.setObjectName("arrayChip")
        self.setStyleSheet(
            f"QFrame#arrayChip {{ background:{_C['card']}; "
            f"border:1px solid {_C['border']}; border-radius:9px; }}"
        )
        self.setToolTip("雙擊可編輯")
        lo = QHBoxLayout(self)
        lo.setContentsMargins(8, 2, 4, 2)
        lo.setSpacing(3)

        self._lbl = QLabel(text if text != "" else "(空)")
        self._lbl.setStyleSheet(
            f"color:{_C['txt']}; background:transparent; border:none; font-size:12px;"
        )
        self._edit = QLineEdit(text)
        self._edit.setStyleSheet(
            f"color:{_C['txt']}; background:transparent; border:none; font-size:12px;"
        )
        self._edit.hide()
        self._edit.editingFinished.connect(self._commit_edit)
        self._edit.textChanged.connect(self._resize_edit)

        btn = QPushButton("✕")
        btn.setFixedSize(18, 18)
        btn.setAutoDefault(False); btn.setDefault(False)
        btn.setCursor(Qt.PointingHandCursor)
        btn.setStyleSheet(
            f"QPushButton {{ background:transparent; border:none; padding:0px; "
            f"margin:0px; color:{_C['txt3']}; font-size:13px; }}"
            f"QPushButton:hover {{ color:{_C['red']}; }}"
        )
        btn.clicked.connect(lambda: self.removed.emit(self))

        lo.addWidget(self._lbl)
        lo.addWidget(self._edit)
        lo.addWidget(btn)

    def mouseDoubleClickEvent(self, event):
        self._start_edit()

    def _start_edit(self):
        self._editing = True
        self._edit.setText(self._text)
        self._resize_edit()
        self._lbl.hide()
        self._edit.show()
        self._edit.setFocus()
        self._edit.selectAll()
        self.updateGeometry()

    def _resize_edit(self):
        fm = self._edit.fontMetrics()
        self._edit.setFixedWidth(max(fm.horizontalAdvance(self._edit.text()) + 8, 24))

    def _commit_edit(self):
        if not self._editing:
            return
        self._editing = False
        new = self._edit.text().strip()
        self._edit.hide()
        if new == "":
            self.removed.emit(self)
            return
        changed = (new != self._text)
        self._text = new
        self._lbl.setText(new)
        self._lbl.show()
        self.updateGeometry()
        if changed:
            self.edited.emit()

    def text(self):
        return self._text


class ChipsEdit(QWidget):
    """Array field editor: wrapping removable chips + add-input + copy button."""
    changed = Signal()

    def __init__(self, parent=None, chip_area_height=64):
        super().__init__(parent)
        self._chips = []
        lo = QVBoxLayout(self)
        lo.setContentsMargins(0, 0, 0, 0)
        lo.setSpacing(4)

        self._chip_box = QWidget()
        sp = self._chip_box.sizePolicy()
        sp.setHeightForWidth(True)
        self._chip_box.setSizePolicy(sp)
        self._chip_box.setStyleSheet(
            f"background:{_C['code']}; border:1px solid {_C['border']}; border-radius:6px;"
        )
        self._flow = FlowLayout(self._chip_box, spacing=4)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedHeight(chip_area_height)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet("QScrollArea{border:none; background:transparent;}")
        scroll.setWidget(self._chip_box)
        lo.addWidget(scroll)

        row = QHBoxLayout()
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(4)
        self._input = QLineEdit()
        self._input.setPlaceholderText("輸入後按 Enter 或逗號新增…")
        self._input.returnPressed.connect(self._commit_input)
        self._input.textChanged.connect(self._on_input_text)
        copy_btn = QPushButton("⧉ 複製")
        copy_btn.setAutoDefault(False); copy_btn.setDefault(False)
        copy_btn.setCursor(Qt.PointingHandCursor)
        copy_btn.setToolTip("複製整個陣列內容（以逗號分隔）")
        copy_btn.setStyleSheet(
            f"background:{_C['card']}; border:1px solid {_C['border']}; "
            f"color:{_C['txt2']}; border-radius:5px; padding:3px 8px; font-size:11px;"
        )
        copy_btn.clicked.connect(self._copy_all)
        row.addWidget(self._input, 1)
        row.addWidget(copy_btn)
        lo.addLayout(row)

    # ── public API ──
    def set_value(self, comma_str):
        """Populate chips from a comma-separated string (does not emit `changed`)."""
        for ch in self._chips:
            self._flow.removeWidget(ch)
            ch.setParent(None)
            ch.deleteLater()
        self._chips = []
        for tok in str(comma_str or "").split(","):
            tok = tok.strip()
            if tok != "":
                self._add_chip(tok)
        self._chip_box.updateGeometry()

    def value(self):
        return ", ".join(c.text() for c in self._chips)

    # ── internal ──
    def _add_chip(self, text):
        chip = _Chip(text, self._chip_box)
        chip.removed.connect(self._remove_chip)
        chip.edited.connect(self.changed.emit)
        self._flow.addWidget(chip)
        self._chips.append(chip)
        chip.show()

    def _remove_chip(self, chip):
        if chip in self._chips:
            self._chips.remove(chip)
        self._flow.removeWidget(chip)
        chip.setParent(None)
        chip.deleteLater()
        self._chip_box.updateGeometry()
        self.changed.emit()

    def _commit_input(self):
        raw = self._input.text()
        self._input.blockSignals(True)
        self._input.clear()
        self._input.blockSignals(False)
        added = False
        for tok in raw.split(","):
            tok = tok.strip()
            if tok != "":
                self._add_chip(tok)
                added = True
        if added:
            self._chip_box.updateGeometry()
            self.changed.emit()

    def _on_input_text(self, text):
        if "," in text:
            self._commit_input()

    def _copy_all(self):
        QApplication.clipboard().setText(self.value())

    def add_token(self, text):
        """外部快速加入一個元素（建議值點擊用）。"""
        text = str(text).strip()
        if text == "":
            return
        self._add_chip(text)
        self._chip_box.updateGeometry()
        self.changed.emit()


class ArrayEditDialog(QDialog):
    """Popup chips editor for array cells in sub-tables.
    有建議來源（suggest_from）時多一塊建議值區，點一下即加入。"""

    def __init__(self, comma_str, parent=None, suggestions=None, source_note=""):
        super().__init__(parent)
        self.setWindowTitle("編輯陣列")
        self.setMinimumWidth(500 if suggestions else 380)
        self.setStyleSheet(APP_QSS)
        lo = QVBoxLayout(self)
        self._chips = ChipsEdit(chip_area_height=150)
        self._chips.set_value(comma_str)
        lo.addWidget(self._chips)

        if suggestions:
            cap = QLabel("建議值（點一下加入）" + (f"　{source_note}" if source_note else ""))
            cap.setStyleSheet(
                f"color:{_C['txt3']}; font-size:11px; background:transparent; "
                f"padding-top:4px;")
            lo.addWidget(cap)
            sug_box = QWidget()
            sug_box.setStyleSheet(
                f"background:{_C['panel']}; border:1px solid {_C['border']}; "
                f"border-radius:6px;")
            sflow = FlowLayout(sug_box, spacing=4)
            for tok in suggestions:
                sflow.addWidget(_suggest_chip_btn(tok, self._chips.add_token))
            sscroll = QScrollArea()
            sscroll.setWidgetResizable(True)
            sscroll.setMaximumHeight(120)
            sscroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            sscroll.setStyleSheet("QScrollArea{border:none; background:transparent;}")
            sscroll.setWidget(sug_box)
            lo.addWidget(sscroll)

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)
        for _b in bb.buttons():
            _b.setAutoDefault(False); _b.setDefault(False)
        lo.addWidget(bb)

    def keyPressEvent(self, event):
        # Enter is for committing/adding an item (handled by the child QLineEdits'
        # returnPressed/editingFinished, which fire first). Swallow it here so it
        # doesn't trigger a default button (close dialog / delete chip). OK closes.
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            event.accept()
            return
        super().keyPressEvent(event)

    def value(self):
        return self._chips.value()


class ArrayDelegate(QStyledItemDelegate):
    """Sub-table array cell — double-click opens the chips editor dialog.
    欄位設有建議來源（suggest_from）時，彈窗附建議值快速選填區。"""

    def __init__(self, parent=None, df_provider=None, this_col="", context_col=""):
        super().__init__(parent)
        self._df_provider = df_provider
        self._this_col    = this_col
        self._context_col = context_col

    def createEditor(self, parent, option, index):
        cur = index.data(Qt.DisplayRole) or ""
        suggestions, note = [], ""
        if self._df_provider is not None and self._this_col and self._context_col:
            try:
                ctx_val = str(index.model()._df.iloc[index.row()][self._context_col])
            except Exception:
                ctx_val = ""
            suggestions = _get_array_suggestions(
                self._df_provider(), self._this_col, self._context_col, ctx_val)
            note = f"來源：{self._context_col}＝{ctx_val} 的既有資料"
        dlg = ArrayEditDialog(cur, parent, suggestions=suggestions, source_note=note)
        if dlg.exec() == QDialog.Accepted:
            index.model().setData(index, dlg.value(), Qt.EditRole)
        return None


# ── FieldEditorWidget ─────────────────────────────────────────────────────────

class FieldEditorWidget(QWidget):
    """Build-once field form with indigo ● dots and dirty-state highlighting."""
    field_changed = Signal(str, object)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._widgets       = {}
        self._col_types     = {}
        self._lbl_widgets   = {}   # col → QLabel (the col name label)
        self._note_tips     = {}   # col → configured note (tooltip restore)
        self._grp_widgets   = {}   # col → (group widget, separator) — 綁定顯示/隱藏用
        self._bool_updaters = {}   # col → update_style(checked: bool)
        self._img_preview_label:  "QLabel | None" = None  # table-level image preview
        self._img_path_segments:  list = []  # [{"type":"col","col":"X"} | {"type":"lit","value":"Y"}]
        self._img_base_folder:    str  = ""  # configured base folder for images
        self._img_ext:            str  = ""  # file extension appended to assembled path e.g. ".png"
        self._text_ref_cfg  = {}   # col → {"json_path","key_col","val_col"} (per-column text_ref)
        self._ref_labels    = {}   # col → QLabel (resolved lookup text for "text_ref" type)
        self._array_sug     = {}   # array col → {box, flow, cap, ctx, chips}（建議值區）
        self._array_sug_ctx = {}   # array col → 上次渲染的 context 值（避免重複重建）
        self._row_idx     = None
        self._table_name  = None
        self._manager     = None
        self._built       = False

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet("background: transparent; border: none;")
        self._scroll = scroll

        self._content = QWidget()
        self._content.setStyleSheet(f"background: {_C['panel']};")
        self._form_lo = QVBoxLayout(self._content)
        self._form_lo.setContentsMargins(0, 4, 0, 12)
        self._form_lo.setSpacing(0)
        scroll.setWidget(self._content)

        lo = QVBoxLayout(self)
        lo.setContentsMargins(0, 0, 0, 0)
        lo.addWidget(scroll)

    def build_for(self, df, cfg, table_name, manager):
        # Clear
        while self._form_lo.count():
            item = self._form_lo.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._widgets.clear()
        self._col_types.clear()
        self._lbl_widgets.clear()
        self._note_tips.clear()
        self._grp_widgets.clear()
        self._bool_updaters.clear()
        self._img_preview_label  = None  # cleared by layout wipe above
        self._img_path_segments  = []
        self._img_base_folder    = ""
        self._img_ext            = ""
        self._text_ref_cfg = {}
        self._ref_labels.clear()
        self._array_sug.clear()
        self._array_sug_ctx.clear()
        self._table_name = table_name
        self._manager    = manager
        cols_cfg  = cfg.get("columns", {})
        _img_cfg  = cfg.get("image_preview", {})
        self._img_base_folder  = _img_cfg.get("base_folder", "")
        self._img_ext          = _img_cfg.get("ext", "")
        _segs = _img_cfg.get("path_segments", [])
        if not _segs:
            # backward compat: old single "col" key
            _old_col = _img_cfg.get("col", "")
            if _old_col and _old_col in df.columns:
                _segs = [{"type": "col", "col": _old_col}]
        self._img_path_segments = _segs
        self._text_ref_cfg = {
            c: cc.get("text_ref", {})
            for c, cc in cols_cfg.items()
            if isinstance(cc, dict) and cc.get("type") == "text_ref"
        }

        # ── Table-level image preview (shown at top if configured) ─────────────
        if self._img_path_segments:
            _seg_desc = "/".join(
                s.get("col", "?") if s.get("type") == "col" else s.get("value", "?")
                for s in self._img_path_segments
            )
            prev_card = QWidget()
            prev_card.setStyleSheet(f"background:{_C['panel']};")
            pclo = QVBoxLayout(prev_card)
            pclo.setContentsMargins(14, 10, 14, 6); pclo.setSpacing(4)
            lbl_img = QLabel(f"● IMAGE  [{_seg_desc}]")
            lbl_img.setStyleSheet(
                f"color:{_C['txt3']}; font-size:10px; font-weight:600; "
                f"letter-spacing:1px; background:transparent;"
            )
            self._img_preview_label = QLabel("No Image")
            self._img_preview_label.setAlignment(Qt.AlignCenter)
            self._img_preview_label.setFixedHeight(160)
            self._img_preview_label.setStyleSheet(
                f"background:{_C['code']}; border:1px solid {_C['border']}; "
                f"border-radius:6px; color:{_C['txt3']}; font-size:12px;"
            )
            pclo.addWidget(lbl_img)
            pclo.addWidget(self._img_preview_label)
            sep_img = QFrame(); sep_img.setFixedHeight(1)
            sep_img.setStyleSheet(f"background:{_C['border']}; border:none;")
            self._form_lo.addWidget(prev_card)
            self._form_lo.addWidget(sep_img)

        for col in df.columns:
            col_conf = cols_cfg.get(col, {})
            col_type = col_conf.get("type", "string")
            self._col_types[col] = col_type

            # ── Field group ──
            grp = QWidget()
            grp.setStyleSheet(f"background: {_C['panel']};")
            glo = QVBoxLayout(grp)
            glo.setContentsMargins(14, 8, 14, 4)
            glo.setSpacing(4)

            # Label row: ● ColName
            lbl_row = QWidget()
            lbl_row.setStyleSheet("background: transparent;")
            lrlo = QHBoxLayout(lbl_row)
            lrlo.setContentsMargins(0, 0, 0, 0)
            lrlo.setSpacing(5)

            dot = QLabel("●")
            dot.setStyleSheet(
                f"color: {_C['accent']}; font-size: 11px; background: transparent;"
            )
            lbl = QLabel(col)
            lbl.setStyleSheet(
                f"color: {_C['txt2']}; font-size: 11px; font-weight: 500; background: transparent;"
            )
            self._lbl_widgets[col] = lbl
            _note = col_conf.get("note", "")
            self._note_tips[col] = _note
            if _note:
                lbl.setToolTip(_note)
                lbl_row.setToolTip(_note)
            lrlo.addWidget(dot)
            lrlo.addWidget(lbl)
            lrlo.addStretch()
            glo.addWidget(lbl_row)

            # ── Input widget ──
            if col_type == "bool":
                w = QPushButton()
                w.setCheckable(True)
                w.setFixedHeight(36)

                def _make_bool_style(btn):
                    def _upd(checked):
                        if checked:
                            btn.setText("  ✓   True")
                            btn.setStyleSheet(
                                f"background:rgba(16,185,129,0.18);"
                                f"border:1px solid rgba(16,185,129,0.65);"
                                f"color:{_C['green']}; border-radius:7px;"
                                f"font-size:13px; font-weight:600; text-align:left; padding:0 12px;"
                            )
                        else:
                            btn.setText("  ✗   False")
                            btn.setStyleSheet(
                                f"background:rgba(239,68,68,0.10);"
                                f"border:1px solid rgba(239,68,68,0.35);"
                                f"color:{_C['red']}; border-radius:7px;"
                                f"font-size:13px; font-weight:600; text-align:left; padding:0 12px;"
                            )
                    return _upd

                updater = _make_bool_style(w)
                self._bool_updaters[col] = updater
                updater(False)
                w.toggled.connect(updater)
                w.toggled.connect(lambda checked, c=col: self.field_changed.emit(c, checked))

            elif col_type == "enum":
                opts = [str(o) for o in (col_conf.get("options") or []) if str(o) != ""]
                w = _NoscrollCombo()
                w.addItem("")
                w.addItems(opts)
                w.currentTextChanged.connect(
                    lambda v, c=col: self.field_changed.emit(c, v)
                )

            elif col_type in ("int", "float"):
                w = QLineEdit()
                w.textChanged.connect(
                    lambda v, c=col, ct=col_type: self._on_numeric(c, v, ct)
                )

            elif col_type == "array":
                w = ChipsEdit()
                w.changed.connect(
                    lambda c=col, _w=w: self.field_changed.emit(c, _w.value())
                )
                sug_ctx = col_conf.get("suggest_from", "")
                if sug_ctx:
                    # 建議值快速選填區（依當前列的 context 值於 load_row 時填入）
                    cap = QLabel("")
                    cap.setStyleSheet(
                        f"color:{_C['txt3']}; font-size:11px; background:transparent;")
                    sug_box = QWidget()
                    sug_box.setStyleSheet(
                        f"background:{_C['code']}; border:1px solid {_C['border']}; "
                        f"border-radius:6px;")
                    sflow = FlowLayout(sug_box, spacing=4)
                    sscroll = QScrollArea()
                    sscroll.setWidgetResizable(True)
                    sscroll.setMaximumHeight(88)
                    sscroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                    sscroll.setStyleSheet(
                        "QScrollArea{border:none; background:transparent;}")
                    sscroll.setWidget(sug_box)
                    glo.addWidget(w)
                    glo.addWidget(cap)
                    glo.addWidget(sscroll)
                    self._widgets[col] = w
                    self._array_sug[col] = {"box": sug_box, "flow": sflow,
                                            "cap": cap, "scroll": sscroll,
                                            "ctx": sug_ctx, "chips": w}
                    sep = QFrame(); sep.setFixedHeight(1)
                    sep.setStyleSheet(f"background: {_C['border']}; border: none;")
                    self._form_lo.addWidget(grp)
                    self._form_lo.addWidget(sep)
                    self._grp_widgets[col] = (grp, sep)
                    continue

            elif col_type == "text_ref":
                # Editable string (same as string type) — value IS the lookup key
                w = QTextEdit()
                w.setMaximumHeight(76)
                w.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

                # Read-only resolved text label (live-updating)
                ref_lbl = QLabel("—")
                ref_lbl.setWordWrap(True)
                ref_lbl.setStyleSheet(
                    f"color:#FFFFFF; font-size:13px; font-weight:700; font-style:normal; "
                    f"background:{_C['code']}; border-radius:4px; padding:4px 8px;"
                )
                self._ref_labels[col] = ref_lbl

                def _on_text_ref(c=col, widget=w, lbl=ref_lbl):
                    val = widget.toPlainText()
                    self.field_changed.emit(c, val)
                    self._update_ref_label(c, lbl, val)

                w.textChanged.connect(_on_text_ref)
                glo.addWidget(w)
                glo.addWidget(ref_lbl)
                self._widgets[col] = w
                sep = QFrame(); sep.setFixedHeight(1)
                sep.setStyleSheet(f"background: {_C['border']}; border: none;")
                self._form_lo.addWidget(grp)   # must parent grp or GC kills the child widgets
                self._form_lo.addWidget(sep)
                self._grp_widgets[col] = (grp, sep)
                continue

            else:
                w = QTextEdit()
                w.setMaximumHeight(76)
                w.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                w.textChanged.connect(
                    lambda c=col, widget=w: self.field_changed.emit(c, widget.toPlainText())
                )

            glo.addWidget(w)
            self._widgets[col] = w

            # Thin separator line below each field
            sep = QFrame()
            sep.setFixedHeight(1)
            sep.setStyleSheet(f"background: {_C['border']}; border: none;")
            self._form_lo.addWidget(grp)
            self._form_lo.addWidget(sep)
            self._grp_widgets[col] = (grp, sep)

        self._form_lo.addStretch(1)
        self._built = True

    # ── Quick-nav focus (from search results / compare view) ────────────────────
    def focus_field(self, col):
        """Scroll the given field into view and briefly flash its label so the
        user can see where the search/compare jump landed. No-op if unknown."""
        w   = self._widgets.get(col)
        lbl = self._lbl_widgets.get(col)
        if w is None:
            return
        self._scroll.ensureWidgetVisible(w, 0, 40)
        try:
            w.setFocus()
        except Exception:
            pass
        if lbl is not None:
            base = (f"color: {_C['txt2']}; font-size: 11px; font-weight: 500; "
                    f"background: transparent;")
            flash = (f"color: {_C['accent']}; font-size: 11px; font-weight: 700; "
                     f"background: rgba(99,102,241,0.20); border-radius: 3px;")
            lbl.setStyleSheet(flash)
            QTimer.singleShot(900, lambda l=lbl, b=base: l.setStyleSheet(b))

    def refresh_ref_labels(self):
        """Re-resolve every text_ref display label from its current field value.
        Used when a referenced external text table was saved."""
        if not self._built:
            return
        for col, lbl in self._ref_labels.items():
            w = self._widgets.get(col)
            if w is None:
                continue
            try:
                val = w.toPlainText()
            except Exception:
                val = ""
            self._update_ref_label(col, lbl, val)

    # ── Validation ────────────────────────────────────────────────────────────

    def _on_numeric(self, col, value, col_type):
        w = self._widgets.get(col)
        if w:
            valid = self._validate(value, col_type)
            w.setProperty("invalid", "true" if not valid else "false")
            w.style().unpolish(w)
            w.style().polish(w)
        self.field_changed.emit(col, value)

    @staticmethod
    def _validate(value, col_type):
        if col_type == "int":
            return value == "" or value.lstrip("-").isdigit()
        if col_type == "float":
            if value in ("", "-", "."): return True
            try: float(value); return True
            except ValueError: return False
        return True

    # ── Text-ref lookup ───────────────────────────────────────────────────────

    def _update_ref_label(self, col: str, lbl: "QLabel", key_val: str) -> None:
        src = self._text_ref_cfg.get(col, {})
        jp  = (src.get("json_path") or "").strip()
        if not jp:
            lbl.setText("（此欄未設定外部文字表）")
            return
        if not self._manager:
            return
        json_dir = os.path.dirname(self._manager.json_path) if self._manager.json_path else ""
        abs_ref  = jp if os.path.isabs(jp) else (os.path.join(json_dir, jp) if json_dir else jp)
        resolved = self._manager.get_ref_text(
            abs_ref, src.get("key_col", "TextID") or "TextID",
            key_val, src.get("val_col", "TextContent") or "TextContent"
        )
        lbl.setText(resolved if resolved else "（找不到對應文字）")

    # ── Load row ──────────────────────────────────────────────────────────────

    def _style_field(self, col, w, is_dirty):
        """Apply violation (rule colour) > dirty (yellow) > normal styling to a
        field widget + its name label."""
        lbl = self._lbl_widgets.get(col)
        vrules = []
        if self._manager is not None and self._row_idx is not None:
            vrules = self._manager.validator.cell_rules(
                self._table_name, self._row_idx, col)
        if vrules:
            qc = QColor(vrules[0]["color"])
            rgb = f"{qc.red()},{qc.green()},{qc.blue()}"
            w.setStyleSheet(
                f"border-color: rgba({rgb},0.85); background: rgba({rgb},0.12);")
            tip = "\n".join(
                f"⚠ [{'錯誤' if ru['severity'] == 'error' else '警告'}] {ru['name']}"
                for ru in vrules)
            w.setToolTip(tip)
            if lbl:
                lbl.setStyleSheet(
                    f"color: {vrules[0]['color']}; font-size: 11px; "
                    f"font-weight: 700; background: transparent;")
                lbl.setToolTip(tip)
            return
        w.setToolTip("")
        if is_dirty:
            w.setStyleSheet(
                "border-color: rgba(234,179,8,0.55); background: rgba(234,179,8,0.07);"
            )
            if lbl:
                lbl.setStyleSheet(
                    f"color: {_C['yellow']}; font-size: 11px; font-weight: 500; background: transparent;"
                )
        else:
            w.setStyleSheet("")
            if lbl:
                lbl.setStyleSheet(
                    f"color: {_C['txt2']}; font-size: 11px; font-weight: 500; background: transparent;"
                )
        if lbl:
            lbl.setToolTip(self._note_tips.get(col, ""))

    def refresh_validation(self):
        """Re-style every field of the current row after validation changed
        (without reloading values — safe to call mid-edit)."""
        if not self._built or self._row_idx is None:
            return
        dirty = self._manager.dirty_cells if self._manager else set()
        for col, w in self._widgets.items():
            self._style_field(col, w,
                              (self._table_name, self._row_idx, col) in dirty)

    def refresh_binding(self):
        """依欄位綁定顯示/隱藏欄位區塊（母表表單）。值不動，只收合畫面。"""
        if not self._built or self._row_idx is None or self._manager is None:
            return
        rel = self._manager.binding.relevant_fields(
            self._table_name, "", self._row_idx)
        for col, (grp, sep) in self._grp_widgets.items():
            show = rel is None or col in rel
            grp.setVisible(show)
            sep.setVisible(show)

    def refresh_array_suggestions(self):
        """重建有建議來源的陣列欄位建議值區（context 值沒變就跳過）。"""
        if not self._built or self._row_idx is None or self._manager is None:
            return
        df = self._manager.tables.get(self._table_name)
        if df is None or self._row_idx not in df.index:
            return
        for col, sug in self._array_sug.items():
            ctx_col = sug["ctx"]
            ctx_val = str(df.at[self._row_idx, ctx_col]) \
                if ctx_col in df.columns else ""
            if self._array_sug_ctx.get(col) == ctx_val:
                continue
            self._array_sug_ctx[col] = ctx_val
            flow = sug["flow"]
            while flow.count():
                it = flow.takeAt(0)
                w = it.widget()
                if w is not None:
                    w.hide()
                    w.deleteLater()
            toks = _get_array_suggestions(df, col, ctx_col, ctx_val)
            for tok in toks:
                flow.addWidget(_suggest_chip_btn(tok, sug["chips"].add_token))
            sug["cap"].setText(
                f"建議值（點一下加入）　來源：{ctx_col}＝{ctx_val} 的既有資料"
                if toks else "")
            sug["scroll"].setVisible(bool(toks))
            sug["cap"].setVisible(bool(toks))
            sug["box"].updateGeometry()

    def load_row(self, row_data, row_idx):
        if not self._built:
            return
        self._row_idx = row_idx
        dirty = self._manager.dirty_cells if self._manager else set()

        for col, w in self._widgets.items():
            w.blockSignals(True)
            try:
                try:    val = row_data[col]
                except: val = ""

                col_type = self._col_types[col]
                is_dirty = (self._table_name, row_idx, col) in dirty

                self._style_field(col, w, is_dirty)

                # Value
                if col_type == "bool":
                    v = val
                    if isinstance(v, str):
                        v = v.lower() in ("true", "1", "yes")
                    checked = bool(v) if val != "" else False
                    w.setChecked(checked)
                    upd = self._bool_updaters.get(col)
                    if upd:
                        upd(checked)
                elif col_type == "enum":
                    w.setCurrentText(str(val) if val is not None else "")
                elif col_type in ("int", "float"):
                    w.setText(str(val) if val is not None else "")
                    w.setProperty("invalid", "false")
                    w.style().unpolish(w); w.style().polish(w)
                elif col_type == "array":
                    w.set_value(str(val) if val is not None else "")
                elif col_type == "text_ref":
                    val_str = str(val) if val is not None else ""
                    w.setPlainText(val_str)
                    ref_lbl = self._ref_labels.get(col)
                    if ref_lbl:
                        self._update_ref_label(col, ref_lbl, val_str)
                else:
                    w.setPlainText(str(val) if val is not None else "")

            finally:
                w.blockSignals(False)

        # Update table-level image preview
        if self._img_path_segments and self._img_preview_label:
            try:
                parts = []
                for _seg in self._img_path_segments:
                    if _seg.get("type") == "col":
                        _c = _seg.get("col", "")
                        parts.append(str(row_data[_c]) if _c and _c in row_data.index else "")
                    else:
                        parts.append(_seg.get("value", ""))
                img_val = "/".join(parts) + self._img_ext
            except Exception:
                img_val = ""
            # Resolve base: configured folder first, else JSON dir
            base = self._img_base_folder
            if base and not os.path.isabs(base) and self._manager and self._manager.json_path:
                base = os.path.join(os.path.dirname(self._manager.json_path), base)
            elif not base and self._manager and self._manager.json_path:
                base = os.path.dirname(self._manager.json_path)
            _update_img_thumb(img_val, self._img_preview_label, base)


# ── SubTablePanel ─────────────────────────────────────────────────────────────

class SubTablePanel(QWidget):
    row_deleted    = Signal(str, object)
    copy_requested = Signal()
    paste_requested = Signal()
    compare_requested = Signal(object)  # df_index of the right-clicked row

    def __init__(self, sheet_full_name, cols_cfg, manager, parent=None):
        super().__init__(parent)
        self._sheet   = sheet_full_name
        self._manager = manager

        _existing = manager.sub_tables.get(sheet_full_name)
        empty_df = pd.DataFrame(columns=list(
            _existing.columns if _existing is not None else []
        ))
        self._model = SubTableModel(empty_df, cols_cfg, manager, sheet_full_name)

        self._view = QTableView()
        self._view.setModel(self._model)
        self._view.setSortingEnabled(True)
        self._view.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._view.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self._view.setEditTriggers(
            QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
        self._view.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._view.horizontalHeader().setStretchLastSection(True)
        self._view.verticalHeader().setDefaultSectionSize(28)
        self._view.verticalHeader().hide()
        self._view.setContextMenuPolicy(Qt.CustomContextMenu)
        self._view.customContextMenuRequested.connect(self._ctx_menu)
        self._view.keyPressEvent = self._key_press
        self._refresh_delegates(cols_cfg)

        lo = QVBoxLayout(self)
        lo.setContentsMargins(0, 0, 0, 0)
        lo.addWidget(self._view)

    def _refresh_delegates(self, cols_cfg):
        for c, col in enumerate(self._model.df.columns):
            col_conf = (cols_cfg or {}).get(col, {})
            col_type = col_conf.get("type", "string")
            if col_type == "enum":
                opts = col_conf.get("options") or [""]
                self._view.setItemDelegateForColumn(c, EnumDelegate(opts, self._view))
            elif col_type == "array":
                df_getter = (lambda sheet=self._sheet:
                             self._manager.sub_tables.get(sheet))
                self._view.setItemDelegateForColumn(c, ArrayDelegate(
                    self._view, df_provider=df_getter, this_col=col,
                    context_col=col_conf.get("suggest_from", "")))
            elif col_conf.get("suggest_from"):
                ctx = col_conf["suggest_from"]
                df_getter = (lambda sheet=self._sheet:
                             self._manager.sub_tables.get(sheet))
                self._view.setItemDelegateForColumn(c, SuggestDelegate(
                    df_getter, col, ctx, self._view
                ))

    def set_validation_cb(self, cb):
        """Editor hook invoked after each cell edit (validation visuals)."""
        self._model._validation_cb = cb

    def flush_pending_edit(self):
        """Commit a still-open cell editor into the model before the view is
        reloaded or the file is saved. The card list is Qt.NoFocus, so clicking
        another item never focus-outs the editor; the model reset in reload()
        would then destroy the editor and silently discard the picked value
        (empty enum values are dropped from the saved JSON entirely)."""
        if self._view.state() != QAbstractItemView.EditingState:
            return
        editor = QApplication.focusWidget()
        # climb to the widget that is the delegate-registered editor
        # (a direct child of the view's viewport)
        vp = self._view.viewport()
        while editor is not None and editor.parentWidget() is not vp:
            editor = editor.parentWidget()
        if editor is None:
            return
        self._view.commitData(editor)
        self._view.closeEditor(editor, QAbstractItemDelegate.NoHint)

    def reload(self, df, cols_cfg=None):
        self.flush_pending_edit()
        self._model.reload(df, cols_cfg)
        if cols_cfg:
            self._refresh_delegates(cols_cfg)
        self.apply_column_binding()
        if not df.empty:
            self._view.resizeColumnsToContents()

    def apply_column_binding(self):
        """依欄位綁定隱藏「目前顯示的列都用不到」的欄位（聯集）。"""
        df = self._model.df
        if "." in self._sheet:
            master, scope = self._sheet.split(".", 1)
        else:
            master, scope = self._sheet, ""
        visible = self._manager.binding.visible_columns(
            master, scope, list(df.index))
        for c, col in enumerate(df.columns):
            self._view.setColumnHidden(c, visible is not None and col not in visible)

    def select_by_df_index(self, df_idx, col=None):
        """Select the row whose underlying DataFrame index == df_idx and scroll
        to `col` (best-effort). Used by global-search quick-nav into sub-tables."""
        model = self._model
        cols  = list(model.df.columns)
        target_col = cols.index(col) if (col and col in cols) else 0
        for r in range(model.rowCount()):
            try:
                if model.df_index(r) == df_idx:
                    idx = model.index(r, target_col)
                    self._view.setCurrentIndex(idx)
                    self._view.selectRow(r)
                    self._view.scrollTo(idx, QAbstractItemView.PositionAtCenter)
                    return True
            except Exception:
                pass
        return False

    def _ctx_menu(self, pos):
        menu = QMenu(self)
        menu.addAction("複製列", self.copy_requested.emit)
        menu.addAction("貼上列", self.paste_requested.emit)
        idx = self._view.indexAt(pos)
        if idx.isValid():
            di = self._model.df_index(idx.row())
            menu.addAction("對照相同效果…", lambda: self.compare_requested.emit(di))
            menu.addSeparator()
            menu.addAction("刪除此列", self._delete_selected)
        menu.exec(self._view.viewport().mapToGlobal(pos))

    def _key_press(self, event):
        if event.key() == Qt.Key_Delete:
            self._delete_selected()
        elif event.matches(QKeySequence.Copy):
            self.copy_requested.emit()
        elif event.matches(QKeySequence.Paste):
            self.paste_requested.emit()
        else:
            QTableView.keyPressEvent(self._view, event)

    def _delete_selected(self):
        sel = self._view.selectionModel().selectedRows()
        if not sel:
            return
        df_idx = self._model.df_index(sel[0].row())
        if df_idx is not None:
            self.row_deleted.emit(self._sheet, df_idx)

    def selected_df_index(self):
        sel = self._view.selectionModel().selectedRows()
        if not sel:
            return None
        return self._model.df_index(sel[0].row())

    def selected_df_indices(self):
        rows = sorted(i.row() for i in self._view.selectionModel().selectedRows())
        out = []
        for r in rows:
            di = self._model.df_index(r)
            if di is not None:
                out.append(di)
        return out


class EffectCompareDialog(QDialog):
    """Read-only cross-skill effect comparison: list every row in the sub-table
    that matches the seed row's key columns (InfluenceStatus / EffectReceive…)."""

    _FILTERS = [("InfluenceStatus", "影響屬性"), ("EffectReceive", "目標"),
                ("SkillComponentID", "組件")]
    _SHOW = ["SkillComponentID", "EffectValue", "InfluenceStatus",
             "EffectReceive", "AddType", "EffectDurationTime"]

    def __init__(self, sub_df, fk_key, seed, name_of, sub_name, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"對照相同效果 — {sub_name}")
        self.resize(700, 540)
        self.setStyleSheet(APP_QSS)
        self._df = sub_df
        self._fk = fk_key
        self._name_of = name_of
        self._cols = [c for c in self._SHOW if c in sub_df.columns]

        outer = QVBoxLayout(self)
        frow = QWidget(); frow.setStyleSheet("background:transparent;")
        fl = QHBoxLayout(frow); fl.setContentsMargins(10, 8, 10, 2); fl.setSpacing(12)
        fl.addWidget(QLabel("篩選："))
        self._filters = {}
        for col, label in self._FILTERS:
            if col not in sub_df.columns:
                continue
            val = str(seed.get(col, "")).strip()
            cb = QCheckBox(f"{label} =")
            cb.setChecked(col in ("InfluenceStatus", "EffectReceive") and val != "")
            cb.stateChanged.connect(self._rebuild)
            combo = _NoscrollCombo()
            combo.setMaximumWidth(150)
            vals = sorted({str(v).strip() for v in sub_df[col] if str(v).strip() != ""})
            combo.addItems(vals)
            if val in vals:
                combo.setCurrentText(val)
            combo.currentTextChanged.connect(self._rebuild)
            self._filters[col] = (cb, combo)
            fl.addWidget(cb); fl.addWidget(combo)
        fl.addStretch(1)
        outer.addWidget(frow)

        self._tbl = QTableWidget()
        self._tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self._tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self._tbl.setSortingEnabled(True)
        self._tbl.verticalHeader().setVisible(False)
        outer.addWidget(self._tbl, 1)

        bb = QDialogButtonBox(QDialogButtonBox.Close)
        bb.rejected.connect(self.reject)
        outer.addWidget(bb)
        self._rebuild()

    def _rebuild(self):
        df = self._df
        mask = pd.Series([True] * len(df), index=df.index)
        for col, (cb, combo) in self._filters.items():
            if cb.isChecked():
                mask &= (df[col].astype(str).str.strip() == combo.currentText())
        sub = df[mask]
        headers = ["來源 ID", "名稱"] + self._cols
        self._tbl.setSortingEnabled(False)
        self._tbl.clear()
        self._tbl.setColumnCount(len(headers))
        self._tbl.setHorizontalHeaderLabels(headers)
        self._tbl.setRowCount(len(sub))
        for r, (_idx, row) in enumerate(sub.iterrows()):
            fkval = str(row[self._fk])
            cells = [fkval, self._name_of(fkval)] + [str(row[c]) for c in self._cols]
            for c, v in enumerate(cells):
                it = QTableWidgetItem(v)
                if headers[c] == "EffectValue":
                    try: it.setData(Qt.EditRole, float(v))
                    except (ValueError, TypeError): pass
                self._tbl.setItem(r, c, it)
        self._tbl.setSortingEnabled(True)
        self._tbl.resizeColumnsToContents()
        self.setWindowTitle(
            self.windowTitle().split("　(")[0] + f"　({len(sub)} 筆符合)")


# ── TableEditor ───────────────────────────────────────────────────────────────

_ALL_GROUPS = "__ALL_GROUPS__"   # sentinel: the "(全部)" pseudo-classification


class CompareView(QWidget):
    """Reference panel comparing several master items side by side (rows = fields,
    columns = items; differing rows highlighted). Items are tracked by primary-key
    value so the set survives reordering/deletes. Can be docked as a tab or floated.
    Read-only — you keep editing in the main panel and this refreshes live."""

    def __init__(self, editor, parent=None):
        super().__init__(parent)
        self._editor = editor
        self._pks = []          # primary-key values being compared
        self.setStyleSheet(f"background:{_C['panel']};")
        v = QVBoxLayout(self); v.setContentsMargins(8, 6, 8, 8); v.setSpacing(6)
        bar = QHBoxLayout(); bar.setSpacing(8)
        self._info = QLabel("尚未加入項目")
        self._info.setStyleSheet(f"color:{_C['txt2']}; font-size:11px; background:transparent;")
        self._diff_only = QCheckBox("只顯示有差異")
        self._diff_only.stateChanged.connect(self.refresh)
        clr = _mk_btn("清空", "ghost"); clr.setFixedHeight(24); clr.clicked.connect(self.clear)
        bar.addWidget(self._info, 1); bar.addWidget(self._diff_only); bar.addWidget(clr)
        v.addLayout(bar)
        # removable chips (one per compared item) — click ✕ to drop it
        self._chips_box = QWidget(); self._chips_box.setStyleSheet("background:transparent;")
        self._chips_flow = FlowLayout(self._chips_box, spacing=4)
        v.addWidget(self._chips_box)
        self._tbl = QTableWidget()
        self._tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self._tbl.horizontalHeader().customContextMenuRequested.connect(self._hdr_menu)
        self._tbl.cellDoubleClicked.connect(self._cell_dblclick)
        self._row_cols = []   # table row → field (column) name, for quick-nav
        v.addWidget(self._tbl, 1)
        hint = QLabel("在「項目」分頁選取後按「加入比較」可累積（可跨分類）；"
                      "右鍵欄位標題可移除；雙擊儲存格可跳到該項目該欄位")
        hint.setWordWrap(True)
        hint.setStyleSheet(f"color:{_C['txt3']}; font-size:10px; background:transparent;")
        v.addWidget(hint)

    def add_indices(self, idxs):
        df = self._editor.df; pk = self._editor.pk_key
        if pk not in df.columns:
            return
        for i in idxs:
            if i in df.index:
                p = str(df.at[i, pk])
                if p not in self._pks:
                    self._pks.append(p)
        self.refresh()

    def clear(self):
        self._pks = []; self.refresh()

    def _hdr_menu(self, pos):
        col = self._tbl.horizontalHeader().logicalIndexAt(pos)
        if col <= 0 or col - 1 >= len(self._pks):
            return
        p = self._pks[col - 1]
        m = QMenu(self)
        m.addAction("從比較移除", lambda: self._remove(p))
        m.exec(self._tbl.horizontalHeader().mapToGlobal(pos))

    def _remove(self, pk):
        if pk in self._pks:
            self._pks.remove(pk)
        self.refresh()

    def _make_chip(self, pk, label):
        w = QFrame()
        w.setStyleSheet(f"QFrame{{background:{_C['card']}; border:1px solid {_C['border']}; border-radius:9px;}}")
        h = QHBoxLayout(w); h.setContentsMargins(8, 2, 4, 2); h.setSpacing(4)
        l = QLabel(label or pk)
        l.setStyleSheet(f"color:{_C['txt']}; background:transparent; border:none; font-size:11px;")
        x = QPushButton("✕"); x.setFixedSize(16, 16); x.setAutoDefault(False)
        x.setStyleSheet(
            f"QPushButton{{background:transparent; border:none; color:{_C['txt3']}; font-size:12px;}}"
            f"QPushButton:hover{{color:{_C['red']};}}")
        x.clicked.connect(lambda *_a, _p=pk: self._remove(_p))
        h.addWidget(l); h.addWidget(x)
        return w

    def _cell_dblclick(self, r, c):
        """Double-click a value cell → jump the main editor to that item's field.
        Column 0 is the field-name column; data columns map to compared items."""
        if c <= 0 or c - 1 >= len(self._pks):
            return
        if not (0 <= r < len(self._row_cols)):
            return
        pk  = self._pks[c - 1]
        col = self._row_cols[r]
        ed  = self._editor
        df  = ed.df
        m = df.index[df[ed.pk_key].astype(str) == str(pk)]
        if not len(m):
            return
        idx = m[0]
        ed.current_cls_val = df.at[idx, ed.cls_key]
        ed._load_cls_list()
        ed._load_item_list()
        ed._load_editor(idx)
        ed.focus_master_field(col)

    def refresh(self, *_a):
        ed = self._editor; df = ed.df; pk = ed.pk_key
        idxs = []
        if pk in df.columns:
            pkser = df[pk].astype(str)
            kept = []
            for p in self._pks:
                m = df.index[pkser == p]
                if len(m):
                    kept.append(p); idxs.append(m[0])
            self._pks = kept
        else:
            self._pks = []
        self._info.setText(
            f"比較 {len(idxs)} 個項目" if idxs
            else "尚未加入項目（在「項目」分頁選取後按「加入比較」）")
        cols_cfg = ed.cfg.get("columns", {})
        disp = lambda c, r: ed._resolve_textref(c, r, cols_cfg)
        diff_only = self._diff_only.isChecked()
        labels = [disp(pk, str(df.at[i, pk])) for i in idxs]
        headers = ["欄位"] + labels
        # rebuild removable chips
        while self._chips_flow.count():
            itm = self._chips_flow.takeAt(0)
            w = itm.widget() if itm else None
            if w is not None:
                w.setParent(None); w.deleteLater()
        for p, lbl in zip(self._pks, labels):
            self._chips_flow.addWidget(self._make_chip(p, lbl))
        self._chips_box.setVisible(bool(self._pks))
        n = len(idxs)
        # entry = (label, vals, differ, nav_col, kind)
        #   kind: "field"=母表欄位 / "sub"=子表欄位 / "hdr"=子表區段 / "subhdr"=第N筆
        entries = []
        for col in list(df.columns):
            vals = [disp(col, str(df.at[i, col])) for i in idxs]
            differ = len(set(vals)) > 1
            if diff_only and not differ:
                continue
            entries.append((col, vals, differ, col, "field"))

        # ── 子表：把每個項目的子表列逐筆並排比較（第 r 筆 vs 第 r 筆）──
        mgr = ed.manager
        for sub_full in [s for s in mgr.sub_tables if s.startswith(ed.table_name + ".")]:
            sub_df = mgr.sub_tables.get(sub_full)
            if sub_df is None or sub_df.empty:
                continue
            sub_name = sub_full[len(ed.table_name) + 1:]
            sub_cfg = ed.cfg.get("sub_tables", {}).get(sub_name, {})
            sub_cols_cfg = sub_cfg.get("columns", {})
            fk = sub_cfg.get("foreign_key") or pk
            if fk not in sub_df.columns:
                continue
            fkser = sub_df[fk].astype(str)
            per_item = [list(sub_df.index[fkser == p]) for p in self._pks]  # 各項目的子表列 index
            maxn = max((len(x) for x in per_item), default=0)
            if maxn == 0:
                continue
            show_cols = [c for c in sub_df.columns if c != fk]
            sub_disp = lambda c, raw: ed._resolve_textref(c, raw, sub_cols_cfg)
            block = []
            for r in range(maxn):
                row_entries = []
                for col in show_cols:
                    vals = [
                        sub_disp(col, str(sub_df.at[per_item[k][r], col]))
                        if r < len(per_item[k]) else ""
                        for k in range(n)
                    ]
                    differ = len(set(vals)) > 1
                    if diff_only and not differ:
                        continue
                    row_entries.append((f"    {col}", vals, differ, None, "sub"))
                if row_entries:
                    block.append((f"  ● 第 {r + 1} 筆", [""] * n, False, None, "subhdr"))
                    block.extend(row_entries)
            if block:
                entries.append((f"▸ 子表：{sub_name}", [""] * n, False, None, "hdr"))
                entries.extend(block)

        self._row_cols = [nav for (_l, _v, _d, nav, _k) in entries]
        self._tbl.clear()
        self._tbl.setColumnCount(len(headers))
        self._tbl.setHorizontalHeaderLabels(headers)
        self._tbl.setRowCount(len(entries))
        for r, (label, vals, differ, _nav, kind) in enumerate(entries):
            name = QTableWidgetItem(label)
            if kind == "hdr":
                name.setForeground(QBrush(QColor(_C["accent"])))
                f = name.font(); f.setBold(True); name.setFont(f)
            elif kind == "subhdr":
                name.setForeground(QBrush(QColor(_C["txt3"])))
            elif differ:
                name.setForeground(QBrush(QColor(_C["yellow"])))
            self._tbl.setItem(r, 0, name)
            for ci, vv in enumerate(vals):
                it = QTableWidgetItem(vv)
                if vv:
                    it.setToolTip(vv)  # 欄寬受限被截斷時，可 hover 看完整內容
                if differ and kind in ("field", "sub"):
                    it.setBackground(QColor(234, 179, 8, 30))
                self._tbl.setItem(r, ci + 1, it)
        # 欄位名稱欄依內容決定寬度（設上限避免過寬）；其餘資料欄依目前視窗寬度平均分配，
        # 這樣單一欄位內容過長也不會把整列撐爆、超出比較視窗的可視範圍。
        hdr = self._tbl.horizontalHeader()
        self._tbl.resizeColumnToContents(0)
        hdr.setSectionResizeMode(0, QHeaderView.Interactive)
        self._tbl.setColumnWidth(0, min(self._tbl.columnWidth(0), 160))
        for c in range(1, len(headers)):
            hdr.setSectionResizeMode(c, QHeaderView.Stretch)


class TableEditor(QWidget):
    status_message = Signal(str, str)
    _sub_clipboard = None  # {"sub": tab_name, "rows": [ {col: val, ...} ]} — shared across items/tables

    def __init__(self, table_name, manager, parent=None):
        super().__init__(parent)
        self.table_name = table_name
        self.manager    = manager
        self.df         = manager.tables[table_name]
        self.cfg        = manager.config.get(table_name, {})
        cols            = list(self.df.columns)
        self.cls_key    = self.cfg.get("classification_key", "") or (cols[0] if cols else "")
        self.pk_key     = self.cfg.get("primary_key",        "") or (cols[0] if cols else "")

        self.current_cls_val    = None
        self.current_master_idx = None
        self.current_master_pk  = None

        self._field_panel: FieldEditorWidget | None = None
        self._sub_panels:  dict[str, SubTablePanel] = {}
        self._sub_tab_order: list[str] = []

        self._setup_ui()
        self._load_cls_list()

    # ── Layout ────────────────────────────────────────────────────────────────

    def _setup_ui(self):
        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ══ LEFT SIDEBAR — classification groups (200px) ══════════════════════
        left = QWidget()
        left.setFixedWidth(200)
        left.setStyleSheet(f"background: {_C['sidebar']};")
        lv = QVBoxLayout(left)
        lv.setContentsMargins(0, 0, 0, 0)
        lv.setSpacing(0)

        lv.addWidget(_sec_lbl(f"Groups · {self.cls_key}"))

        _clsf_wrap = QWidget(); _clsf_wrap.setStyleSheet("background:transparent;")
        _cfl = QHBoxLayout(_clsf_wrap); _cfl.setContentsMargins(8, 0, 8, 6); _cfl.setSpacing(0)
        self._cls_filter = QLineEdit()
        self._cls_filter.setPlaceholderText("篩選分類…")
        self._cls_filter.setClearButtonEnabled(True)
        self._cls_filter.textChanged.connect(self._apply_cls_filter)
        _cfl.addWidget(self._cls_filter)
        lv.addWidget(_clsf_wrap)

        self._cls_list = QListWidget()
        self._cls_list.setObjectName("cls-list")
        self._cls_list.setFocusPolicy(Qt.NoFocus)
        self._cls_list.currentItemChanged.connect(self._on_cls_changed)
        self._cls_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self._cls_list.customContextMenuRequested.connect(self._cls_ctx_menu)
        lv.addWidget(self._cls_list, 1)

        lv.addWidget(_hsep())

        cls_btns = QHBoxLayout()
        cls_btns.setContentsMargins(8, 6, 8, 6)
        cls_btns.setSpacing(4)
        b_add_cls = _mk_btn("分類", "ghost", icon="folder-plus", icon_color="#10B981"); b_add_cls.clicked.connect(self.add_classification)
        b_del_cls = _mk_btn("", "ghost", icon="trash", icon_color="#EF4444"); b_del_cls.setFixedWidth(30); b_del_cls.setToolTip("刪除分類"); b_del_cls.clicked.connect(self.delete_classification)
        b_up_cls  = _mk_btn("", "ghost", icon="arrow-up");   b_up_cls.setFixedWidth(30); b_up_cls.setToolTip("上移"); b_up_cls.clicked.connect(lambda: self.move_classification(-1))
        b_dn_cls  = _mk_btn("", "ghost", icon="arrow-down"); b_dn_cls.setFixedWidth(30); b_dn_cls.setToolTip("下移"); b_dn_cls.clicked.connect(lambda: self.move_classification(1))
        for b in [b_add_cls, b_del_cls, b_up_cls, b_dn_cls]:
            cls_btns.addWidget(b)
        lv.addLayout(cls_btns)

        root.addWidget(left)
        root.addWidget(_vsep())

        # ══ MIDDLE — search + card list (flex) ════════════════════════════════
        mid = QWidget()
        mid.setMinimumWidth(240)
        mid.setStyleSheet(f"background: {_C['bg']};")
        mv = QVBoxLayout(mid)
        mv.setContentsMargins(0, 0, 0, 0)
        mv.setSpacing(0)

        # Search + action bar
        action_bar = QWidget()
        action_bar.setStyleSheet(
            f"background: {_C['sidebar']}; border-bottom: 1px solid {_C['border']};"
        )
        action_bar.setFixedHeight(50)
        alo = QHBoxLayout(action_bar)
        alo.setContentsMargins(10, 0, 10, 0)
        alo.setSpacing(6)

        search_icon = QLabel("⌕")
        search_icon.setStyleSheet(f"color:{_C['txt2']}; background:transparent; font-size:16px;")
        self._filter_edit = QLineEdit()
        self._filter_edit.setPlaceholderText("搜尋項目…")
        self._filter_edit.setClearButtonEnabled(True)
        self._filter_edit.textChanged.connect(self._apply_filter)
        alo.addWidget(search_icon)
        alo.addWidget(self._filter_edit, 1)
        alo.addWidget(_vsep())

        for text, role, icon, tip, slot in [
            ("新增", "success", "circle-plus", "",   self.add_master_item),
            ("範本", "",        "clipboard-plus",
             "從範本新增（含子表列）；在項目上按右鍵可「設為範本」", self._template_menu),
            ("複製", "",        "copy",        "",   self.copy_master_item),
            ("",     "ghost",   "arrow-up",    "上移", lambda: self.move_master_item(-1)),
            ("",     "ghost",   "arrow-down",  "下移", lambda: self.move_master_item(1)),
            ("刪除", "danger",  "trash",       "",   self.delete_master_item),
            ("加入比較", "ghost", "columns",   "把選取項目加入「比較」分頁（Ctrl／Shift 多選，可跨分類）", self.add_to_compare),
        ]:
            b = _mk_btn(text, role, icon=icon)
            b.setFixedHeight(32)
            if tip:
                b.setToolTip(tip)
            if not text:
                b.setFixedWidth(36)
            b.clicked.connect(slot)
            alo.addWidget(b)

        items_page = QWidget(); items_page.setStyleSheet("background:transparent;")
        ipv = QVBoxLayout(items_page); ipv.setContentsMargins(0, 0, 0, 0); ipv.setSpacing(0)
        ipv.addWidget(action_bar)

        # Card list
        self._card_list = QListWidget()
        self._card_list.setObjectName("card-list")
        self._card_list.setItemDelegate(ItemCardDelegate(self._card_list))
        self._card_list.setMouseTracking(True)
        self._card_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self._card_list.setFocusPolicy(Qt.NoFocus)
        self._card_list.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self._card_list.currentItemChanged.connect(self._on_item_changed)
        self._card_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self._card_list.customContextMenuRequested.connect(self._item_ctx_menu)
        ipv.addWidget(self._card_list, 1)

        # 項目 / 比較 tabs, with a float-out corner button
        self._mid_tabs = QTabWidget()
        self._mid_tabs.setDocumentMode(True)
        self._mid_tabs.addTab(items_page, "項目")
        self._compare_view = CompareView(self)
        self._mid_tabs.addTab(self._compare_view, "比較")
        self._compare_float = None
        self._compare_float_btn = _mk_btn("⤢ 浮動", "ghost")
        self._compare_float_btn.setFixedHeight(24)
        self._compare_float_btn.setToolTip("把「比較」浮出成獨立視窗（可一邊編輯一邊參考）／收回")
        self._compare_float_btn.clicked.connect(self._toggle_float_compare)
        self._mid_tabs.setCornerWidget(self._compare_float_btn)
        mv.addWidget(self._mid_tabs, 1)

        # mid is placed into the draggable splitter below

        # ══ RIGHT — field editor + JSON preview + sub-tables ══════════════════
        right = QWidget()
        right.setMinimumWidth(320)
        right.setStyleSheet(f"background: {_C['panel']};")
        self._right_panel = right
        rv = QVBoxLayout(right)
        rv.setContentsMargins(0, 0, 0, 0)
        rv.setSpacing(0)

        # Panel header (item ID) + ⊞ 欄位 button
        _ph_widget = QWidget()
        _ph_widget.setFixedHeight(44)
        _ph_widget.setStyleSheet(
            f"background: {_C['panel']}; border-bottom: 1px solid {_C['border']};"
        )
        _ph_lo = QHBoxLayout(_ph_widget)
        _ph_lo.setContentsMargins(14, 0, 8, 0)
        _ph_lo.setSpacing(6)
        self._panel_hdr = QLabel("— 請選擇項目 —")
        self._panel_hdr.setStyleSheet(
            f"color: {_C['txt2']}; font-size: 13px; font-weight: 500; background: transparent;"
        )
        _col_btn = _mk_btn("欄位 ▾", "ghost", icon="columns")
        _col_btn.setFixedHeight(28)
        _col_menu = QMenu(_col_btn)
        _col_menu.addAction(_ti_icon("plus", "#10B981"), "新增欄位", self.add_master_column)
        _col_menu.addAction(_ti_icon("pencil", "#A5B4FC"), "重新命名", self.rename_master_column)
        _col_menu.addAction(_ti_icon("trash", "#EF4444"), "刪除欄位", self.delete_master_column)
        _col_btn.setMenu(_col_menu)
        _ph_lo.addWidget(self._panel_hdr, 1)
        _ph_lo.addWidget(_col_btn)
        rv.addWidget(_ph_widget)

        # Right splitter: field editor (top) / sub-tables (bottom)
        rsplit = QSplitter(Qt.Vertical)
        rsplit.setHandleWidth(2)
        rsplit.setStyleSheet(f"QSplitter::handle:vertical {{ background: {_C['border']}; }}")

        # ── Field + JSON area ──
        field_area = QWidget()
        field_area.setStyleSheet(f"background: {_C['panel']};")
        fv = QVBoxLayout(field_area)
        fv.setContentsMargins(0, 0, 0, 0)
        fv.setSpacing(0)

        self._field_container = QWidget()
        self._field_container.setStyleSheet(f"background: {_C['panel']};")
        fclo = QVBoxLayout(self._field_container)
        fclo.setContentsMargins(0, 0, 0, 0)
        fv.addWidget(self._field_container, 1)

        # JSON Preview (collapsible)
        self._json_toggle = QPushButton("▶  JSON PREVIEW")
        self._json_toggle.setStyleSheet(
            f"QPushButton {{ background:{_C['sidebar']}; border:none; "
            f"border-top:1px solid {_C['border']}; border-bottom:1px solid {_C['border']}; "
            f"color:{_C['txt3']}; font-size:10px; font-weight:600; letter-spacing:1px; "
            f"text-align:left; padding:7px 14px; }}"
            f"QPushButton:hover {{ background:{_C['cardH']}; color:{_C['txt2']}; }}"
        )
        self._json_toggle.clicked.connect(self._toggle_json)
        self._json_preview = QTextEdit()
        self._json_preview.setObjectName("code-view")
        self._json_preview.setReadOnly(True)
        self._json_preview.setMaximumHeight(180)
        self._json_preview.hide()
        fv.addWidget(self._json_toggle)
        fv.addWidget(self._json_preview)
        rsplit.addWidget(field_area)

        # ── Sub-tables area ──
        sub_area = QWidget()
        sub_area.setStyleSheet(f"background: {_C['panel']};")
        sv = QVBoxLayout(sub_area)
        sv.setContentsMargins(0, 0, 0, 0)
        sv.setSpacing(0)

        sub_hdr = QWidget()
        self._sub_hdr = sub_hdr
        sub_hdr.setFixedHeight(40)
        sub_hdr.setStyleSheet(
            f"background:{_C['sidebar']}; border-top:1px solid {_C['border']}; border-bottom:1px solid {_C['border']};"
        )
        sh = QHBoxLayout(sub_hdr)
        sh.setContentsMargins(12, 0, 8, 0)
        sh.setSpacing(4)
        lbl_sub = QLabel("SUB-TABLES")
        lbl_sub.setStyleSheet(
            f"color:{_C['txt3']}; font-size:10px; font-weight:600; letter-spacing:1px; background:transparent;"
        )
        sh.addWidget(lbl_sub)
        self._sub_picker = _NoscrollCombo()
        self._sub_picker.setFixedWidth(150)
        self._sub_picker.setToolTip("跳到子表")
        self._sub_picker.activated.connect(self._on_sub_picker)
        sh.addWidget(self._sub_picker)
        sh.addStretch(1)

        for text, role, icon, tip, slot in [
            ("新增列", "success", "circle-plus", "",   self.add_sub_item),
            ("複製",   "ghost",   "copy",        "原地複製一列", self.copy_sub_item),
            ("",       "ghost",   "arrow-up",    "上移", lambda: self.move_sub_item(-1)),
            ("",       "ghost",   "arrow-down",  "下移", lambda: self.move_sub_item(1)),
            ("刪除",   "danger",  "trash",       "",   self.delete_sub_item),
            ("複製列", "ghost",   "copy",        "複製選取列到剪貼簿（可跨筆/跨子表貼上）", self.copy_sub_rows),
            ("貼上列", "ghost",   "clipboard-plus", "把剪貼簿的列貼到目前子表", self.paste_sub_rows),
        ]:
            b = _mk_btn(text, role, icon=icon)
            b.setFixedHeight(28)
            if tip:
                b.setToolTip(tip)
            if not text:
                b.setFixedWidth(32)
            b.clicked.connect(slot)
            sh.addWidget(b)

        sh.addWidget(_vsep())

        _scol_btn = _mk_btn("欄位 ▾", "ghost", icon="columns"); _scol_btn.setFixedHeight(28)
        _scol_menu = QMenu(_scol_btn)
        _scol_menu.addAction(_ti_icon("plus", "#10B981"), "新增欄位", self.add_sub_column)
        _scol_menu.addAction(_ti_icon("pencil", "#A5B4FC"), "重新命名", self.rename_sub_column)
        _scol_menu.addAction(_ti_icon("trash", "#EF4444"), "刪除欄位", self.delete_sub_column)
        _scol_btn.setMenu(_scol_menu)
        sh.addWidget(_scol_btn)

        _stbl_btn = _mk_btn("子表 ▾", "ghost", icon="table"); _stbl_btn.setFixedHeight(28)
        _stbl_menu = QMenu(_stbl_btn)
        _stbl_menu.addAction(_ti_icon("plus", "#10B981"), "新增子表", self.add_sub_table)
        _stbl_menu.addAction(_ti_icon("trash", "#EF4444"), "刪除子表", self.delete_sub_table)
        _stbl_btn.setMenu(_stbl_menu)
        sh.addWidget(_stbl_btn)

        sv.addWidget(sub_hdr)
        self._sub_tabs = QTabWidget()
        self._sub_tabs.setObjectName("sub-tabs")
        self._sub_tabs.setDocumentMode(True)
        self._sub_tabs.currentChanged.connect(self._sync_sub_picker)
        sv.addWidget(self._sub_tabs, 1)
        sub_area.setMinimumHeight(110)     # ensure header + tab bar always visible
        rsplit.addWidget(sub_area)

        rsplit.setSizes([360, 200])
        rsplit.setCollapsible(0, False)
        rsplit.setCollapsible(1, False)
        rv.addWidget(rsplit, 1)

        # Draggable divider: middle list ↔ right panel (field editor + sub-tables)
        mid_right_split = QSplitter(Qt.Horizontal)
        mid_right_split.setObjectName("mid-right-split")
        mid_right_split.setStyleSheet(
            f"QSplitter#mid-right-split::handle:horizontal "
            f"{{ background: {_C['border']}; width: 4px; }}"
        )
        mid_right_split.setHandleWidth(4)
        mid_right_split.setChildrenCollapsible(False)
        mid_right_split.addWidget(mid)
        mid_right_split.addWidget(right)
        mid_right_split.setStretchFactor(0, 1)
        mid_right_split.setStretchFactor(1, 0)
        mid_right_split.setSizes([620, 440])
        root.addWidget(mid_right_split, 1)

        self._build_sub_tabs()

    # ── JSON preview ──────────────────────────────────────────────────────────

    def _toggle_json(self):
        visible = self._json_preview.isVisible()
        self._json_preview.setVisible(not visible)
        self._json_toggle.setText(
            ("▼" if not visible else "▶") + "  JSON PREVIEW"
        )

    def _update_json(self, row_data):
        data = {col: val for col, val in row_data.items()}
        raw  = json.dumps(data, ensure_ascii=False, indent=2, default=str)
        hl   = _json_highlight(raw)
        self._json_preview.setHtml(
            f'<div style="font-family:Consolas;font-size:11px;line-height:1.6;">'
            f'<pre style="margin:0;">{hl}</pre></div>'
        )

    # ── Classification ────────────────────────────────────────────────────────

    def _load_cls_list(self):
        self._cls_list.blockSignals(True)
        self._cls_list.clear()
        if self.cls_key not in self.df.columns:
            self._cls_list.blockSignals(False)
            return
        # "(全部)" pseudo-group on top → middle list shows every item (cross-class)
        all_item = QListWidgetItem(f"  (全部)  ({len(self.df)})")
        all_item.setData(Qt.UserRole, _ALL_GROUPS)
        all_item.setForeground(QBrush(QColor(_C["txtAcc"])))
        self._cls_list.addItem(all_item)
        if self.current_cls_val == _ALL_GROUPS:
            self._cls_list.setCurrentItem(all_item)
        groups = self.df[self.cls_key].unique()
        for g in groups:
            cat   = _cat_for(str(g))
            count = int((self.df[self.cls_key] == g).sum())
            item  = QListWidgetItem(f"  {g}  ({count})")
            item.setData(Qt.UserRole, g)
            item.setForeground(QBrush(QColor(cat["text"])))
            self._cls_list.addItem(item)
            if self.current_cls_val != _ALL_GROUPS and str(g) == str(self.current_cls_val):
                self._cls_list.setCurrentItem(item)
        self._cls_list.blockSignals(False)
        self._apply_cls_filter()

    def _apply_cls_filter(self, text=None):
        if not hasattr(self, "_cls_filter"):
            return
        q = (text if text is not None else self._cls_filter.text()).strip().lower()
        for i in range(self._cls_list.count()):
            it = self._cls_list.item(i)
            g = it.data(Qt.UserRole)
            if g == _ALL_GROUPS:
                it.setHidden(False); continue
            it.setHidden(q != "" and q not in str(g).lower())

    def _on_cls_changed(self, cur, _prev):
        if cur is None:
            return
        self.current_cls_val = cur.data(Qt.UserRole)
        self._load_item_list()

    def add_classification(self):
        if self.cls_key not in self.df.columns:
            col, ok = QInputDialog.getText(self, "設定分類欄位", "分類欄位名稱:")
            if not ok or not col.strip():
                return
            col = col.strip()
            self.manager.add_column(self.table_name, col)
            self.df = self.manager.tables[self.table_name]
            self.cls_key = col
            self.cfg["classification_key"] = col
            if not self.pk_key or self.pk_key not in self.df.columns:
                self.pk_key = col
                self.cfg["primary_key"] = col
        name, ok = QInputDialog.getText(self, "新增分類", "分類名稱:")
        if not ok or not name.strip():
            return
        name = name.strip()
        if name in self.df[self.cls_key].values:
            QMessageBox.warning(self, "錯誤", "此分類已存在"); return
        new_row = {col: "" for col in self.df.columns}
        new_row[self.cls_key] = name
        if self.pk_key in self.df.columns and self.pk_key != self.cls_key:
            new_id, ok2 = QInputDialog.getText(self, "新增分類", f"首筆資料的 {self.pk_key}:")
            if not ok2 or not new_id.strip(): return
            new_row[self.pk_key] = new_id.strip()
        self.df = pd.concat([self.df, pd.DataFrame([new_row])], ignore_index=True)
        self.manager.tables[self.table_name] = self.df
        self.manager.dirty = True
        self._reload_all(select_cls=name)

    def delete_classification(self):
        g = self.current_cls_val
        if g is None or g == _ALL_GROUPS: return
        if QMessageBox.question(self, "確認刪除", f"刪除分類 [{g}] 及其所有資料？",
                                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        self.df = self.df[self.df[self.cls_key] != g].reset_index(drop=True)
        self.manager.tables[self.table_name] = self.df
        self.manager.dirty = True
        self.current_cls_val = None
        self.current_master_idx = None
        self._reload_all()

    def move_classification(self, delta):
        if self.current_cls_val is None or self.current_cls_val == _ALL_GROUPS: return
        groups = list(self.df[self.cls_key].unique())
        try:    pos = [str(g) for g in groups].index(str(self.current_cls_val))
        except: return
        new_pos = pos + delta
        if new_pos < 0 or new_pos >= len(groups): return
        groups[pos], groups[new_pos] = groups[new_pos], groups[pos]
        self.df = pd.concat(
            [self.df[self.df[self.cls_key] == g] for g in groups], ignore_index=True
        )
        self.manager.tables[self.table_name] = self.df
        self.manager.dirty = True
        self._reload_all(select_cls=self.current_cls_val)

    # ── Item list (card list) ─────────────────────────────────────────────────

    def _load_item_list(self):
        query    = self._filter_edit.text().strip().lower()
        self._card_list.blockSignals(True)
        self._card_list.clear()
        if self.current_cls_val is None or self.pk_key not in self.df.columns:
            self._card_list.blockSignals(False)
            return

        all_mode = (self.current_cls_val == _ALL_GROUPS)
        sub_df   = self.df if all_mode else self.df[self.df[self.cls_key] == self.current_cls_val]
        sub_col  = next(
            (c for c in self.df.columns if c != self.pk_key and c != self.cls_key),
            None
        )

        # Build text-ref resolver for display (per-column source)
        cols_cfg = self.cfg.get("columns", {})

        def _disp(col, raw):
            return self._resolve_textref(col, raw, cols_cfg)

        tpl_pks = self._template_pks()

        for df_idx, row in sub_df.iterrows():
            pk_raw   = str(row[self.pk_key])
            sub_raw  = str(row[sub_col]) if sub_col else ""
            pk_disp  = _disp(self.pk_key, pk_raw)
            if pk_raw in tpl_pks:
                pk_disp = f"⭐ {pk_disp}"
            sub_disp = _disp(sub_col, sub_raw) if sub_col else ""

            if query and query not in pk_raw.lower() and query not in pk_disp.lower() \
                     and query not in sub_raw.lower() and query not in sub_disp.lower():
                continue

            cat_val = str(row[self.cls_key]) if (all_mode and self.cls_key in self.df.columns) \
                      else str(self.current_cls_val)
            item = QListWidgetItem()
            item.setData(Qt.UserRole,            df_idx)
            item.setData(ItemCardDelegate.R_PK,  pk_disp)
            item.setData(ItemCardDelegate.R_SUB, sub_disp if not all_mode else f"{cat_val} · {sub_disp}")
            item.setData(ItemCardDelegate.R_CAT, cat_val)
            item.setData(ItemCardDelegate.R_VIO,
                         self.manager.validator.record_violation_severity(
                             self.table_name, df_idx))
            self._card_list.addItem(item)
            if df_idx == self.current_master_idx:
                self._card_list.setCurrentItem(item)

        self._card_list.blockSignals(False)
        if hasattr(self, "_compare_view"):
            self._compare_view.refresh()

    def _apply_filter(self, _text=""):
        self._load_item_list()

    def _on_item_changed(self, cur, _prev):
        if cur is None: return
        self._load_editor(cur.data(Qt.UserRole))

    def add_master_item(self):
        if self.current_cls_val is None or self.current_cls_val == _ALL_GROUPS:
            QMessageBox.warning(self, "提示", "請先選擇一個分類（「(全部)」檢視下無法新增）"); return
        new_id, ok = QInputDialog.getText(self, "新增項目", f"{self.pk_key}:")
        if not ok or not new_id.strip(): return
        new_id = new_id.strip()
        if new_id in self.df[self.pk_key].astype(str).values:
            QMessageBox.warning(self, "錯誤", "此 ID 已存在"); return
        new_row = {col: "" for col in self.df.columns}
        new_row[self.cls_key] = self.current_cls_val
        new_row[self.pk_key]  = new_id
        cls_rows     = self.df[self.df[self.cls_key] == self.current_cls_val]
        insert_after = cls_rows.index.max() + 1 if not cls_rows.empty else len(self.df)
        top, bot = self.df.iloc[:insert_after], self.df.iloc[insert_after:]
        self.df  = pd.concat([top, pd.DataFrame([new_row]), bot], ignore_index=True)
        self.manager.tables[self.table_name] = self.df
        self.manager.dirty = True
        new_df_idx = self.df[self.df[self.pk_key].astype(str) == str(new_id)].index[0]
        self._reload_all(select_cls=self.current_cls_val, select_idx=new_df_idx)

    def copy_master_item(self):
        if self.current_master_idx is None: return
        new_id, ok = QInputDialog.getText(self, "複製項目", f"新的 {self.pk_key}:")
        if not ok or not new_id.strip(): return
        new_id = new_id.strip()
        if new_id in self.df[self.pk_key].astype(str).values:
            QMessageBox.warning(self, "錯誤", "此 ID 已存在"); return
        new_row = self.df.loc[self.current_master_idx].copy()
        new_row[self.pk_key] = new_id
        self.df = pd.concat([self.df, pd.DataFrame([new_row])], ignore_index=True)
        self.manager.tables[self.table_name] = self.df
        self.manager.dirty = True
        new_df_idx = self.df[self.df[self.pk_key].astype(str) == str(new_id)].index[0]
        self._reload_all(select_cls=self.current_cls_val, select_idx=new_df_idx)

    def delete_master_item(self):
        if self.current_master_idx is None: return
        pk = str(self.df.at[self.current_master_idx, self.pk_key])
        self.df.drop(self.current_master_idx, inplace=True)
        self.df.reset_index(drop=True, inplace=True)
        self.manager.tables[self.table_name] = self.df
        self.manager.dirty = True
        self.current_master_idx = None
        self._reload_all(select_cls=self.current_cls_val)
        self.status_message.emit(f"已刪除 {pk}", _C["yellow"])

    def move_master_item(self, delta):
        if self.current_master_idx is None or self.current_cls_val is None: return
        cls_idxs = list(self.df[self.df[self.cls_key] == self.current_cls_val].index)
        try:    pos = cls_idxs.index(self.current_master_idx)
        except: return
        new_pos = pos + delta
        if new_pos < 0 or new_pos >= len(cls_idxs): return
        cls_idxs[pos], cls_idxs[new_pos] = cls_idxs[new_pos], cls_idxs[pos]
        other_idxs = [i for i in self.df.index if i not in cls_idxs]
        all_cls    = list(self.df[self.df[self.cls_key] == self.current_cls_val].index)
        first_cls  = all_cls[0]
        before = [i for i in other_idxs if i < first_cls]
        after  = [i for i in other_idxs if i > max(all_cls)]
        self.df = self.df.loc[before + cls_idxs + after].reset_index(drop=True)
        self.manager.tables[self.table_name] = self.df
        self.manager.dirty = True
        new_idx = self.df[self.df[self.pk_key].astype(str) == str(self.current_master_pk)].index
        self._reload_all(select_cls=self.current_cls_val,
                         select_idx=new_idx[0] if not new_idx.empty else None)

    # ── Field editor ──────────────────────────────────────────────────────────

    def _load_editor(self, df_idx):
        self.current_master_idx = df_idx
        if df_idx not in self.df.index:
            return
        row_data = self.df.loc[df_idx]
        self.current_master_pk = row_data[self.pk_key]

        # Update panel header
        cat = _cat_for(str(self.current_cls_val))
        self._panel_hdr.setText(f"{self.current_master_pk}")
        self._panel_hdr.setStyleSheet(
            f"color:{cat['text']}; font-family:Consolas; font-size:13px; "
            f"font-weight:700; background:transparent;"
        )

        if self._field_panel is None:
            self._field_panel = FieldEditorWidget(self._field_container)
            self._field_panel.field_changed.connect(self._on_field_change)
            self._field_container.layout().addWidget(self._field_panel)
            self._field_panel.build_for(self.df, self.cfg, self.table_name, self.manager)

        self._field_panel.load_row(row_data, df_idx)
        self._field_panel.refresh_binding()
        self._field_panel.refresh_array_suggestions()
        self._update_json(row_data)
        self._refresh_sub_tables()

    def add_to_compare(self):
        idxs = [it.data(Qt.UserRole) for it in self._card_list.selectedItems()]
        idxs = [i for i in idxs if i in self.df.index]
        if not idxs:
            self.status_message.emit("請先選取要加入比較的項目", _C["yellow"]); return
        self._compare_view.add_indices(idxs)
        if self._compare_float is None:
            self._mid_tabs.setCurrentWidget(self._compare_view)
        else:
            self._compare_float.raise_()
        self.status_message.emit(f"已加入 {len(idxs)} 個到比較", _C["green"])

    def _toggle_float_compare(self):
        if self._compare_float is None:
            # dock → float: pull the compare view out of the tab into a window
            idx = self._mid_tabs.indexOf(self._compare_view)
            if idx >= 0:
                self._mid_tabs.removeTab(idx)
            win = QDialog(self)
            win.setWindowTitle("比較（參考）")
            win.setStyleSheet(APP_QSS)
            win.resize(760, 600)
            lo = QVBoxLayout(win); lo.setContentsMargins(0, 0, 0, 0)
            lo.addWidget(self._compare_view)
            self._compare_view.show()        # removeTab hid it — show again
            self._compare_view.refresh()
            win.finished.connect(lambda *_: self._dock_compare())
            self._compare_float = win
            self._compare_float_btn.setText("⤡ 收回")
            self._mid_tabs.setCurrentIndex(0)
            win.show()
        else:
            self._compare_float.close()  # triggers finished → _dock_compare

    def _dock_compare(self):
        if self._compare_float is None:
            return
        win = self._compare_float
        self._compare_float = None
        self._compare_view.setParent(None)
        self._mid_tabs.addTab(self._compare_view, "比較")
        self._compare_float_btn.setText("⤢ 浮動")
        win.deleteLater()

    def _on_field_change(self, col, value):
        if self.current_master_idx is None: return
        self.manager.update_cell(self.table_name, self.current_master_idx, col, value)
        # If the edited column drives the left/middle lists, refresh them so the
        # grouping / labels reflect the new value.
        if col == self.cls_key:
            self._load_cls_list()    # groups + counts
            self._load_item_list()   # current group's items (regrouped item leaves)
        elif col == self.pk_key:
            self.current_master_pk = str(value)
            if self._panel_hdr is not None:
                self._panel_hdr.setText(f"{self.current_master_pk}")
            self._load_item_list()
        # keep the live comparison reference in sync with edits
        if hasattr(self, "_compare_view"):
            self._compare_view.refresh()
        self._refresh_validation_visuals()

    # ── Sub-tables ────────────────────────────────────────────────────────────

    def _build_sub_tabs(self):
        self._sub_tabs.clear()
        self._sub_panels.clear()
        self._sub_tab_order = []
        prefix = self.table_name + "."
        for key in self.manager.sub_tables:
            if not key.startswith(prefix):
                continue
            tab_name = key[len(prefix):]
            self._sub_tab_order.append(tab_name)
            # Create the real panel immediately — no placeholder swap needed
            sub_cfg  = self.cfg.get("sub_tables", {}).get(tab_name, {})
            cols_cfg = sub_cfg.get("columns", {})
            panel = SubTablePanel(key, cols_cfg, self.manager)
            panel.row_deleted.connect(self._on_sub_delete)
            panel.copy_requested.connect(self.copy_sub_rows)
            panel.paste_requested.connect(self.paste_sub_rows)
            panel.compare_requested.connect(self.compare_sub_effect)
            panel.set_validation_cb(self._refresh_validation_visuals)
            self._sub_panels[tab_name] = panel
            idx = self._sub_tabs.addTab(panel, tab_name)
            note = sub_cfg.get("note", "")
            if note:
                self._sub_tabs.setTabToolTip(idx, note)

        if not self._sub_tab_order:
            no_sub = QLabel("此表格無巢狀子表")
            no_sub.setAlignment(Qt.AlignCenter)
            no_sub.setStyleSheet(
                f"color:{_C['txt3']}; font-size:12px; background:{_C['panel']};"
            )
            self._sub_tabs.addTab(no_sub, "—")

        # Sub-table quick-jump picker (handy when there are many sub-tables)
        if hasattr(self, "_sub_picker"):
            self._sub_picker.blockSignals(True)
            self._sub_picker.clear()
            self._sub_picker.addItems(self._sub_tab_order)
            self._sub_picker.blockSignals(False)
            self._sub_picker.setVisible(len(self._sub_tab_order) > 1)

        self._fit_right_panel()

        self.status_message.emit(
            f"從表: 偵測到 {len(self._sub_tab_order)} 個"
            + (f"  ({', '.join(self._sub_tab_order)})" if self._sub_tab_order else ""),
            _C["txt2"],
        )

    def flush_pending_edits(self):
        """Commit any still-open sub-table cell editor (Ctrl+S doesn't move
        focus, so a pending combo/line edit would otherwise miss the save)."""
        for panel in self._sub_panels.values():
            if panel is not None:
                panel.flush_pending_edit()

    def _refresh_sub_tables(self, revalidate=False):
        """Reload each sub-table panel with rows matching the currently selected master pk."""
        if revalidate:
            # sub-row add/delete/move/copy resets sub indexes → rebuild violations
            self.manager.validator.validate_table(self.table_name)
        if self.current_master_pk is None:
            return
        prefix = self.table_name + "."
        for tab_name in self._sub_tab_order:
            panel = self._sub_panels.get(tab_name)
            if panel is None:
                continue
            full   = prefix + tab_name
            sub_df = self.manager.sub_tables.get(full)
            if sub_df is None:
                continue
            sub_cfg  = self.cfg.get("sub_tables", {}).get(tab_name, {})
            fk_key   = sub_cfg.get("foreign_key", self.pk_key)
            cols_cfg = sub_cfg.get("columns", {})
            try:
                filtered = sub_df[sub_df[fk_key].astype(str) == str(self.current_master_pk)]
            except KeyError:
                # FK column not found — fall back to first column
                fk_key   = sub_df.columns[0] if len(sub_df.columns) > 0 else None
                if fk_key is None:
                    continue
                filtered = sub_df[sub_df[fk_key].astype(str) == str(self.current_master_pk)]
            panel.reload(filtered, cols_cfg)
        self._update_sub_tab_vio_colors()

    # ── Validation visuals ─────────────────────────────────────────────────────

    def _refresh_validation_visuals(self):
        """Refresh every validation/binding visual for the current record:
        field styles+visibility, the item card's dot, sub-tab colors, sub-view
        column hiding & repaints. Called after cell edits and rules change."""
        if self._field_panel is not None:
            self._field_panel.refresh_validation()
            self._field_panel.refresh_binding()
            self._field_panel.refresh_array_suggestions()   # context 欄改了會換建議
        it = self._card_list.currentItem()
        if it is not None:
            df_idx = it.data(Qt.UserRole)
            it.setData(ItemCardDelegate.R_VIO,
                       self.manager.validator.record_violation_severity(
                           self.table_name, df_idx))
        self._update_sub_tab_vio_colors()
        for panel in self._sub_panels.values():
            if panel is not None:
                panel.apply_column_binding()
                panel._view.viewport().update()

    def _update_sub_tab_vio_colors(self):
        """Colour sub-tab titles red when the current record has violating rows
        in that sub table (count shown in the tooltip alongside the note)."""
        validator = self.manager.validator
        bar = self._sub_tabs.tabBar()
        prefix = self.table_name + "."
        for i in range(self._sub_tabs.count()):
            tab_name = self._sub_tabs.tabText(i)
            panel = self._sub_panels.get(tab_name)
            if panel is None:
                continue
            full = prefix + tab_name
            cnt = sum(1 for ri in panel._model.df.index
                      if validator.row_has_violation(full, ri))
            note = self.cfg.get("sub_tables", {}).get(tab_name, {}).get("note", "")
            if cnt:
                bar.setTabTextColor(i, QColor(_C["red"]))
                self._sub_tabs.setTabToolTip(
                    i, f"⚠ {cnt} 列驗證未通過" + (f"\n{note}" if note else ""))
            else:
                bar.setTabTextColor(i, QColor())   # invalid → default colour
                self._sub_tabs.setTabToolTip(i, note)

    def _on_sub_delete(self, sheet_full, df_idx):
        sub_df = self.manager.sub_tables.get(sheet_full)
        if sub_df is None or df_idx not in sub_df.index: return
        sub_df.drop(df_idx, inplace=True)
        sub_df.reset_index(drop=True, inplace=True)
        self.manager.sub_tables[sheet_full] = sub_df
        self.manager.dirty = True
        self._refresh_sub_tables(revalidate=True)
        self._refresh_validation_visuals()
        self.status_message.emit("已刪除子表列", _C["yellow"])

    def _current_sub_panel(self) -> SubTablePanel | None:
        idx = self._sub_tabs.currentIndex()
        if idx < 0: return None
        return self._sub_panels.get(self._sub_tabs.tabText(idx))

    def add_sub_item(self):
        if self.current_master_pk is None: return
        panel = self._current_sub_panel()
        if panel is None: return
        tab_name = self._sub_tabs.tabText(self._sub_tabs.currentIndex())
        full     = self.table_name + "." + tab_name
        sub_df   = self.manager.sub_tables.get(full)
        if sub_df is None: return
        sub_cfg  = self.cfg.get("sub_tables", {}).get(tab_name, {})
        fk_key   = sub_cfg.get("foreign_key", self.pk_key)
        new_row  = {col: "" for col in sub_df.columns}
        new_row[fk_key] = self.current_master_pk
        siblings = sub_df[sub_df[fk_key].astype(str) == str(self.current_master_pk)]
        insert_at = siblings.index.max() + 1 if not siblings.empty else len(sub_df)
        top, bot = sub_df.iloc[:insert_at], sub_df.iloc[insert_at:]
        self.manager.sub_tables[full] = pd.concat([top, pd.DataFrame([new_row]), bot], ignore_index=True)
        self.manager.dirty = True
        self._refresh_sub_tables(revalidate=True)
        self._refresh_validation_visuals()

    def delete_sub_item(self):
        panel = self._current_sub_panel()
        if panel is None: return
        tab_name = self._sub_tabs.tabText(self._sub_tabs.currentIndex())
        full     = self.table_name + "." + tab_name
        df_idx   = panel.selected_df_index()
        if df_idx is None: return
        self._on_sub_delete(full, df_idx)

    def move_sub_item(self, delta):
        panel = self._current_sub_panel()
        if panel is None: return
        tab_name = self._sub_tabs.tabText(self._sub_tabs.currentIndex())
        full     = self.table_name + "." + tab_name
        sub_df   = self.manager.sub_tables.get(full)
        if sub_df is None: return
        df_idx   = panel.selected_df_index()
        if df_idx is None: return
        sub_cfg  = self.cfg.get("sub_tables", {}).get(tab_name, {})
        fk_key   = sub_cfg.get("foreign_key", self.pk_key)
        siblings = list(sub_df[sub_df[fk_key].astype(str) == str(self.current_master_pk)].index)
        try:    pos = siblings.index(df_idx)
        except: return
        new_pos = pos + delta
        if new_pos < 0 or new_pos >= len(siblings): return
        siblings[pos], siblings[new_pos] = siblings[new_pos], siblings[pos]
        others = [i for i in sub_df.index if i not in siblings]
        first  = siblings[0] if siblings else 0
        before = [i for i in others if i < first]
        after  = [i for i in others if i > max(siblings)]
        self.manager.sub_tables[full] = sub_df.loc[before + siblings + after].reset_index(drop=True)
        self.manager.dirty = True
        self._refresh_sub_tables(revalidate=True)

    def copy_sub_item(self):
        panel = self._current_sub_panel()
        if panel is None: return
        tab_name = self._sub_tabs.tabText(self._sub_tabs.currentIndex())
        full     = self.table_name + "." + tab_name
        sub_df   = self.manager.sub_tables.get(full)
        if sub_df is None: return
        df_idx = panel.selected_df_index()
        if df_idx is None: return
        new_row = sub_df.loc[df_idx].copy()
        self.manager.sub_tables[full] = pd.concat([sub_df, pd.DataFrame([new_row])], ignore_index=True)
        self.manager.dirty = True
        self._refresh_sub_tables(revalidate=True)
        self._refresh_validation_visuals()

    # ── Cross-item row copy / paste ─────────────────────────────────────────────
    def copy_sub_rows(self):
        """Copy the selected sub-table row(s) to a shared buffer (FK excluded),
        so they can be pasted into another master item's (or sheet's) sub-table."""
        panel = self._current_sub_panel()
        if panel is None: return
        tab_name, full = self._current_sub_full()
        if full is None: return
        sub_df = self.manager.sub_tables.get(full)
        if sub_df is None: return
        idxs = panel.selected_df_indices()
        if not idxs:
            self.status_message.emit("請先選取要複製的子表列", _C["yellow"]); return
        fk_key = self.cfg.get("sub_tables", {}).get(tab_name, {}).get("foreign_key", self.pk_key)
        rows = []
        for di in idxs:
            if di not in sub_df.index: continue
            row = {c: sub_df.at[di, c] for c in sub_df.columns if c != fk_key}
            rows.append(row)
        TableEditor._sub_clipboard = {"sub": tab_name, "rows": rows}
        self.status_message.emit(f"已複製 {len(rows)} 列（子表: {tab_name}）", _C["green"])

    def paste_sub_rows(self):
        """Paste buffered rows into the current sub-table under the current master
        item's FK. Only columns that exist in the target sub-table are filled."""
        if self.current_master_pk is None:
            self.status_message.emit("請先選擇一個母表項目", _C["yellow"]); return
        buf = TableEditor._sub_clipboard
        if not buf or not buf.get("rows"):
            self.status_message.emit("列剪貼簿是空的", _C["yellow"]); return
        tab_name, full = self._current_sub_full()
        if full is None: return
        sub_df = self.manager.sub_tables.get(full)
        if sub_df is None: return
        fk_key = self.cfg.get("sub_tables", {}).get(tab_name, {}).get("foreign_key", self.pk_key)
        cols = list(sub_df.columns)
        new_rows = []
        for src in buf["rows"]:
            nr = {c: "" for c in cols}
            nr[fk_key] = self.current_master_pk
            for c in cols:
                if c != fk_key and c in src:
                    nr[c] = src[c]
            new_rows.append(nr)
        if not new_rows: return
        siblings = sub_df[sub_df[fk_key].astype(str) == str(self.current_master_pk)]
        insert_at = siblings.index.max() + 1 if not siblings.empty else len(sub_df)
        top, bot = sub_df.iloc[:insert_at], sub_df.iloc[insert_at:]
        self.manager.sub_tables[full] = pd.concat(
            [top, pd.DataFrame(new_rows), bot], ignore_index=True)
        self.manager.dirty = True
        self._refresh_sub_tables(revalidate=True)
        self._refresh_validation_visuals()
        note = "" if buf["sub"] == tab_name else f"（來源子表: {buf['sub']}，只貼同名欄位）"
        self.status_message.emit(f"已貼上 {len(new_rows)} 列{note}", _C["green"])

    def _resolve_textref(self, col, raw, cols_cfg=None):
        """Resolve a text_ref column's value via that column's own text_ref source."""
        cols_cfg = cols_cfg if cols_cfg is not None else self.cfg.get("columns", {})
        cc = cols_cfg.get(col, {})
        if not isinstance(cc, dict) or cc.get("type") != "text_ref":
            return str(raw)
        src = cc.get("text_ref", {})
        jp = (src.get("json_path") or "").strip()
        if not jp or not self.manager.json_path:
            return str(raw)
        jd = os.path.dirname(self.manager.json_path)
        ab = jp if os.path.isabs(jp) else os.path.join(jd, jp)
        r = self.manager.get_ref_text(ab, src.get("key_col", "TextID") or "TextID",
                                      str(raw), src.get("val_col", "TextContent") or "TextContent")
        return r if r else str(raw)

    def _make_name_resolver(self):
        """Return fk_value -> readable name (via master 'Name' column + its text_ref)."""
        df = self.df; pk = self.pk_key
        namecol = "Name" if "Name" in df.columns else None
        raw_map = {}
        if namecol and pk in df.columns:
            for _, row in df.iterrows():
                raw_map[str(row[pk])] = str(row[namecol])
        def resolve(fk):
            raw = raw_map.get(str(fk), "")
            if namecol and raw:
                return self._resolve_textref(namecol, raw)
            return raw or str(fk)
        return resolve

    def compare_sub_effect(self, df_idx):
        tab_name, full = self._current_sub_full()
        if full is None: return
        sub_df = self.manager.sub_tables.get(full)
        if sub_df is None or df_idx not in sub_df.index: return
        seed = {c: sub_df.at[df_idx, c] for c in sub_df.columns}
        fk_key = self.cfg.get("sub_tables", {}).get(tab_name, {}).get("foreign_key", self.pk_key)
        dlg = EffectCompareDialog(sub_df, fk_key, seed, self._make_name_resolver(), tab_name, self)
        dlg.exec()

    # ── Context menus ─────────────────────────────────────────────────────────

    def _cls_ctx_menu(self, pos):
        item = self._cls_list.itemAt(pos)
        if item is None: return
        g = item.data(Qt.UserRole)
        menu = QMenu(self)
        menu.addAction("重新命名", lambda: self._rename_cls(g))
        menu.addAction("刪除此分類", self.delete_classification)
        menu.exec(self._cls_list.mapToGlobal(pos))

    def _rename_cls(self, old_val):
        new_val, ok = QInputDialog.getText(self, "重新命名", "新名稱:", text=str(old_val))
        if not ok or not new_val.strip() or new_val.strip() == str(old_val): return
        self.df[self.cls_key] = self.df[self.cls_key].replace(old_val, new_val.strip())
        self.manager.tables[self.table_name] = self.df
        self.manager.dirty = True
        if str(self.current_cls_val) == str(old_val):
            self.current_cls_val = new_val.strip()
        self._reload_all(select_cls=self.current_cls_val)

    def _item_ctx_menu(self, pos):
        item = self._card_list.itemAt(pos)
        if item is None: return
        menu = QMenu(self)
        menu.addAction("複製", self.copy_master_item)
        menu.addAction("刪除", self.delete_master_item)
        df_idx = item.data(Qt.UserRole)
        if df_idx in self.df.index and self.pk_key in self.df.columns:
            pk_val = str(self.df.at[df_idx, self.pk_key])
            menu.addSeparator()
            tpl = next((t for t in self._templates()
                        if str(t.get("pk")) == pk_val), None)
            if tpl is None:
                menu.addAction("⭐ 設為範本…", lambda: self.set_template(df_idx))
            else:
                menu.addAction(f"☆ 移除範本（{tpl.get('name', pk_val)}）",
                               lambda: self.remove_template(pk_val))
        sel = len(self._card_list.selectedItems())
        if sel >= 1:
            menu.addSeparator()
            menu.addAction(f"加入比較（{sel} 項）", self.add_to_compare)
        menu.exec(self._card_list.mapToGlobal(pos))

    # ── Templates（範本列：具名原型，從範本新增＝深拷貝母列＋子表列） ──────────

    def _templates(self):
        """Mutable template list stored in this table's config:
        [{"name": 顯示名, "pk": 範本列的主鍵值}, …]"""
        tpls = self.cfg.get("templates")
        if not isinstance(tpls, list):
            tpls = []
            self.cfg["templates"] = tpls
        return tpls

    def _template_pks(self):
        return {str(t.get("pk")) for t in self._templates()}

    def set_template(self, df_idx):
        if df_idx not in self.df.index:
            return
        pk_val = str(self.df.at[df_idx, self.pk_key])
        name, ok = QInputDialog.getText(
            self, "設為範本", "範本名稱：", text=pk_val)
        if not ok or not name.strip():
            return
        self._templates().append({"name": name.strip(), "pk": pk_val})
        self.manager.save_config()
        self._load_item_list()
        self.status_message.emit(f"⭐ 已設為範本：{name.strip()}", _C["green"])

    def remove_template(self, pk_val):
        tpls = self._templates()
        tpls[:] = [t for t in tpls if str(t.get("pk")) != str(pk_val)]
        self.manager.save_config()
        self._load_item_list()
        self.status_message.emit("已移除範本", _C["yellow"])

    def _template_menu(self):
        menu = QMenu(self)
        tpls = self._templates()
        for t in tpls:
            exists = str(t.get("pk")) in self.df[self.pk_key].astype(str).values \
                if self.pk_key in self.df.columns else False
            act = menu.addAction(f"⭐ {t.get('name', t.get('pk'))}")
            if exists:
                act.triggered.connect(lambda _=False, tt=t: self.add_from_template(tt))
            else:
                act.setText(act.text() + "（來源已不存在）")
                act.setEnabled(False)
        if not tpls:
            hint = menu.addAction("（尚無範本 — 在項目上按右鍵「設為範本」）")
            hint.setEnabled(False)
        menu.exec(QCursor.pos())

    def add_from_template(self, tpl):
        src = self.df[self.df[self.pk_key].astype(str) == str(tpl.get("pk"))]
        if src.empty:
            QMessageBox.warning(self, "提示", "範本來源列已不存在（ID 可能被改掉了）")
            return
        new_id, ok = QInputDialog.getText(
            self, "從範本新增",
            f"範本：{tpl.get('name')}\n新的 {self.pk_key}:")
        if not ok or not new_id.strip():
            return
        new_id = new_id.strip()
        if new_id in self.df[self.pk_key].astype(str).values:
            QMessageBox.warning(self, "錯誤", "此 ID 已存在"); return
        self._create_from_template(tpl, new_id)

    def _create_from_template(self, tpl, new_id):
        """深拷貝範本母列＋其所有子表列（FK 改成 new_id），並選中新項目。"""
        src = self.df[self.df[self.pk_key].astype(str) == str(tpl.get("pk"))]
        if src.empty:
            return
        src_idx = src.index[0]
        new_row = self.df.loc[src_idx].copy()
        new_row[self.pk_key] = new_id
        cls_val = new_row[self.cls_key] if self.cls_key in self.df.columns else None

        # 插在範本同分類的最後一列後面（跟「新增項目」同邏輯）
        if cls_val is not None:
            cls_rows = self.df[self.df[self.cls_key] == cls_val]
            insert_after = cls_rows.index.max() + 1 if not cls_rows.empty else len(self.df)
        else:
            insert_after = len(self.df)
        top, bot = self.df.iloc[:insert_after], self.df.iloc[insert_after:]
        self.df = pd.concat([top, pd.DataFrame([new_row]), bot], ignore_index=True)
        self.manager.tables[self.table_name] = self.df

        # 子表列跟著深拷貝（FK 換成新 ID）
        copied_subs = 0
        prefix = self.table_name + "."
        for full in list(self.manager.sub_tables):
            if not full.startswith(prefix):
                continue
            sub_name = full[len(prefix):]
            sub_df = self.manager.sub_tables[full]
            fk = (self.cfg.get("sub_tables", {}).get(sub_name, {})
                  .get("foreign_key") or self.pk_key)
            if fk not in sub_df.columns:
                continue
            rows = sub_df[sub_df[fk].astype(str) == str(tpl.get("pk"))]
            if rows.empty:
                continue
            copies = rows.copy()
            copies[fk] = new_id
            self.manager.sub_tables[full] = pd.concat(
                [sub_df, copies], ignore_index=True)
            copied_subs += len(copies)

        self.manager.dirty = True
        new_df_idx = self.df[self.df[self.pk_key].astype(str) == str(new_id)].index[0]
        self._reload_all(select_cls=cls_val if cls_val is not None else self.current_cls_val,
                         select_idx=new_df_idx)
        self.status_message.emit(
            f"⭐ 已從範本「{tpl.get('name')}」建立 {new_id}"
            + (f"（含 {copied_subs} 列子表）" if copied_subs else ""), _C["green"])

    # ── Refresh ───────────────────────────────────────────────────────────────

    def _reload_all(self, select_cls=None, select_idx=None):
        # master mutations (add/copy/delete/paste/column ops) reset row indexes
        # → the validator's violation map must be rebuilt for this table
        self.manager.validator.validate_table(self.table_name)
        self.df = self.manager.tables[self.table_name]
        if select_cls is not None:  self.current_cls_val    = select_cls
        if select_idx is not None:  self.current_master_idx = select_idx
        self._load_cls_list()
        if self.current_cls_val is not None:
            self._load_item_list()
        if self.current_master_idx is not None and self.current_master_idx in self.df.index:
            self._load_editor(self.current_master_idx)

    def add_master_column(self):
        col_name, ok = QInputDialog.getText(self, "新增欄位", "欄位名稱:")
        if not ok or not col_name.strip(): return
        col_name = col_name.strip()
        if col_name in self.df.columns:
            QMessageBox.warning(self, "錯誤", f"欄位 [{col_name}] 已存在"); return
        self.df[col_name] = ""
        self.manager.tables[self.table_name] = self.df
        self.manager.dirty = True
        if self._field_panel:
            self._field_panel.deleteLater()
            self._field_panel = None
        self._reload_all(select_cls=self.current_cls_val, select_idx=self.current_master_idx)
        self.status_message.emit(f"欄位 [{col_name}] 已新增", _C["green"])

    def add_sub_column(self):
        tab_idx = self._sub_tabs.currentIndex()
        if tab_idx < 0: return
        tab_name = self._sub_tabs.tabText(tab_idx)
        if not tab_name or tab_name == "—": return
        full   = self.table_name + "." + tab_name
        sub_df = self.manager.sub_tables.get(full)
        if sub_df is None: return
        col_name, ok = QInputDialog.getText(self, "新增欄位", f"欄位名稱（子表: {tab_name}）:")
        if not ok or not col_name.strip(): return
        col_name = col_name.strip()
        if col_name in sub_df.columns:
            QMessageBox.warning(self, "錯誤", f"欄位 [{col_name}] 已存在"); return
        sub_df[col_name] = ""
        self.manager.sub_tables[full] = sub_df
        self.manager.dirty = True
        self._build_sub_tabs()
        self._select_sub_tab(tab_name)
        self._refresh_sub_tables()
        self.status_message.emit(f"從表欄位 [{col_name}] 已新增", _C["green"])

    def add_sub_table(self):
        if self.table_name not in self.manager.tables:
            return
        if not self.pk_key:
            QMessageBox.warning(self, "提示", "母表沒有主鍵，無法建立子表"); return
        name, ok = QInputDialog.getText(self, "新增子表", "子表名稱（巢狀陣列欄位名）:")
        if not ok or not name.strip(): return
        name = name.strip()
        full = self.table_name + "." + name
        if full in self.manager.sub_tables:
            QMessageBox.warning(self, "錯誤", f"子表 [{name}] 已存在"); return
        if name in self.df.columns:
            QMessageBox.warning(self, "錯誤", f"名稱 [{name}] 與母表欄位重複"); return
        if not self.manager.add_sub_table(self.table_name, name):
            QMessageBox.warning(self, "錯誤", "建立子表失敗"); return
        self._build_sub_tabs()
        self._select_sub_tab(name)
        self._refresh_sub_tables()
        self.status_message.emit(f"子表 [{name}]（FK={self.pk_key}）已新增", _C["green"])

    def delete_sub_table(self):
        tab_name, full = self._current_sub_full()
        if full is None:
            QMessageBox.information(self, "提示", "目前沒有可刪除的子表"); return
        if QMessageBox.question(
                self, "確認刪除",
                f"確定刪除整張子表 [{tab_name}]？\n會移除其所有資料列與欄位定義。") != QMessageBox.Yes:
            return
        self.manager.delete_sub_table(self.table_name, tab_name)
        self._build_sub_tabs()
        self._refresh_sub_tables()
        self.status_message.emit(f"子表 [{tab_name}] 已刪除", _C["yellow"])

    # ── Column rename / delete (master + sub) ───────────────────────────────────

    def _select_sub_tab(self, name):
        for i in range(self._sub_tabs.count()):
            if self._sub_tabs.tabText(i) == name:
                self._sub_tabs.setCurrentIndex(i); return

    def focus_master_field(self, col):
        """Scroll/flash a master field — used by search & compare quick-nav."""
        if col and self._field_panel is not None:
            self._field_panel.focus_field(col)

    def focus_sub(self, tab_name, df_idx, col=None):
        """Open a sub-table tab and select the given sub-row (quick-nav)."""
        self._select_sub_tab(tab_name)
        panel = self._sub_panels.get(tab_name)
        if panel is not None:
            panel.select_by_df_index(df_idx, col)

    def refresh_ref_display(self):
        """Re-resolve external text_ref display (field labels + compare view)
        after a referenced text table changed. Item-list names refresh on the
        next reselect."""
        if self._field_panel is not None:
            self._field_panel.refresh_ref_labels()
        cv = getattr(self, "_compare_view", None)
        if cv is not None:
            try:
                cv.refresh()
            except Exception:
                pass

    def _fit_right_panel(self):
        """Make the right panel's min width track what the sub-table toolbar
        actually needs, so its buttons never get squeezed (hardcoded → adaptive)."""
        if not hasattr(self, "_sub_hdr") or not hasattr(self, "_right_panel"):
            return
        self._sub_hdr.layout().activate()
        need = self._sub_hdr.sizeHint().width() + 24
        self._right_panel.setMinimumWidth(max(320, need))

    def _on_sub_picker(self, idx):
        name = self._sub_picker.itemText(idx)
        if name:
            self._select_sub_tab(name)

    def _sync_sub_picker(self, idx):
        if not hasattr(self, "_sub_picker"):
            return
        name = self._sub_tabs.tabText(idx) if idx >= 0 else ""
        pi = self._sub_picker.findText(name)
        if pi >= 0 and pi != self._sub_picker.currentIndex():
            self._sub_picker.blockSignals(True)
            self._sub_picker.setCurrentIndex(pi)
            self._sub_picker.blockSignals(False)

    def _current_sub_full(self):
        idx = self._sub_tabs.currentIndex()
        if idx < 0:
            return None, None
        tab_name = self._sub_tabs.tabText(idx)
        if not tab_name or tab_name == "—":
            return None, None
        return tab_name, self.table_name + "." + tab_name

    def _pick_column(self, columns, exclude, title):
        choices = [c for c in columns if c not in exclude]
        if not choices:
            QMessageBox.information(self, "提示", "沒有可操作的欄位"); return None
        col, ok = QInputDialog.getItem(self, title, "選擇欄位:", choices, 0, False)
        if not ok or not col:
            return None
        return col

    def rename_master_column(self):
        col = self._pick_column(list(self.df.columns), {self.pk_key, self.cls_key}, "重新命名欄位")
        if col is None: return
        new, ok = QInputDialog.getText(self, "重新命名", f"新欄位名稱（{col} →）:")
        if not ok or not new.strip(): return
        new = new.strip()
        if new in self.df.columns:
            QMessageBox.warning(self, "錯誤", f"欄位 [{new}] 已存在"); return
        self.manager.rename_column(self.table_name, col, new)
        self.df = self.manager.tables[self.table_name]
        if self._field_panel:
            self._field_panel.deleteLater(); self._field_panel = None
        self._reload_all(select_cls=self.current_cls_val, select_idx=self.current_master_idx)
        self.status_message.emit(f"欄位 [{col}] → [{new}]", _C["green"])

    def delete_master_column(self):
        col = self._pick_column(list(self.df.columns), {self.pk_key, self.cls_key}, "刪除欄位")
        if col is None: return
        if QMessageBox.question(self, "確認刪除", f"確定刪除欄位 [{col}]？") != QMessageBox.Yes:
            return
        self.manager.delete_column(self.table_name, col)
        self.df = self.manager.tables[self.table_name]
        if self._field_panel:
            self._field_panel.deleteLater(); self._field_panel = None
        self._reload_all(select_cls=self.current_cls_val, select_idx=self.current_master_idx)
        self.status_message.emit(f"欄位 [{col}] 已刪除", _C["yellow"])

    def rename_sub_column(self):
        tab_name, full = self._current_sub_full()
        if full is None: return
        sub_df = self.manager.sub_tables.get(full)
        if sub_df is None: return
        fk = self.cfg.get("sub_tables", {}).get(tab_name, {}).get("foreign_key", self.pk_key)
        col = self._pick_column(list(sub_df.columns), {fk}, f"重新命名欄位（子表: {tab_name}）")
        if col is None: return
        new, ok = QInputDialog.getText(self, "重新命名", f"新欄位名稱（{col} →）:")
        if not ok or not new.strip(): return
        new = new.strip()
        if new in sub_df.columns:
            QMessageBox.warning(self, "錯誤", f"欄位 [{new}] 已存在"); return
        self.manager.rename_column(full, col, new)
        self._build_sub_tabs(); self._select_sub_tab(tab_name); self._refresh_sub_tables()
        self.status_message.emit(f"子表欄位 [{col}] → [{new}]", _C["green"])

    def delete_sub_column(self):
        tab_name, full = self._current_sub_full()
        if full is None: return
        sub_df = self.manager.sub_tables.get(full)
        if sub_df is None: return
        fk = self.cfg.get("sub_tables", {}).get(tab_name, {}).get("foreign_key", self.pk_key)
        col = self._pick_column(list(sub_df.columns), {fk}, f"刪除欄位（子表: {tab_name}）")
        if col is None: return
        if QMessageBox.question(self, "確認刪除", f"確定刪除子表欄位 [{col}]？") != QMessageBox.Yes:
            return
        self.manager.delete_column(full, col)
        self._build_sub_tabs(); self._select_sub_tab(tab_name); self._refresh_sub_tables()
        self.status_message.emit(f"子表欄位 [{col}] 已刪除", _C["yellow"])

    def reload_after_config(self):
        self.cfg     = self.manager.config.get(self.table_name, {})
        cols         = list(self.df.columns)
        self.cls_key = self.cfg.get("classification_key", "") or (cols[0] if cols else "")
        self.pk_key  = self.cfg.get("primary_key",        "") or (cols[0] if cols else "")
        if self._field_panel:
            self._field_panel.deleteLater()
            self._field_panel = None
        self._build_sub_tabs()
        self._reload_all()


# ── WelcomeWidget ─────────────────────────────────────────────────────────────

class WelcomeWidget(QWidget):
    open_file   = Signal()
    new_file    = Signal()
    open_recent = Signal(str)

    def __init__(self, manager: JsonDataManager, parent=None):
        super().__init__(parent)
        self._manager = manager
        self.setStyleSheet(f"background: {_C['bg']};")
        self._setup_ui()

    def _setup_ui(self):
        outer = QVBoxLayout(self)
        outer.setAlignment(Qt.AlignCenter)

        card = QWidget()
        card.setFixedWidth(440)
        card.setStyleSheet(
            f"background:{_C['card']}; border-radius:12px; "
            f"border:1px solid {_C['border']};"
        )
        lo = QVBoxLayout(card)
        lo.setContentsMargins(40, 36, 40, 36)
        lo.setSpacing(0)

        # Logo
        logo = QLabel("{ }")
        logo.setAlignment(Qt.AlignCenter)
        logo.setFixedSize(60, 60)
        logo.setStyleSheet(
            f"font-family:Consolas; font-size:22px; font-weight:bold; color:white; "
            f"background:qlineargradient(x1:0,y1:0,x2:1,y2:1,"
            f"stop:0 {_C['accent']},stop:1 #8B5CF6); border-radius:12px; border:none;"
        )
        logo_wrap = QHBoxLayout()
        logo_wrap.addStretch()
        logo_wrap.addWidget(logo)
        logo_wrap.addStretch()
        lo.addLayout(logo_wrap)
        lo.addSpacing(16)

        title = QLabel("JsonEditor")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet(
            f"font-size:20px; font-weight:700; color:{_C['txt']}; background:transparent; border:none;"
        )
        lo.addWidget(title)

        subtitle = QLabel("輕量 JSON 資料編輯器")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet(
            f"color:{_C['txt2']}; font-size:12px; background:transparent; border:none;"
        )
        lo.addWidget(subtitle)
        lo.addSpacing(28)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)
        b_open = _mk_btn("📂  開啟 JSON", "primary")
        b_open.setFixedHeight(38)
        b_open.clicked.connect(self.open_file)
        b_new = _mk_btn("📄  新建 JSON")
        b_new.setFixedHeight(38)
        b_new.clicked.connect(self.new_file)
        btn_row.addWidget(b_open, 1)
        btn_row.addWidget(b_new, 1)
        lo.addLayout(btn_row)

        recent = self._manager._recent_files
        if recent:
            lo.addSpacing(20)
            sep = QFrame()
            sep.setFixedHeight(1)
            sep.setStyleSheet(f"background:{_C['border']}; border:none;")
            lo.addWidget(sep)
            lo.addSpacing(12)

            hdr = QLabel("最近開啟")
            hdr.setStyleSheet(
                f"color:{_C['txt3']}; font-size:10px; font-weight:600; "
                f"letter-spacing:1px; background:transparent; border:none;"
            )
            lo.addWidget(hdr)
            lo.addSpacing(6)

            for path in recent[:6]:
                fname = os.path.basename(path)
                dirn  = os.path.dirname(path)
                row   = QWidget()
                row.setStyleSheet(
                    f"QWidget {{ background:transparent; border-radius:6px; border:none; }}"
                    f"QWidget:hover {{ background:{_C['cardH']}; }}"
                )
                row.setCursor(Qt.PointingHandCursor)
                rlo = QHBoxLayout(row)
                rlo.setContentsMargins(8, 6, 8, 6)
                lbl_fn  = QLabel(fname)
                lbl_fn.setStyleSheet(
                    f"color:{_C['txtAcc']}; font-weight:600; background:transparent; border:none;"
                )
                lbl_dir = QLabel(dirn)
                lbl_dir.setStyleSheet(
                    f"color:{_C['txt3']}; font-size:10px; background:transparent; border:none;"
                )
                lbl_dir.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
                rlo.addWidget(lbl_fn)
                rlo.addWidget(lbl_dir, 1)
                row.mousePressEvent = lambda e, p=path: self.open_recent.emit(p)
                lo.addWidget(row)

        outer.addWidget(card)

    def refresh(self, manager):
        self._manager = manager
        while self.layout().count():
            item = self.layout().takeAt(0)
            if item.widget(): item.widget().deleteLater()
        self._setup_ui()


# ── Notes (free-form text / table pages, saved into config with the DB) ────────

class NotePage(QWidget):
    """A single note page. Three kinds:
      • text  — free-form multi-line text
      • table — editable grid (Excel range copy/paste supported)
      • group — holds a one-level QTabWidget of child text/table pages
    Content is flushed into config and persisted by the normal 儲存 action."""
    changed = Signal()

    def __init__(self, note, parent=None):
        super().__init__(parent)
        self._note_id = note.get("id", "")
        self._type    = note.get("type", "text")
        v = QVBoxLayout(self)
        v.setContentsMargins(10, 8, 10, 10); v.setSpacing(8)

        if self._type == "group":
            bar = QHBoxLayout(); bar.setSpacing(6)
            b = _mk_btn("＋ 小分頁", "ghost"); b.setFixedHeight(26)
            b.clicked.connect(self._add_child)
            bar.addWidget(b); bar.addStretch(1)
            tip = QLabel("右鍵小分頁標題可改名／刪除")
            tip.setStyleSheet(f"color:{_C['txt3']}; font-size:10px; background:transparent;")
            bar.addWidget(tip)
            v.addLayout(bar)
            self._inner = QTabWidget(); self._inner.setObjectName("sub-tabs")
            self._inner.setDocumentMode(True)
            self._inner.tabBar().setContextMenuPolicy(Qt.CustomContextMenu)
            self._inner.tabBar().customContextMenuRequested.connect(self._inner_menu)
            v.addWidget(self._inner, 1)
            for child in (note.get("children") or []):
                self._add_child_page(child)

        elif self._type == "table":
            bar = QHBoxLayout(); bar.setSpacing(6)
            for txt, fn in (("＋ 列", self._add_row), ("－ 列", self._del_row),
                            ("＋ 欄", self._add_col), ("－ 欄", self._del_col)):
                b = _mk_btn(txt, "ghost"); b.setFixedHeight(26); b.clicked.connect(fn)
                bar.addWidget(b)
            bar.addStretch(1)
            tip = QLabel("雙擊儲存格編輯；雙擊欄標題改名；可從 Excel 複製範圍 Ctrl+V 貼上")
            tip.setStyleSheet(f"color:{_C['txt3']}; font-size:10px; background:transparent;")
            bar.addWidget(tip)
            v.addLayout(bar)

            content = note.get("content") or [["", ""], ["", ""]]
            ncols   = max((len(r) for r in content), default=2) if content else 2
            headers = note.get("headers") or [f"欄{i+1}" for i in range(ncols)]
            if len(headers) < ncols:
                headers += [f"欄{i+1}" for i in range(len(headers), ncols)]
            self._tbl = QTableWidget()
            self._tbl.setColumnCount(len(headers))
            self._tbl.setHorizontalHeaderLabels(headers)
            self._tbl.setRowCount(len(content))
            for r, rowv in enumerate(content):
                for c in range(len(headers)):
                    val = rowv[c] if c < len(rowv) else ""
                    self._tbl.setItem(r, c, QTableWidgetItem(str(val)))
            self._tbl.verticalHeader().setDefaultSectionSize(28)
            self._tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            self._tbl.horizontalHeader().setStretchLastSection(True)
            self._tbl.setSelectionMode(QAbstractItemView.ExtendedSelection)
            self._tbl.keyPressEvent = self._table_key   # Excel-style copy/paste
            v.addWidget(self._tbl, 1)
            # connect AFTER populating so initial fill doesn't mark dirty
            self._tbl.itemChanged.connect(lambda *_: self.changed.emit())
            self._tbl.horizontalHeader().sectionDoubleClicked.connect(self._rename_col)

        else:
            self._edit = QPlainTextEdit()
            self._edit.setPlainText(note.get("content", "") or "")
            self._edit.setStyleSheet(
                f"background:{_C['code']}; color:{_C['txt']}; "
                f"border:1px solid {_C['border']}; border-radius:6px; "
                f"padding:8px; font-size:13px;")
            v.addWidget(self._edit, 1)
            self._edit.textChanged.connect(self.changed.emit)

    # ── table operations ──
    def _add_row(self):
        self._tbl.insertRow(self._tbl.rowCount()); self.changed.emit()

    def _del_row(self):
        r = self._tbl.currentRow()
        if r < 0:
            r = self._tbl.rowCount() - 1
        if r >= 0:
            self._tbl.removeRow(r); self.changed.emit()

    def _add_col(self):
        c = self._tbl.columnCount()
        self._tbl.insertColumn(c)
        self._tbl.setHorizontalHeaderItem(c, QTableWidgetItem(f"欄{c + 1}"))
        self.changed.emit()

    def _del_col(self):
        c = self._tbl.currentColumn()
        if c < 0:
            c = self._tbl.columnCount() - 1
        if c >= 0:
            self._tbl.removeColumn(c); self.changed.emit()

    def _rename_col(self, c):
        cur = self._tbl.horizontalHeaderItem(c)
        old = cur.text() if cur else f"欄{c + 1}"
        name, ok = QInputDialog.getText(self, "欄位改名", "名稱:", text=old)
        if ok and name.strip():
            self._tbl.setHorizontalHeaderItem(c, QTableWidgetItem(name.strip()))
            self.changed.emit()

    def _table_key(self, event):
        if event.matches(QKeySequence.Paste):
            self._paste_clipboard()
        elif event.matches(QKeySequence.Copy):
            self._copy_selection()
        else:
            QTableWidget.keyPressEvent(self._tbl, event)

    def _paste_clipboard(self):
        """Paste an Excel range (TSV: tabs = columns, newlines = rows), auto-
        expanding the grid so any N×M selection fits from the current cell."""
        text = QApplication.clipboard().text()
        if not text:
            return
        grid = [row.split("\t") for row in text.replace("\r\n", "\n").rstrip("\n").split("\n")]
        if not grid:
            return
        r0 = max(0, self._tbl.currentRow())
        c0 = max(0, self._tbl.currentColumn())
        need_r = r0 + len(grid)
        need_c = c0 + max(len(row) for row in grid)
        if self._tbl.rowCount() < need_r:
            self._tbl.setRowCount(need_r)
        if self._tbl.columnCount() < need_c:
            old = self._tbl.columnCount()
            self._tbl.setColumnCount(need_c)
            for c in range(old, need_c):
                self._tbl.setHorizontalHeaderItem(c, QTableWidgetItem(f"欄{c + 1}"))
        self._tbl.blockSignals(True)
        for dr, row in enumerate(grid):
            for dc, val in enumerate(row):
                self._tbl.setItem(r0 + dr, c0 + dc, QTableWidgetItem(val))
        self._tbl.blockSignals(False)
        self.changed.emit()

    def _copy_selection(self):
        """Copy the selected rectangle back out as Excel-compatible TSV."""
        ranges = self._tbl.selectedRanges()
        if not ranges:
            return
        rng = ranges[0]
        lines = []
        for r in range(rng.topRow(), rng.bottomRow() + 1):
            cells = []
            for c in range(rng.leftColumn(), rng.rightColumn() + 1):
                it = self._tbl.item(r, c)
                cells.append(it.text() if it else "")
            lines.append("\t".join(cells))
        QApplication.clipboard().setText("\n".join(lines))

    # ── group (inner sub-tab) operations ──
    def _add_child_page(self, child):
        page = NotePage(child)
        page.changed.connect(self.changed.emit)
        self._inner.addTab(page, child.get("name") or "小分頁")

    def _add_child(self):
        name, ok = QInputDialog.getText(self, "新增小分頁", "名稱:")
        if not ok or not name.strip():
            return
        kind, ok2 = QInputDialog.getItem(
            self, "新增小分頁", "型態:", ["純文字", "表格"], 0, False)
        if not ok2:
            return
        child = ({"name": name.strip(), "type": "text", "content": ""}
                 if kind == "純文字" else
                 {"name": name.strip(), "type": "table",
                  "headers": ["欄1", "欄2"], "content": [["", ""], ["", ""]]})
        self._add_child_page(child)
        self._inner.setCurrentIndex(self._inner.count() - 1)
        self.changed.emit()

    def _inner_menu(self, pos):
        bar = self._inner.tabBar(); idx = bar.tabAt(pos)
        if idx < 0:
            return
        m = QMenu(self)
        m.addAction("重新命名", lambda: self._rename_child(idx))
        m.addAction("刪除", lambda: self._del_child(idx))
        m.exec(bar.mapToGlobal(pos))

    def _rename_child(self, idx):
        name, ok = QInputDialog.getText(
            self, "重新命名", "名稱:", text=self._inner.tabText(idx))
        if ok and name.strip():
            self._inner.setTabText(idx, name.strip()); self.changed.emit()

    def _del_child(self, idx):
        w = self._inner.widget(idx)
        self._inner.removeTab(idx)
        if w is not None:
            w.deleteLater()
        self.changed.emit()

    def to_data(self):
        """Serialize back to the config note dict (type-specific payload)."""
        if self._type == "group":
            children = []
            for i in range(self._inner.count()):
                w = self._inner.widget(i)
                if isinstance(w, NotePage):
                    d = w.to_data()
                    d["name"] = self._inner.tabText(i)
                    children.append(d)
            return {"type": "group", "children": children}
        if self._type == "table":
            headers = [
                (self._tbl.horizontalHeaderItem(c).text()
                 if self._tbl.horizontalHeaderItem(c) else f"欄{c + 1}")
                for c in range(self._tbl.columnCount())
            ]
            content = []
            for r in range(self._tbl.rowCount()):
                content.append([
                    (self._tbl.item(r, c).text() if self._tbl.item(r, c) else "")
                    for c in range(self._tbl.columnCount())
                ])
            return {"type": "table", "headers": headers, "content": content}
        return {"type": "text", "content": self._edit.toPlainText()}


# ── Global search (results window with quick-nav into the editor) ──────────────

class GlobalSearchDialog(QDialog):
    """Non-modal global search: type a keyword, list every matching record across
    all tables (master + sub); double-click a row to jump the main editor to that
    record's matching field."""

    def __init__(self, app):
        super().__init__(app)
        self._app = app
        self.setWindowTitle("全域搜尋")
        self.setStyleSheet(APP_QSS)
        self.setWindowFlag(Qt.Window, True)
        self.resize(760, 540)
        self._rows = []  # [(table_name, is_sub, row_idx, col)]

        v = QVBoxLayout(self); v.setContentsMargins(12, 12, 12, 12); v.setSpacing(8)
        row = QHBoxLayout(); row.setSpacing(6)
        self._input = QLineEdit()
        self._input.setPlaceholderText("輸入關鍵字，搜尋所有資料表…")
        self._input.returnPressed.connect(self._run)
        btn = _mk_btn("搜尋", "primary"); btn.setFixedHeight(30); btn.clicked.connect(self._run)
        row.addWidget(self._input, 1); row.addWidget(btn)
        v.addLayout(row)

        self._info = QLabel("輸入關鍵字後按 Enter")
        self._info.setStyleSheet(f"color:{_C['txt2']}; font-size:11px; background:transparent;")
        v.addWidget(self._info)

        self._tbl = QTableWidget()
        self._tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self._tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self._tbl.setColumnCount(4)
        self._tbl.setHorizontalHeaderLabels(["資料表", "來源", "命中欄位", "命中內容"])
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.doubleClicked.connect(self._open)
        v.addWidget(self._tbl, 1)

        hint = QLabel("雙擊任一列 → 跳到主編輯區的該筆該欄位")
        hint.setStyleSheet(f"color:{_C['txt3']}; font-size:10px; background:transparent;")
        v.addWidget(hint)

    def focus_input(self):
        self._input.setFocus(); self._input.selectAll()

    def _run(self):
        q = self._input.text().strip()
        self._tbl.setRowCount(0); self._rows = []
        if not q:
            self._info.setText("請輸入關鍵字"); return
        results = self._app.manager.search_index(q)
        ql = q.lower()
        display = []
        for table_name, is_sub, row_idx, matched in results:
            for col, val in matched.items():
                display.append((table_name, is_sub, row_idx, col, val))
        self._tbl.setRowCount(len(display))
        for r, (tn, is_sub, ridx, col, val) in enumerate(display):
            cells = [tn, self._source_label(tn, is_sub, ridx), col,
                     self._snippet(val, ql)]
            for c, txt in enumerate(cells):
                it = QTableWidgetItem(txt)
                it.setToolTip(str(val) if c == 3 else txt)
                self._tbl.setItem(r, c, it)
            self._rows.append((tn, is_sub, ridx, col))
        self._tbl.resizeColumnsToContents()
        self._tbl.horizontalHeader().setStretchLastSection(True)
        self._info.setText(
            f"找到 {len(display)} 筆命中（{len(results)} 筆資料）" if display else "無結果")

    def _source_label(self, table_name, is_sub, row_idx):
        mgr = self._app.manager
        try:
            if is_sub:
                master, sub = table_name.split(".", 1)
                df = mgr.sub_tables.get(table_name)
                fk = (mgr.config.get(master, {}).get("sub_tables", {})
                      .get(sub, {}).get("foreign_key"))
                if df is not None and fk in (df.columns if df is not None else []) \
                        and row_idx in df.index:
                    return str(df.at[row_idx, fk])
            else:
                df = mgr.tables.get(table_name)
                pk = mgr.config.get(table_name, {}).get("primary_key")
                if df is not None and pk in (df.columns if df is not None else []) \
                        and row_idx in df.index:
                    return str(df.at[row_idx, pk])
        except Exception:
            pass
        return str(row_idx)

    @staticmethod
    def _snippet(val, ql, width=40):
        s = str(val); i = s.lower().find(ql)
        if i < 0:
            return s[:width]
        start = max(0, i - 12); end = min(len(s), i + len(ql) + 24)
        return f"{'…' if start else ''}{s[start:end]}{'…' if end < len(s) else ''}"

    def _open(self, index):
        r = index.row()
        if 0 <= r < len(self._rows):
            tn, is_sub, ridx, col = self._rows[r]
            self._app.navigate_to(tn, is_sub, ridx, col)


# ── Validation rules editor ───────────────────────────────────────────────────

class _VCondRow(QWidget):
    """One editable condition row: 欄位｜運算子｜值(｜值2)，或母表 scope 的
    「子表聚合」變體（子表｜欄位｜運算子｜值｜列數比較）。"""
    removed = Signal(object)

    def __init__(self, dlg, cond=None, allow_agg=False, parent=None):
        super().__init__(parent)
        self._dlg = dlg                     # ValidationRulesDialog (欄位清單來源)
        self._allow_agg = allow_agg
        lo = QHBoxLayout(self)
        lo.setContentsMargins(0, 0, 0, 0)
        lo.setSpacing(6)

        self.kind = _NoscrollCombo()
        self.kind.addItem("欄位")
        if allow_agg:
            self.kind.addItem("子表聚合")
        self.kind.currentIndexChanged.connect(self._kind_changed)
        lo.addWidget(self.kind)
        if not allow_agg:
            self.kind.hide()

        # ── field variant ──
        self.field = _NoscrollCombo(); self.field.setEditable(True)
        self.field.setMinimumWidth(150)
        self.op = _NoscrollCombo()
        for o in _V_OPS:
            self.op.addItem(_V_OP_LABELS[o], o)
        self.op.currentIndexChanged.connect(self._op_changed)
        self.value  = QLineEdit(); self.value.setMinimumWidth(110)
        self.value2 = QLineEdit(); self.value2.setFixedWidth(70)
        self.value2.setPlaceholderText("上限")
        lo.addWidget(self.field); lo.addWidget(self.op)
        lo.addWidget(self.value, 1); lo.addWidget(self.value2)

        # ── agg variant ──
        self.agg_sub = _NoscrollCombo(); self.agg_sub.setMinimumWidth(110)
        self.agg_field = _NoscrollCombo(); self.agg_field.setEditable(True)
        self.agg_field.setMinimumWidth(120)
        self.agg_op = _NoscrollCombo()
        for o in _V_OPS:
            if o not in ("between",):
                self.agg_op.addItem(_V_OP_LABELS[o], o)
        self.agg_val = QLineEdit(); self.agg_val.setMinimumWidth(80)
        self.agg_cnt_lbl = QLabel("的列數")
        self.agg_cnt_lbl.setStyleSheet(f"color:{_C['txt2']}; background:transparent;")
        self.agg_cnt_op = _NoscrollCombo()
        for k, v in _V_COUNT_OPS.items():
            self.agg_cnt_op.addItem(v, k)
        self.agg_cnt = QSpinBox(); self.agg_cnt.setRange(0, 9999); self.agg_cnt.setValue(1)
        self.agg_sub.currentIndexChanged.connect(self._agg_sub_changed)
        for w in (self.agg_sub, self.agg_field, self.agg_op, self.agg_val,
                  self.agg_cnt_lbl, self.agg_cnt_op, self.agg_cnt):
            lo.addWidget(w)

        rm = _mk_btn("", "danger", icon="trash")
        rm.setFixedSize(26, 26)
        rm.setToolTip("刪除此條件")
        rm.clicked.connect(lambda: self.removed.emit(self))
        lo.addWidget(rm)

        self._load(cond or {})
        self._kind_changed()

    # ── UI state ──
    def _kind_changed(self, *_):
        is_agg = self._allow_agg and self.kind.currentIndex() == 1
        for w in (self.field, self.op, self.value, self.value2):
            w.setVisible(not is_agg)
        for w in (self.agg_sub, self.agg_field, self.agg_op, self.agg_val,
                  self.agg_cnt_lbl, self.agg_cnt_op, self.agg_cnt):
            w.setVisible(is_agg)
        if not is_agg:
            self._op_changed()

    def _op_changed(self, *_):
        op = self.op.currentData()
        self.value.setVisible(op not in ("empty", "not_empty"))
        self.value2.setVisible(op == "between")

    def _agg_sub_changed(self, *_):
        cur = self.agg_field.currentText()
        self.agg_field.clear()
        self.agg_field.addItem("")          # 空 = 數全部列
        self.agg_field.addItems(self._dlg.sub_columns(self.agg_sub.currentText()))
        self.agg_field.setCurrentText(cur)

    def set_field_choices(self, fields, subs):
        cur = self.field.currentText()
        self.field.clear(); self.field.addItems(fields)
        self.field.setCurrentText(cur)
        cur_sub = self.agg_sub.currentText()
        self.agg_sub.clear(); self.agg_sub.addItems(subs)
        if cur_sub:
            self.agg_sub.setCurrentText(cur_sub)
        self._agg_sub_changed()

    # ── cond dict ↔ form ──
    def _load(self, cond):
        agg = cond.get("agg")
        if agg and self._allow_agg:
            self.kind.setCurrentIndex(1)
            self.agg_sub.setCurrentText(str(agg.get("sub", "")))
            self.agg_field.setCurrentText(str(agg.get("field", "")))
            i = self.agg_op.findData(agg.get("op", "eq"))
            self.agg_op.setCurrentIndex(max(0, i))
            self.agg_val.setText(str(agg.get("value", "")))
            i = self.agg_cnt_op.findData(agg.get("count_op", "ge"))
            self.agg_cnt_op.setCurrentIndex(max(0, i))
            try:
                self.agg_cnt.setValue(int(float(agg.get("count", 1))))
            except (ValueError, TypeError):
                self.agg_cnt.setValue(1)
        else:
            self.field.setCurrentText(str(cond.get("field", "")))
            i = self.op.findData(cond.get("op", "eq"))
            self.op.setCurrentIndex(max(0, i))
            self.value.setText(str(cond.get("value", "")))
            self.value2.setText(str(cond.get("value2", "")))

    def to_cond(self):
        if self._allow_agg and self.kind.currentIndex() == 1:
            return {"agg": {
                "sub": self.agg_sub.currentText(),
                "field": self.agg_field.currentText().strip(),
                "op": self.agg_op.currentData(),
                "value": self.agg_val.text(),
                "count_op": self.agg_cnt_op.currentData(),
                "count": self.agg_cnt.value(),
            }}
        c = {"field": self.field.currentText().strip(),
             "op": self.op.currentData(),
             "value": self.value.text()}
        if c["op"] == "between":
            c["value2"] = self.value2.text()
        return c


class _BindingTab(QWidget):
    """欄位綁定編輯器（驗證視窗第二分頁）：每表一條綁定＝驅動欄位＋
    「driver值 × 欄位」勾選矩陣。編輯工作複本，按「套用」才由對話框寫回。"""

    def __init__(self, manager, table_name, parent=None):
        super().__init__(parent)
        self.manager = manager
        self.table_name = table_name
        src = manager.config.get(table_name, {}).get("field_bindings", {})
        self._edits = json.loads(json.dumps(src)) if isinstance(src, dict) else {}
        self._cur_scope = None
        self._loading = False
        self._cards = []       # [{value_cb, boxes:{欄位:QCheckBox}, widget}]

        lo = QVBoxLayout(self)
        lo.setContentsMargins(14, 12, 14, 12)
        lo.setSpacing(8)

        def _cap(t):
            lb = QLabel(t)
            lb.setStyleSheet(f"color:{_C['txt2']}; font-size:12px; background:transparent;")
            return lb

        # ── 引導列：① 選表 → ② 選切換依據的欄位 → ③ 啟用 ──
        top = QHBoxLayout(); top.setSpacing(8)
        top.addWidget(_cap("① 要設定的表"))
        self.f_scope = _NoscrollCombo()
        self.f_scope.addItem(f"母表（{table_name}）", "")
        prefix = table_name + "."
        for k in manager.sub_tables:
            if k.startswith(prefix):
                sub = k[len(prefix):]
                self.f_scope.addItem(f"子表 [{sub}]", sub)
        self.f_scope.currentIndexChanged.connect(self._on_scope_changed)
        top.addWidget(self.f_scope)
        top.addSpacing(14)
        top.addWidget(_cap("② 依哪個欄位的值切換"))
        self.f_driver = _NoscrollCombo(); self.f_driver.setMinimumWidth(160)
        self.f_driver.currentIndexChanged.connect(self._on_driver_changed)
        top.addWidget(self.f_driver)
        top.addSpacing(14)
        self.f_enabled = QCheckBox("③ 啟用")
        self.f_enabled.setStyleSheet(f"color:{_C['txt']}; background:transparent;")
        top.addWidget(self.f_enabled)
        top.addStretch(1)
        lo.addLayout(top)

        hint = QLabel("按「＋ 新增卡片」設定一個值：「當 <欄位>＝<值> 時，顯示勾選的欄位」。"
                      "沒建卡片的值不隱藏任何欄位；沒被任何卡片勾到的欄位＝共用欄位（永遠顯示）；"
                      "ID/FK 與切換欄位本身也永遠顯示。")
        hint.setWordWrap(True)
        hint.setStyleSheet(f"color:{_C['txt3']}; font-size:11px; background:transparent;")
        lo.addWidget(hint)

        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setStyleSheet("background:transparent; border:none;")
        self._cards_host = QWidget()
        self._cards_host.setStyleSheet("background:transparent;")
        self._cards_lo = QVBoxLayout(self._cards_host)
        self._cards_lo.setContentsMargins(0, 2, 0, 2)
        self._cards_lo.setSpacing(8)
        self._scroll.setWidget(self._cards_host)
        lo.addWidget(self._scroll, 1)

        self._load_scope(self.f_scope.currentData())

    # ── helpers ──
    def _scope_columns(self, scope):
        if scope:
            df = self.manager.sub_tables.get(f"{self.table_name}.{scope}")
        else:
            df = self.manager.tables.get(self.table_name)
        return [] if df is None else [str(c) for c in df.columns]

    def _protected_cols(self, scope):
        cfg = self.manager.config.get(self.table_name, {})
        keep = {cfg.get("primary_key", "")}
        if scope:
            keep.add(cfg.get("sub_tables", {}).get(scope, {})
                     .get("foreign_key") or cfg.get("primary_key", ""))
        return keep

    # ── scope / driver 切換 ──
    def _on_scope_changed(self, *_):
        if self._loading:
            return
        self._save_scope(self._cur_scope)
        self._load_scope(self.f_scope.currentData())

    def _on_driver_changed(self, *_):
        if self._loading:
            return
        self._rebuild_cards(keep_checks=False)

    def _load_scope(self, scope):
        self._loading = True
        try:
            self._cur_scope = scope
            b = self._edits.get(scope) or {}
            self.f_enabled.setChecked(bool(b.get("enabled", True)) if b else False)
            self.f_driver.clear()
            self.f_driver.addItem("（不綁定）", "")
            for c in self._scope_columns(scope):
                if c not in self._protected_cols(scope):
                    self.f_driver.addItem(c, c)
            i = self.f_driver.findData(b.get("driver", ""))
            self.f_driver.setCurrentIndex(max(0, i))
        finally:
            self._loading = False
        self._rebuild_cards(keep_checks=True)

    def _bindable_fields(self):
        driver = self.f_driver.currentData() or ""
        skip = self._protected_cols(self._cur_scope) | {driver}
        return [c for c in self._scope_columns(self._cur_scope) if c not in skip]

    def _clear_cards(self):
        while self._cards_lo.count():
            it = self._cards_lo.takeAt(0)
            w = it.widget()
            if w is not None:
                w.hide()
                w.deleteLater()
        self._cards = []

    def _rebuild_cards(self, keep_checks):
        """卡片只在使用者按「＋ 新增卡片」（或 config 已有設定）時存在：
        「當 <欄位>＝<值> 時，顯示勾選的欄位」。值用可打字下拉選，
        候選值優先列現有資料的內容。"""
        self._clear_cards()
        driver = self.f_driver.currentData() or ""
        if not driver:
            ph = QLabel("↑ 先選「② 依哪個欄位的值切換」（例如 SkillComponent），"
                        "再按「＋ 新增卡片」設定各個值要顯示的欄位。")
            ph.setStyleSheet(f"color:{_C['txt3']}; font-size:12px; background:transparent;")
            ph.setWordWrap(True)
            self._cards_lo.addWidget(ph)
            self._cards_lo.addStretch(1)
            return
        b = self._edits.get(self._cur_scope) or {}
        groups = b.get("groups", {}) if (keep_checks and
                                         b.get("driver") == driver) else {}
        for val, fields in groups.items():
            self._add_card(val, set(fields if isinstance(fields, list) else []))
        b_add = _mk_btn("＋ 新增卡片", "success", icon="plus")
        b_add.setFixedHeight(30)
        b_add.setToolTip("新增一個「當 值＝…」的顯示欄位設定")
        b_add.clicked.connect(lambda _=False: self._add_card("", set()))
        self._cards_lo.addWidget(b_add)
        self._cards_lo.addStretch(1)

    def _value_candidates(self, driver):
        """值下拉的候選：現有資料的值優先，其後補 enum options。"""
        vals = []
        if self._cur_scope:
            df = self.manager.sub_tables.get(f"{self.table_name}.{self._cur_scope}")
        else:
            df = self.manager.tables.get(self.table_name)
        if df is not None and driver in df.columns:
            for v in sorted(set(df[driver].astype(str).str.strip())):
                if v and v not in vals:
                    vals.append(v)
        cfg = self.manager.config.get(self.table_name, {})
        if self._cur_scope:
            col_cfg = (cfg.get("sub_tables", {}).get(self._cur_scope, {})
                       .get("columns", {}).get(driver, {}))
        else:
            col_cfg = cfg.get("columns", {}).get(driver, {})
        for o in (col_cfg.get("options") or []):
            if str(o).strip() and str(o) not in vals:
                vals.append(str(o))
        return vals[:60]

    def _add_card(self, value, checked):
        driver = self.f_driver.currentData() or ""
        card = QWidget()
        card.setObjectName("bindCard")
        card.setStyleSheet(
            f"#bindCard {{ background:{_C['card']}; border:1px solid {_C['border']};"
            f" border-radius:8px; }}")
        v = QVBoxLayout(card); v.setContentsMargins(12, 10, 12, 10); v.setSpacing(6)

        hdr = QHBoxLayout(); hdr.setSpacing(8)
        t1 = QLabel(f"當 <b style='color:{_C['txtAcc']}'>{driver}</b> ＝")
        t1.setStyleSheet(f"color:{_C['txt']}; font-size:12px; background:transparent; border:none;")
        hdr.addWidget(t1)
        val_cb = _NoscrollCombo()
        val_cb.setEditable(True)              # 可挑現有值，也可自己打
        val_cb.setMinimumWidth(180)
        val_cb.addItems(self._value_candidates(driver))
        val_cb.setCurrentText(str(value))
        if not value:
            val_cb.setCurrentIndex(-1)
            val_cb.lineEdit().setPlaceholderText("選或輸入值…")
        hdr.addWidget(val_cb)
        t2 = QLabel("時，顯示：")
        t2.setStyleSheet(f"color:{_C['txt']}; font-size:12px; background:transparent; border:none;")
        hdr.addWidget(t2)
        hdr.addStretch(1)
        b_all = _mk_btn("全選", "ghost"); b_none = _mk_btn("全不選", "ghost")
        for btn in (b_all, b_none):
            btn.setFixedHeight(22)
        b_del = _mk_btn("", "danger", icon="trash")
        b_del.setFixedSize(24, 24)
        b_del.setToolTip("移除這張卡片（該值恢復成「顯示全部欄位」）")
        hdr.addWidget(b_all); hdr.addWidget(b_none); hdr.addWidget(b_del)
        v.addLayout(hdr)

        grid = QGridLayout(); grid.setHorizontalSpacing(14); grid.setVerticalSpacing(4)
        boxes = {}
        for i, f in enumerate(self._bindable_fields()):
            cb = QCheckBox(f)
            cb.setChecked(f in checked)
            cb.setStyleSheet(f"color:{_C['txt']}; background:transparent; border:none;")
            grid.addWidget(cb, i // 4, i % 4)
            boxes[f] = cb
        v.addLayout(grid)

        b_all.clicked.connect(lambda _=False, bs=boxes: [cb.setChecked(True) for cb in bs.values()])
        b_none.clicked.connect(lambda _=False, bs=boxes: [cb.setChecked(False) for cb in bs.values()])

        entry = {"value_cb": val_cb, "boxes": boxes, "widget": card}
        self._cards.append(entry)

        def _remove():
            self._cards.remove(entry)
            card.hide()
            card.deleteLater()

        b_del.clicked.connect(lambda _=False: _remove())

        # 新卡片插在「＋ 新增卡片」按鈕前；rebuild 期間按鈕還沒建 → 直接 append
        if self._cards_lo.count() >= 2:
            self._cards_lo.insertWidget(self._cards_lo.count() - 2, card)
        else:
            self._cards_lo.addWidget(card)

    # ── 工作複本 ↔ UI ──
    def _save_scope(self, scope):
        if scope is None:
            return
        driver = self.f_driver.currentData() or ""
        if not driver:
            self._edits.pop(scope, None)
            return
        groups = {}
        for entry in self._cards:
            val = entry["value_cb"].currentText().strip()
            if not val:
                continue              # 沒選值的卡片不寫入
            groups[val] = [f for f, cb in entry["boxes"].items()
                           if cb.isChecked()]
        self._edits[scope] = {"enabled": self.f_enabled.isChecked(),
                              "driver": driver, "groups": groups}

    def result(self):
        """套用時呼叫：回傳 {scope: binding}（空 dict = 全部無綁定）。"""
        self._save_scope(self._cur_scope)
        return {k: v for k, v in self._edits.items() if v.get("driver")}


class ValidationRulesDialog(QDialog):
    """自訂資料驗證規則編輯器（左：規則清單／右：規則內容）。
    編輯的是 config[table]["validations"] 的複本，按「套用」才寫回。"""

    def __init__(self, parent, manager, table_name):
        super().__init__(parent)
        self.manager = manager
        self.table_name = table_name
        self.setWindowTitle(f"驗證規則與欄位綁定 — {table_name}")
        self.setStyleSheet(APP_QSS)
        self.resize(980, 640)

        src = manager.config.get(table_name, {}).get("validations", [])
        self.rules = [_v_norm(json.loads(json.dumps(r))) for r in src
                      if isinstance(r, dict)]
        self._cur = None          # index into self.rules currently in the form
        self._loading = False

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0); outer.setSpacing(0)
        rules_page = QWidget(); rules_page.setStyleSheet("background:transparent;")
        body = QHBoxLayout(rules_page)
        body.setContentsMargins(0, 0, 0, 0); body.setSpacing(0)

        # ── left: rule list ──
        left = QWidget(); left.setFixedWidth(250)
        left.setStyleSheet(f"background:{_C['sidebar']}; border-right:1px solid {_C['border']};")
        llo = QVBoxLayout(left); llo.setContentsMargins(10, 10, 10, 10); llo.setSpacing(8)
        self._list = QListWidget()
        self._list.setStyleSheet(
            f"QListWidget{{background:{_C['panel']}; border:1px solid {_C['border']};"
            f"border-radius:6px;}} QListWidget::item{{padding:7px 8px; border-radius:4px;}}"
            f"QListWidget::item:selected{{background:{_C['card']};}}")
        self._list.currentRowChanged.connect(self._on_select)
        llo.addWidget(self._list, 1)
        brow = QHBoxLayout(); brow.setSpacing(6)
        b_add = _mk_btn("新增", "success", icon="plus");  b_add.clicked.connect(self._add_rule)
        b_cp  = _mk_btn("", "ghost", icon="copy"); b_cp.setFixedWidth(30)
        b_cp.setToolTip("複製選中規則"); b_cp.clicked.connect(self._copy_rule)
        b_del = _mk_btn("", "danger", icon="trash"); b_del.setFixedWidth(30)
        b_del.setToolTip("刪除選中規則"); b_del.clicked.connect(self._del_rule)
        for b in (b_add, b_cp, b_del):
            b.setFixedHeight(30); brow.addWidget(b)
        llo.addLayout(brow)
        body.addWidget(left)

        # ── right: rule form ──
        right = QWidget(); right.setStyleSheet(f"background:{_C['panel']};")
        rscroll = QScrollArea(); rscroll.setWidgetResizable(True)
        rscroll.setStyleSheet("background:transparent; border:none;")
        rscroll.setWidget(right)
        rlo = QVBoxLayout(right); rlo.setContentsMargins(16, 14, 16, 14); rlo.setSpacing(10)

        def _cap(text):
            lb = QLabel(text)
            lb.setStyleSheet(f"color:{_C['txt3']}; font-size:10px; font-weight:600;"
                             f"letter-spacing:1px; background:transparent;")
            return lb

        # 名稱 / 啟用
        row1 = QHBoxLayout(); row1.setSpacing(8)
        self.f_name = QLineEdit(); self.f_name.setPlaceholderText("規則名稱")
        self.f_enabled = QCheckBox("啟用")
        self.f_enabled.setStyleSheet(f"color:{_C['txt']}; background:transparent;")
        row1.addWidget(self.f_name, 1); row1.addWidget(self.f_enabled)
        rlo.addLayout(row1)

        # scope / 嚴重度 / 顏色
        row2 = QHBoxLayout(); row2.setSpacing(8)
        row2.addWidget(_cap("作用範圍"))
        self.f_scope = _NoscrollCombo(); self.f_scope.setMinimumWidth(160)
        self.f_scope.addItem(f"母表（{table_name}）", "")
        for sub in self._sub_names():
            self.f_scope.addItem(f"子表 [{sub}]", sub)
        self.f_scope.currentIndexChanged.connect(self._scope_changed)
        row2.addWidget(self.f_scope)
        row2.addSpacing(10)
        row2.addWidget(_cap("嚴重度"))
        self.f_sev = _NoscrollCombo()
        self.f_sev.addItem("錯誤（擋存檔）", "error")
        self.f_sev.addItem("警告（僅提醒）", "warn")
        row2.addWidget(self.f_sev)
        row2.addSpacing(10)
        row2.addWidget(_cap("顏色"))
        self.f_color = QPushButton(); self.f_color.setFixedSize(46, 26)
        self.f_color.setToolTip("違規儲存格的標記顏色")
        self.f_color.clicked.connect(self._pick_color)
        self._color_val = _V_DEFAULT_COLOR
        row2.addWidget(self.f_color)
        row2.addStretch(1)
        rlo.addLayout(row2)

        # 模式切換
        mode_row = QHBoxLayout(); mode_row.setSpacing(10)
        self.f_mode_builder = QRadioButton("一般（條件組合）")
        self.f_mode_expr    = QRadioButton("進階（表達式）")
        for rb in (self.f_mode_builder, self.f_mode_expr):
            rb.setStyleSheet(f"color:{_C['txt']}; background:transparent;")
            mode_row.addWidget(rb)
        mode_row.addStretch(1)
        self._mode_grp = QButtonGroup(self)
        self._mode_grp.addButton(self.f_mode_builder)
        self._mode_grp.addButton(self.f_mode_expr)
        self.f_mode_builder.toggled.connect(self._mode_changed)
        rlo.addLayout(mode_row)

        self._stack = QStackedWidget()
        rlo.addWidget(self._stack, 1)

        # ── builder page ──
        bpage = QWidget(); bpage.setStyleSheet("background:transparent;")
        blo = QVBoxLayout(bpage); blo.setContentsMargins(0, 0, 0, 0); blo.setSpacing(8)
        when_hdr = QHBoxLayout()
        when_hdr.addWidget(_cap("當（條件成立時檢查）"))
        self.f_logic = _NoscrollCombo()
        self.f_logic.addItem("全部成立 (AND)", "and")
        self.f_logic.addItem("任一成立 (OR)", "or")
        when_hdr.addWidget(self.f_logic)
        b_wadd = _mk_btn("＋ 條件", "ghost"); b_wadd.setFixedHeight(24)
        b_wadd.clicked.connect(lambda: self._add_cond(self._when_lo, {}))
        when_hdr.addWidget(b_wadd); when_hdr.addStretch(1)
        blo.addLayout(when_hdr)
        hint_w = QLabel("（留空 = 每一列都檢查「則」）")
        hint_w.setStyleSheet(f"color:{_C['txt3']}; font-size:11px; background:transparent;")
        blo.addWidget(hint_w)
        self._when_box = QWidget(); self._when_box.setStyleSheet("background:transparent;")
        self._when_lo = QVBoxLayout(self._when_box)
        self._when_lo.setContentsMargins(0, 0, 0, 0); self._when_lo.setSpacing(4)
        blo.addWidget(self._when_box)

        then_hdr = QHBoxLayout()
        then_hdr.addWidget(_cap("則（必須全部成立，否則違規）"))
        b_tadd = _mk_btn("＋ 要求", "ghost"); b_tadd.setFixedHeight(24)
        b_tadd.clicked.connect(lambda: self._add_cond(self._then_lo, {}))
        then_hdr.addWidget(b_tadd); then_hdr.addStretch(1)
        blo.addLayout(then_hdr)
        self._then_box = QWidget(); self._then_box.setStyleSheet("background:transparent;")
        self._then_lo = QVBoxLayout(self._then_box)
        self._then_lo.setContentsMargins(0, 0, 0, 0); self._then_lo.setSpacing(4)
        blo.addWidget(self._then_box)
        blo.addStretch(1)
        self._stack.addWidget(bpage)

        # ── expr page ──
        epage = QWidget(); epage.setStyleSheet("background:transparent;")
        elo = QVBoxLayout(epage); elo.setContentsMargins(0, 0, 0, 0); elo.setSpacing(6)
        elo.addWidget(_cap("表達式（回傳 True＝通過，False＝違規）"))
        self.f_expr = QPlainTextEdit()
        self.f_expr.setPlaceholderText(
            'SkillComponent != "ContinueBuff" or not empty(EffectDurationTime)')
        self.f_expr.setMaximumHeight(96)
        elo.addWidget(self.f_expr)
        help_lbl = QLabel(
            "可用：本列欄位直接寫欄名｜master.欄位（子表規則讀母表）｜"
            "empty(x)、num(x)、match(regex, x)、len/str/int/float/abs/min/max/round｜"
            "any_sub(\"子表\", \"條件式\")、count_sub(\"子表\", \"條件式\")（母表規則）｜"
            "and / or / not、比較與四則運算")
        help_lbl.setWordWrap(True)
        help_lbl.setStyleSheet(f"color:{_C['txt3']}; font-size:11px; background:transparent;")
        elo.addWidget(help_lbl)
        elo.addStretch(1)
        self._stack.addWidget(epage)

        # 標記欄位
        mark_row = QHBoxLayout(); mark_row.setSpacing(8)
        mark_row.addWidget(_cap("標記欄位"))
        self.f_mark = QLineEdit()
        self.f_mark.setPlaceholderText("逗號分隔；留空＝自動（用「則」的欄位）")
        mark_row.addWidget(self.f_mark, 1)
        rlo.addLayout(mark_row)

        # 測試列
        test_row = QHBoxLayout(); test_row.setSpacing(8)
        b_test = _mk_btn("▶ 測試規則", "primary"); b_test.setFixedHeight(28)
        b_test.clicked.connect(self._test_rule)
        self._test_lbl = QLabel("")
        self._test_lbl.setStyleSheet(f"color:{_C['txt2']}; font-size:12px; background:transparent;")
        self._test_lbl.setWordWrap(True)
        test_row.addWidget(b_test); test_row.addWidget(self._test_lbl, 1)
        rlo.addLayout(test_row)

        body.addWidget(rscroll, 1)

        tabs = QTabWidget()
        tabs.setDocumentMode(True)
        tabs.addTab(rules_page, "驗證規則")
        self._binding_tab = _BindingTab(manager, table_name)
        tabs.addTab(self._binding_tab, "欄位綁定")
        outer.addWidget(tabs, 1)

        # ── bottom bar ──
        bb = QWidget(); bb.setStyleSheet(
            f"background:{_C['sidebar']}; border-top:1px solid {_C['border']};")
        bl = QHBoxLayout(bb); bl.setContentsMargins(16, 10, 16, 10)
        bl.addStretch(1)
        b_cancel = _mk_btn("取消"); b_cancel.setFixedHeight(32)
        b_cancel.clicked.connect(self.reject)
        b_ok = _mk_btn("套用", "primary"); b_ok.setFixedHeight(32)
        b_ok.clicked.connect(self._apply)
        bl.addWidget(b_cancel); bl.addWidget(b_ok)
        outer.addWidget(bb)

        self._right = right
        self._reload_list(select=0 if self.rules else -1)
        right.setEnabled(bool(self.rules))

    # ── helpers ──
    def _sub_names(self):
        prefix = self.table_name + "."
        return [k[len(prefix):] for k in self.manager.sub_tables
                if k.startswith(prefix)]

    def sub_columns(self, sub):
        df = self.manager.sub_tables.get(f"{self.table_name}.{sub}")
        return [] if df is None else [str(c) for c in df.columns]

    def _master_columns(self):
        df = self.manager.tables.get(self.table_name)
        return [] if df is None else [str(c) for c in df.columns]

    def _field_choices(self, scope):
        if scope:
            return (self.sub_columns(scope)
                    + [f"master.{c}" for c in self._master_columns()])
        return self._master_columns()

    def _cond_rows(self, lo):
        return [lo.itemAt(i).widget() for i in range(lo.count())
                if isinstance(lo.itemAt(i).widget(), _VCondRow)]

    def _clear_conds(self, lo):
        for w in self._cond_rows(lo):
            lo.removeWidget(w)
            w.hide()          # removeWidget 後仍是可見子元件，先藏再刪避免殘影
            w.deleteLater()

    def _add_cond(self, lo, cond):
        scope = self.f_scope.currentData()
        row = _VCondRow(self, cond, allow_agg=(scope == ""))
        row.set_field_choices(self._field_choices(scope), self._sub_names())
        row._load(cond or {})
        row._kind_changed()
        row.removed.connect(
            lambda w, l=lo: (l.removeWidget(w), w.hide(), w.deleteLater()))
        lo.addWidget(row)
        return row

    # ── list handling ──
    def _rule_item_text(self, r):
        sev = "🔴" if r["severity"] == "error" else "🟡"
        off = "" if r.get("enabled") else "（停用）"
        sc  = f"[{r['scope']}] " if r.get("scope") else ""
        return f"{sev} {sc}{r['name']}{off}"

    def _reload_list(self, select=None):
        cur = self._list.currentRow() if select is None else select
        self._list.blockSignals(True)
        self._list.clear()
        for r in self.rules:
            it = QListWidgetItem(self._rule_item_text(r))
            pm = QPixmap(12, 12); pm.fill(QColor(r.get("color", _V_DEFAULT_COLOR)))
            it.setIcon(QIcon(pm))
            self._list.addItem(it)
        self._list.blockSignals(False)
        if 0 <= cur < len(self.rules):
            self._list.setCurrentRow(cur)     # triggers _on_select
        else:
            self._cur = None

    def _on_select(self, row):
        if self._loading:
            return
        self._save_form()                     # persist previous rule's form
        if 0 <= row < len(self.rules):
            self._cur = row
            self._load_form(self.rules[row])
            self._right.setEnabled(True)
        else:
            self._cur = None
            self._right.setEnabled(False)

    def _add_rule(self):
        self._save_form()
        r = _v_new_rule()
        self.rules.append(r)
        self._cur = None                      # avoid re-saving old form over it
        self._reload_list(select=len(self.rules) - 1)

    def _copy_rule(self):
        i = self._list.currentRow()
        if not (0 <= i < len(self.rules)):
            return
        self._save_form()
        r = json.loads(json.dumps(self.rules[i]))
        r["id"] = uuid.uuid4().hex[:8]
        r["name"] += "（複製）"
        self.rules.insert(i + 1, r)
        self._cur = None
        self._reload_list(select=i + 1)

    def _del_rule(self):
        i = self._list.currentRow()
        if not (0 <= i < len(self.rules)):
            return
        del self.rules[i]
        self._cur = None
        self._reload_list(select=min(i, len(self.rules) - 1))
        if not self.rules:
            self._right.setEnabled(False)

    # ── form ↔ rule ──
    def _load_form(self, r):
        self._loading = True
        try:
            self.f_name.setText(r["name"])
            self.f_enabled.setChecked(bool(r.get("enabled", True)))
            i = self.f_scope.findData(r.get("scope", ""))
            self.f_scope.setCurrentIndex(max(0, i))
            i = self.f_sev.findData(r.get("severity", "error"))
            self.f_sev.setCurrentIndex(max(0, i))
            self._set_color(r.get("color", _V_DEFAULT_COLOR))
            i = self.f_logic.findData(r.get("when", {}).get("logic", "and"))
            self.f_logic.setCurrentIndex(max(0, i))
            self._clear_conds(self._when_lo)
            self._clear_conds(self._then_lo)
            for c in r.get("when", {}).get("conds", []):
                self._add_cond(self._when_lo, c)
            for c in r.get("then", []):
                self._add_cond(self._then_lo, c)
            self.f_expr.setPlainText(r.get("expr", ""))
            self.f_mark.setText(", ".join(r.get("mark", [])))
            if r.get("mode") == "expr":
                self.f_mode_expr.setChecked(True)
            else:
                self.f_mode_builder.setChecked(True)
            self._mode_changed()
            self._test_lbl.setText("")
        finally:
            self._loading = False

    def _save_form(self):
        if self._cur is None or not (0 <= self._cur < len(self.rules)):
            return
        self.rules[self._cur] = self._form_rule(self.rules[self._cur])
        it = self._list.item(self._cur)
        if it is not None:
            it.setText(self._rule_item_text(self.rules[self._cur]))
            pm = QPixmap(12, 12); pm.fill(QColor(self._color_val))
            it.setIcon(QIcon(pm))

    def _form_rule(self, base):
        r = dict(base)
        r["name"] = self.f_name.text().strip() or "未命名規則"
        r["enabled"] = self.f_enabled.isChecked()
        r["scope"] = self.f_scope.currentData() or ""
        r["severity"] = self.f_sev.currentData()
        r["color"] = self._color_val
        r["mode"] = "expr" if self.f_mode_expr.isChecked() else "builder"
        r["when"] = {"logic": self.f_logic.currentData(),
                     "conds": [w.to_cond() for w in self._cond_rows(self._when_lo)]}
        r["then"] = [w.to_cond() for w in self._cond_rows(self._then_lo)]
        r["expr"] = self.f_expr.toPlainText().strip()
        r["mark"] = [t.strip() for t in self.f_mark.text().split(",") if t.strip()]
        return r

    def _scope_changed(self, *_):
        if self._loading:
            return
        scope = self.f_scope.currentData()
        # scope 變更 → 條件列的欄位選項全部換掉（agg 只在母表可用）
        for lo in (self._when_lo, self._then_lo):
            conds = [w.to_cond() for w in self._cond_rows(lo)]
            self._clear_conds(lo)
            for c in conds:
                if c.get("agg") and scope:
                    continue                  # 子表 scope 不支援聚合條件
                self._add_cond(lo, c)

    def _mode_changed(self, *_):
        self._stack.setCurrentIndex(1 if self.f_mode_expr.isChecked() else 0)

    def _set_color(self, color):
        self._color_val = color
        self.f_color.setStyleSheet(
            f"background:{color}; border:1px solid {_C['border']}; border-radius:5px;")

    def _pick_color(self):
        c = QColorDialog.getColor(QColor(self._color_val), self, "選擇標記顏色")
        if c.isValid():
            self._set_color(c.name())

    def _test_rule(self):
        if self._cur is None:
            return
        rule = self._form_rule(self.rules[self._cur])
        cnt, samples, err = self.manager.validator.test_rule(self.table_name, rule)
        if err:
            self._test_lbl.setText(f'<span style="color:{_C["red"]}">✗ {err}</span>')
            return
        if cnt == 0:
            self._test_lbl.setText(
                f'<span style="color:{_C["green"]}">✓ 目前資料全部通過</span>')
        else:
            ex = "、".join(samples[:6]) + ("…" if cnt > 6 else "")
            self._test_lbl.setText(
                f'<span style="color:{_C["yellow"]}">⚠ 目前資料有 {cnt} 列違規'
                f'（{ex}）</span>')

    def bindings(self):
        return self._binding_tab.result()

    def _apply(self):
        self._save_form()
        self.accept()


# ── App ───────────────────────────────────────────────────────────────────────

class _Doc:
    """One open JSON file: its own data manager, tables/notes tab widget, lazily
    built table editors, and floating note windows."""
    __slots__ = ("manager", "tab_widget", "editors", "note_floats")

    def __init__(self, manager, tab_widget):
        self.manager = manager
        self.tab_widget = tab_widget
        self.editors: dict[str, "TableEditor | None"] = {}
        self.note_floats: dict[str, dict] = {}


class App(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("JsonEditor")
        self.resize(1280, 800)
        self.setMinimumSize(900, 600)
        # Multi-document: each open file is a _Doc (own manager + tab widget +
        # editors + note floats). A "boot" manager holds shared config / recent /
        # ref-cache and acts as the fallback when no file is open.
        self._boot_mgr       = JsonDataManager()
        self._shared         = self._boot_mgr.shared_state()
        self._docs:          list["_Doc"] = []
        self._active         = -1
        self._empty_editors  = {}
        self._empty_floats   = {}
        self._active_worker  = None
        self._snackbar_timer = QTimer(self)
        self._snackbar_timer.setSingleShot(True)
        self._snackbar_timer.timeout.connect(self._clear_status)

        self._setup_content()

        # Keyboard shortcuts
        self.addAction(QAction(parent=self, shortcut=QKeySequence("Ctrl+S"), triggered=self.save_file))
        self.addAction(QAction(parent=self, shortcut=QKeySequence("Ctrl+O"), triggered=self.load_file))
        self.addAction(QAction(parent=self, shortcut=QKeySequence("Ctrl+W"), triggered=self._close_active_doc))
        self.addAction(QAction(parent=self, shortcut=QKeySequence("Ctrl+F"), triggered=self._show_search))

        self._show_welcome()
        self._update_sync()

    # ── Active-document accessors (most methods operate on the active doc) ──────
    @property
    def manager(self):
        if 0 <= self._active < len(self._docs):
            return self._docs[self._active].manager
        return self._boot_mgr

    @property
    def _tab_widget(self):
        if 0 <= self._active < len(self._docs):
            return self._docs[self._active].tab_widget
        return self._empty_tabs

    @property
    def _editors(self):
        if 0 <= self._active < len(self._docs):
            return self._docs[self._active].editors
        return self._empty_editors

    @property
    def _note_floats(self):
        if 0 <= self._active < len(self._docs):
            return self._docs[self._active].note_floats
        return self._empty_floats

    def _active_doc(self):
        return self._docs[self._active] if 0 <= self._active < len(self._docs) else None

    # ── Content layout ────────────────────────────────────────────────────────

    def _setup_content(self):
        central = QWidget()
        central.setStyleSheet(f"background:{_C['bg']};")
        self.setCentralWidget(central)
        clo = QVBoxLayout(central)
        clo.setContentsMargins(0, 0, 0, 0)
        clo.setSpacing(0)

        # ── Top bar ──
        topbar = QWidget()
        topbar.setObjectName("topbar")
        topbar.setFixedHeight(54)
        topbar.setStyleSheet(
            f"QWidget#topbar {{ background:{_C['sidebar']}; border-bottom:1px solid {_C['border']}; }}"
        )
        tlo = QHBoxLayout(topbar)
        tlo.setContentsMargins(16, 0, 16, 0)
        tlo.setSpacing(10)

        # Logo
        logo_box = QLabel("{ }")
        logo_box.setFixedSize(34, 34)
        logo_box.setAlignment(Qt.AlignCenter)
        logo_box.setStyleSheet(
            f"font-family:Consolas; font-size:14px; font-weight:bold; color:white; "
            f"background:qlineargradient(x1:0,y1:0,x2:1,y2:1,"
            f"stop:0 {_C['accent']},stop:1 #8B5CF6); border-radius:9px;"
        )
        tlo.addWidget(logo_box)

        app_name = QLabel("JsonEditor")
        app_name.setStyleSheet(
            f"font-size:15px; font-weight:700; color:{_C['txt']}; background:transparent;"
        )
        tlo.addWidget(app_name)

        app_sub = QLabel("Pro")
        app_sub.setStyleSheet(
            f"font-size:11px; color:{_C['txt2']}; background:transparent;"
        )
        tlo.addWidget(app_sub)
        tlo.addStretch(1)

        # Action buttons
        for text, role, slot in [
            ("📂  開啟",   "",        self.load_file),
            ("+ 新建",    "",        self._new_file),
            ("💾  儲存",  "success", self.save_file),
            ("📤  匯出 Excel", "",  self._export_to_excel),
        ]:
            b = _mk_btn(text, role)
            b.setFixedHeight(34)
            b.clicked.connect(slot)
            tlo.addWidget(b)

        tlo.addSpacing(10)
        tlo.addWidget(_vsep())
        tlo.addSpacing(10)

        # Sync indicator
        self._sync_dot = QLabel("●")
        self._sync_dot.setFixedWidth(14)
        self._sync_dot.setStyleSheet(f"color:{_C['green']}; background:transparent; font-size:9px;")
        self._sync_lbl = QLabel("已儲存")
        self._sync_lbl.setStyleSheet(
            f"font-size:12px; font-weight:500; color:{_C['green']}; background:transparent;"
        )
        tlo.addWidget(self._sync_dot)
        tlo.addWidget(self._sync_lbl)

        tlo.addSpacing(14)

        # Util buttons
        for text, slot, tip in [
            ("🔍", self._show_search,      "全域搜尋（Ctrl+F）"),
            ("📝", self._note_menu,        "筆記：新增／重新開啟已關閉"),
            ("🩺", self.open_health_check, "資料健檢"),
            ("✅", self.open_validation_rules, "驗證規則（自訂資料驗證）"),
            ("⚙",  self.open_config,       "配置設定"),
            ("🕓", self._show_recent_menu, "最近開啟"),
        ]:
            b = _mk_btn(text, "ghost")
            b.setFixedSize(36, 34)
            b.setToolTip(tip)
            b.clicked.connect(slot)
            tlo.addWidget(b)

        clo.addWidget(topbar)

        # ── Stack: welcome  ↔  documents page ──
        self._stack = QStackedWidget()
        clo.addWidget(self._stack, 1)

        self._welcome = WelcomeWidget(self._boot_mgr)
        self._welcome.open_file.connect(self.load_file)
        self._welcome.new_file.connect(self._new_file)
        self._welcome.open_recent.connect(self._load_recent)
        self._stack.addWidget(self._welcome)

        # documents page = file tab bar (one tab per open json) + a stack of the
        # per-file table/note tab widgets
        docs_page = QWidget(); docs_page.setStyleSheet("background:transparent;")
        dpl = QVBoxLayout(docs_page); dpl.setContentsMargins(0, 0, 0, 0); dpl.setSpacing(0)
        file_row = QWidget()
        file_row.setStyleSheet(f"background:{_C['bg']}; border-bottom:1px solid {_C['border']};")
        frl = QHBoxLayout(file_row); frl.setContentsMargins(6, 4, 6, 0); frl.setSpacing(6)
        self._file_bar = QTabBar()
        self._file_bar.setExpanding(False)
        self._file_bar.setTabsClosable(True)
        self._file_bar.setMovable(False)   # keep index aligned with _docs/_doc_stack
        self._file_bar.setDrawBase(False)
        self._file_bar.currentChanged.connect(self._switch_doc)
        self._file_bar.tabCloseRequested.connect(self._close_doc)
        add_btn = _mk_btn("＋ 開檔", "ghost"); add_btn.setFixedHeight(26)
        add_btn.setToolTip("開啟另一個 JSON（多開）")
        add_btn.clicked.connect(self.load_file)
        frl.addWidget(self._file_bar, 1); frl.addWidget(add_btn)
        dpl.addWidget(file_row)
        self._doc_stack = QStackedWidget()
        dpl.addWidget(self._doc_stack, 1)
        self._stack.addWidget(docs_page)
        self._docs_page = docs_page
        self._empty_tabs = QTabWidget()   # harmless fallback when no doc is active

        # ── Status strip ──
        status_bar = QWidget()
        status_bar.setFixedHeight(26)
        status_bar.setStyleSheet(
            f"background:{_C['sidebar']}; border-top:1px solid {_C['border']};"
        )
        slo = QHBoxLayout(status_bar)
        slo.setContentsMargins(14, 0, 14, 0)
        slo.setSpacing(0)
        self._status_lbl = QLabel("就緒")
        self._status_lbl.setStyleSheet(
            f"font-size:11px; color:{_C['txt2']}; background:transparent;"
        )
        self._path_lbl = QLabel("")
        self._path_lbl.setStyleSheet(
            f"font-size:11px; color:{_C['txt3']}; background:transparent;"
        )
        self._path_lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        slo.addWidget(self._status_lbl, 1)
        slo.addWidget(self._path_lbl)
        clo.addWidget(status_bar)

    def _show_welcome(self):
        self._stack.setCurrentWidget(self._welcome)

    def _show_editor(self):
        self._stack.setCurrentWidget(self._docs_page)

    # ── Sync / status ─────────────────────────────────────────────────────────

    def _update_sync(self):
        dirty = self.manager.dirty
        if dirty:
            self._sync_dot.setStyleSheet(
                f"color:{_C['yellow']}; background:transparent; font-size:9px;"
            )
            self._sync_lbl.setText("未儲存")
            self._sync_lbl.setStyleSheet(
                f"font-size:12px; font-weight:500; color:{_C['yellow']}; background:transparent;"
            )
        else:
            self._sync_dot.setStyleSheet(
                f"color:{_C['green']}; background:transparent; font-size:9px;"
            )
            self._sync_lbl.setText("已儲存")
            self._sync_lbl.setStyleSheet(
                f"font-size:12px; font-weight:500; color:{_C['green']}; background:transparent;"
            )
        if self.manager.json_path:
            fname = os.path.basename(self.manager.json_path)
            self.setWindowTitle(f"JsonEditor — {fname}" + (" *" if dirty else ""))
            self._path_lbl.setText(self.manager.json_path)
        else:
            self.setWindowTitle("JsonEditor")
            self._path_lbl.setText("")
        # reflect dirty state on the active file tab
        if 0 <= self._active < len(self._docs) and hasattr(self, "_file_bar"):
            fname = os.path.basename(self.manager.json_path or "未命名")
            self._file_bar.setTabText(self._active, fname + (" *" if dirty else ""))

    def show_snackbar(self, text: str, duration_ms: int = 3000, color: str = ""):
        self._status_lbl.setText(text)
        self._status_lbl.setStyleSheet(
            f"font-size:11px; color:{color or _C['txtAcc']}; background:transparent;"
        )
        self._snackbar_timer.start(duration_ms)

    def _clear_status(self):
        self._status_lbl.setText("就緒")
        self._status_lbl.setStyleSheet(
            f"font-size:11px; color:{_C['txt2']}; background:transparent;"
        )

    # Alias for backwards-compat with TableEditor signal
    def _update_title(self):
        self._update_sync()

    # ── Loading ───────────────────────────────────────────────────────────────

    def _set_loading(self, loading: bool, msg: str = "就緒"):
        self.setEnabled(not loading)
        self._status_lbl.setText(msg)
        self._status_lbl.setStyleSheet(
            f"font-size:11px; color:{_C['cyan'] if loading else _C['txt2']}; background:transparent;"
        )
        QApplication.processEvents()

    # ── File I/O ──────────────────────────────────────────────────────────────

    def load_file(self):
        last_dir = os.path.dirname(self.manager._full_config.get("_last_file", "")) or ""
        path, _ = QFileDialog.getOpenFileName(
            self, "開啟 JSON", last_dir, "JSON 檔案 (*.json);;所有檔案 (*.*)"
        )
        if path:
            self._load_path(path)

    def _new_file(self):
        path, _ = QFileDialog.getSaveFileName(self, "新建 JSON", "", "JSON 檔案 (*.json)")
        if not path: return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump([], f, indent=4, ensure_ascii=False)
        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e)); return
        self._load_path(path)

    def _load_recent(self, path):
        if not os.path.exists(path):
            QMessageBox.warning(self, "錯誤", f"找不到檔案：\n{path}")
            rf = self.manager._recent_files
            rf[:] = [p for p in rf if p != path]   # in place (shared list)
            self.manager.save_config()
            return
        self._load_path(path)

    def _load_path(self, path):
        # already open → just switch to that file tab
        norm = os.path.normpath(path)
        for i, d in enumerate(self._docs):
            if d.manager.json_path and os.path.normpath(d.manager.json_path) == norm:
                self._file_bar.setCurrentIndex(i)
                return
        mgr = JsonDataManager(shared=self._shared)
        self._set_loading(True, f"載入 {os.path.basename(path)}…")
        orig = sys.getswitchinterval()
        sys.setswitchinterval(0.001)
        worker = _LoadWorker(mgr, path)
        self._active_worker = worker

        def _done():
            sys.setswitchinterval(orig)
            self._active_worker = None
            self._set_loading(False)
            self._add_document(mgr)       # new tab widget + file tab, made active
            mgr._full_config["_last_file"] = path
            mgr.save_config()
            self._check_config_paths()

        def _err(msg):
            sys.setswitchinterval(orig)
            self._active_worker = None
            self._set_loading(False)
            QMessageBox.critical(self, "載入失敗", msg)

        worker.done.connect(_done)
        worker.error.connect(_err)
        worker.start()

    # ── Multi-document management ───────────────────────────────────────────────
    def _add_document(self, manager):
        tabw = QTabWidget()
        tabw.setObjectName("main-tabs")
        tabw.setDocumentMode(True)
        tabw.currentChanged.connect(self._on_tab_changed)
        tabw.tabBar().setContextMenuPolicy(Qt.CustomContextMenu)
        tabw.tabBar().customContextMenuRequested.connect(self._note_tab_menu)
        self._docs.append(_Doc(manager, tabw))
        self._doc_stack.addWidget(tabw)
        fname = os.path.basename(manager.json_path or "未命名")
        self._file_bar.blockSignals(True)
        fi = self._file_bar.addTab(fname)
        self._file_bar.setTabToolTip(fi, manager.json_path or "")
        self._file_bar.setCurrentIndex(fi)
        self._file_bar.blockSignals(False)
        self._active = len(self._docs) - 1
        self._doc_stack.setCurrentWidget(tabw)
        self._refresh_ui()            # build this doc's table/note tabs

    def _switch_doc(self, i):
        if not (0 <= i < len(self._docs)):
            return
        self._active = i
        self._doc_stack.setCurrentWidget(self._docs[i].tab_widget)
        self._show_editor()
        self._ensure_editor(self._tab_widget.currentIndex())
        self._update_sync()

    def _close_active_doc(self):
        if self._docs:
            self._close_doc(self._active)

    def _close_doc(self, i):
        if not (0 <= i < len(self._docs)):
            return
        doc = self._docs[i]
        if doc.manager.dirty:
            nm = os.path.basename(doc.manager.json_path or "未命名")
            if QMessageBox.question(
                    self, "關閉檔案", f"「{nm}」有未儲存變更，仍要關閉？",
                    QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
                return
        self._active = i                    # so _close_all_note_floats targets it
        self._close_all_note_floats()
        self._doc_stack.removeWidget(doc.tab_widget)
        doc.tab_widget.deleteLater()
        del self._docs[i]
        self._file_bar.blockSignals(True)
        self._file_bar.removeTab(i)
        self._file_bar.blockSignals(False)
        if not self._docs:
            self._active = -1
            self._welcome.refresh(self._boot_mgr)
            self._show_welcome()
            self._update_sync()
            return
        new_i = min(i, len(self._docs) - 1)
        self._active = -1                   # force _switch_doc to re-apply
        self._file_bar.blockSignals(True)
        self._file_bar.setCurrentIndex(new_i)
        self._file_bar.blockSignals(False)
        self._switch_doc(new_i)

    # ── Config path validation (offer to re-pick stale external paths) ──────────
    def _store_path(self, p, base):
        """Store relative to the json dir when possible (mirrors config dialog)."""
        if not base:
            return p
        try:
            return os.path.relpath(p, base)
        except ValueError:
            return p

    def _check_config_paths(self):
        """After load: if config has external paths that no longer exist, offer
        to re-pick them (外部文字表 json_path / 圖片資料夾 base_folder)."""
        if not self.manager.json_path:
            return
        json_dir = os.path.dirname(self.manager.json_path)
        changed = False
        for tname, cfg in list(self.manager.config.items()):
            if not isinstance(cfg, dict):
                continue
            # per-column text_ref sources (master + sub-table columns)
            scopes = [(cfg.get("columns", {}), "")]
            for sname, scfg in cfg.get("sub_tables", {}).items():
                if isinstance(scfg, dict):
                    scopes.append((scfg.get("columns", {}), f"（子表 {sname}）"))
            for cols, suffix in scopes:
                for cname, ccfg in cols.items():
                    if not (isinstance(ccfg, dict) and ccfg.get("type") == "text_ref"):
                        continue
                    src = ccfg.get("text_ref", {}) if isinstance(ccfg.get("text_ref"), dict) else {}
                    ref = (src.get("json_path") or "").strip()
                    if not ref:
                        continue
                    abs_ref = ref if os.path.isabs(ref) else os.path.join(json_dir, ref)
                    if os.path.isfile(abs_ref):
                        continue
                    if self._prompt_repath(tname, f"欄位「{cname}」{suffix}的外部文字表", abs_ref):
                        new, _ = QFileDialog.getOpenFileName(
                            self, f"重新選擇外部文字表（{cname}）", json_dir,
                            "JSON 檔案 (*.json);;所有檔案 (*.*)")
                        if new:
                            src["json_path"] = self._store_path(new, json_dir)
                            ccfg["text_ref"] = src
                            changed = True
            imgp = cfg.get("image_preview", {})
            base = (imgp.get("base_folder") or "").strip() if isinstance(imgp, dict) else ""
            if base:
                abs_base = base if os.path.isabs(base) else os.path.join(json_dir, base)
                if not os.path.isdir(abs_base):
                    if self._prompt_repath(tname, "圖片資料夾", abs_base):
                        new = QFileDialog.getExistingDirectory(
                            self, f"重新選擇圖片資料夾（{tname}）", json_dir)
                        if new:
                            imgp["base_folder"] = self._store_path(new, json_dir)
                            cfg["image_preview"] = imgp
                            changed = True
        if changed:
            self.manager.save_config()
            self.manager.invalidate_ref_cache()
            idx = self._tab_widget.currentIndex()
            if idx >= 0:
                ed = self._editors.get(self._tab_widget.tabText(idx))
                if ed:
                    ed.reload_after_config()
            self.show_snackbar("✓ 路徑已更新", color=_C["green"])

    def _prompt_repath(self, tname, label, badpath):
        return QMessageBox.question(
            self, "路徑不存在",
            f"[{tname}] 的{label}路徑找不到：\n{badpath}\n\n是否要重新設定？",
            QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes

    def save_file(self):
        if not self.manager.json_path:
            QMessageBox.warning(self, "提示", "尚未載入任何 JSON 檔案"); return
        # a sub-table cell editor may still be open (Ctrl+S / NoFocus card list
        # don't focus-out) — commit it so the pending value makes the save
        for ed in self._editors.values():
            if ed is not None:
                ed.flush_pending_edits()
        # ── 資料驗證閘門：先全量重驗，錯誤擋存檔、警告可放行 ──
        self.manager.validator.validate_all()
        v_items = self.manager.validator.summary()
        if v_items:
            errs = sum(1 for i in v_items if i["severity"] == "error")
            warns = len(v_items) - errs
            proceed = self._show_validation_gate(v_items, errs, warns)
            for ed in self._editors.values():        # gate 重驗後刷新標示
                if ed is not None:
                    ed._refresh_validation_visuals()
            if not proceed:
                return
        # notes live in config → flush widget state and persist alongside the DB
        self._flush_notes()
        self.manager.save_config()
        self._set_loading(True, "儲存中…")
        orig = sys.getswitchinterval()
        sys.setswitchinterval(0.001)
        worker = _SaveWorker(self.manager)
        self._active_worker = worker

        saved_path = self.manager.json_path

        def _done():
            sys.setswitchinterval(orig)
            self._active_worker = None
            self._set_loading(False)
            # the file just saved may be another file's external text table →
            # drop its cached rows and re-resolve references in every open doc
            self.manager.invalidate_ref_cache(saved_path)
            self._refresh_all_ref_displays()
            self.show_snackbar("✓ 已儲存", color=_C["green"])
            self._update_sync()

        def _err(msg):
            sys.setswitchinterval(orig)
            self._active_worker = None
            self._set_loading(False)
            QMessageBox.critical(self, "存檔失敗", msg)

        worker.done.connect(_done)
        worker.error.connect(_err)
        worker.start()

    def _refresh_all_ref_displays(self):
        """Re-resolve text_ref display across every open document (after an
        external text table was saved and its cache was invalidated)."""
        for doc in self._docs:
            for ed in doc.editors.values():
                if ed is not None:
                    try:
                        ed.refresh_ref_display()
                    except Exception:
                        pass

    # ── Export to Excel ───────────────────────────────────────────────────────

    def _export_to_excel(self):
        if not self.manager.json_path:
            QMessageBox.warning(self, "提示", "尚未載入任何 JSON 檔案"); return
        if not self.manager.tables:
            QMessageBox.warning(self, "提示", "沒有資料可以匯出"); return
        base = os.path.splitext(os.path.basename(self.manager.json_path))[0]
        default = os.path.join(os.path.dirname(self.manager.json_path), base + ".xlsx")
        path, _ = QFileDialog.getSaveFileName(
            self, "匯出為 Excel（可選現有檔案直接覆蓋）",
            default, "Excel 檔案 (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            self._write_xlsx(path)
        except Exception as e:
            QMessageBox.critical(self, "匯出失敗", str(e))
            return
        self.show_snackbar(f"✓ 已匯出 Excel: {os.path.basename(path)}", color=_C["green"])

    def _write_xlsx(self, path):
        import openpyxl
        from openpyxl.comments import Comment

        def _safe_sheet(name):
            bad = '\\/*?:[]'
            s = "".join("_" if c in bad else c for c in str(name))
            return s[:31] or "Sheet1"

        def _cell(val, col_type):
            s = "" if val is None else str(val).strip()
            if s == "":
                if col_type == "int":   return 0
                if col_type == "float": return 0.0
                if col_type == "bool":  return False
                return None
            if col_type == "int":
                try: return int(float(s))
                except (ValueError, TypeError): return s
            if col_type == "float":
                try: return float(s)
                except (ValueError, TypeError): return s
            if col_type == "bool":
                return s.lower() in ("true", "1", "yes")
            return s  # string / enum / array(comma-joined) / text_ref

        def _write_table(ws, df, cols_cfg):
            cols = list(df.columns)
            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=1, column=ci, value=str(col))
                note = (cols_cfg.get(col, {}) or {}).get("note", "")
                if note:
                    cell.comment = Comment(note, "JsonEditor")
            for ri, (_, row) in enumerate(df.iterrows(), 2):
                for ci, col in enumerate(cols, 1):
                    ct = (cols_cfg.get(col, {}) or {}).get("type", "string")
                    ws.cell(row=ri, column=ci, value=_cell(row[col], ct))

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        for tname, tdf in self.manager.tables.items():
            cfg = self.manager.config.get(tname, {})
            ws = wb.create_sheet(title=_safe_sheet(tname))
            _write_table(ws, tdf, cfg.get("columns", {}))
            prefix = tname + "."
            for sub_full, sdf in self.manager.sub_tables.items():
                if not sub_full.startswith(prefix):
                    continue
                sub_name = sub_full[len(prefix):]
                ws2 = wb.create_sheet(title=_safe_sheet(sub_name))
                _write_table(
                    ws2, sdf,
                    cfg.get("sub_tables", {}).get(sub_name, {}).get("columns", {})
                )

        wb.save(path)

    # ── UI refresh ────────────────────────────────────────────────────────────

    def _refresh_ui(self):
        # (category colors accumulate across open files so they stay consistent)
        self._close_all_note_floats()   # drop any floats from the previous build
        tabw = self._tab_widget
        tabw.blockSignals(True)
        tabw.clear()
        self._editors.clear()
        tables = list(self.manager.tables.keys())
        for tname in tables:
            self._editors[tname] = None
            tabw.addTab(QWidget(), tname)
        tabw.blockSignals(False)
        self._rebuild_note_tabs()
        self._show_editor()
        if tables:
            QTimer.singleShot(0, lambda: self._ensure_editor(0))
        self._update_sync()
        self.show_snackbar(f"已載入 {len(tables)} 個資料表")

    def _on_tab_changed(self, idx):
        self._ensure_editor(idx)

    def _ensure_editor(self, idx):
        if idx < 0 or idx >= self._tab_widget.count():
            return
        tname = self._tab_widget.tabText(idx)
        if tname not in self.manager.tables:
            return  # note tab (or other non-table page) — nothing to build
        if self._editors.get(tname) is not None:
            return
        editor = TableEditor(tname, self.manager)
        editor.status_message.connect(
            lambda text, color: self.show_snackbar(text, color=color)
        )
        self._editors[tname] = editor          # guard before tab swap
        self._tab_widget.blockSignals(True)
        self._tab_widget.removeTab(idx)
        self._tab_widget.insertTab(idx, editor, tname)
        self._tab_widget.blockSignals(False)
        self._tab_widget.setCurrentIndex(idx)

    # ── Recent / search / config ──────────────────────────────────────────────

    def _show_recent_menu(self):
        recent = self.manager._recent_files
        if not recent:
            self.show_snackbar("沒有最近開啟的檔案"); return
        menu = QMenu(self)
        for path in recent:
            fname = os.path.basename(path)
            menu.addAction(f"{fname}  —  {os.path.dirname(path)}",
                           lambda p=path: self._load_recent(p))
        menu.addSeparator()
        menu.addAction("清除記錄", lambda: (
            self.manager._recent_files.clear(),
            self.manager.save_config(),
            self.show_snackbar("已清除最近記錄"),
        ))
        menu.exec(self.mapToGlobal(self.rect().topLeft()))

    def _show_search(self):
        if not self.manager.json_path:
            self.show_snackbar("尚未載入任何 JSON 檔案"); return
        dlg = getattr(self, "_search_dlg", None)
        if dlg is None:
            dlg = GlobalSearchDialog(self)
            dlg.finished.connect(lambda *_: setattr(self, "_search_dlg", None))
            self._search_dlg = dlg
            dlg.show()
        else:
            dlg.raise_(); dlg.activateWindow()
        dlg.focus_input()

    def navigate_to(self, table_name, is_sub, row_idx, col=None):
        """Jump the main editor to a record (and field). Shared by global-search
        results and the compare view. Sub-table hits open the parent master record
        and switch to the matching sub-table tab."""
        master_table, sub_tab, sub_df = table_name, None, None
        if is_sub:
            if "." not in table_name:
                return
            master_table, sub_tab = table_name.split(".", 1)
            sub_df = self.manager.sub_tables.get(table_name)
            if sub_df is None or row_idx not in sub_df.index:
                return
        if master_table not in self.manager.tables:
            return
        # activate the master tab (build its editor lazily)
        for i in range(self._tab_widget.count()):
            if self._tab_widget.tabText(i) == master_table:
                self._tab_widget.setCurrentIndex(i)
                self._ensure_editor(i)
                break
        else:
            return
        editor = self._editors.get(master_table)
        mdf = self.manager.tables.get(master_table)
        if editor is None or mdf is None:
            return
        if is_sub:
            fk_key = (self.manager.config.get(master_table, {})
                      .get("sub_tables", {}).get(sub_tab, {})
                      .get("foreign_key") or editor.pk_key)
            if fk_key not in sub_df.columns or fk_key not in mdf.columns:
                return
            fk_val = str(sub_df.at[row_idx, fk_key])
            m = mdf.index[mdf[fk_key].astype(str) == fk_val]
            if not len(m):
                return
            master_idx = m[0]
        else:
            if row_idx not in mdf.index:
                return
            master_idx = row_idx
        editor.current_cls_val = mdf.at[master_idx, editor.cls_key]
        editor._load_cls_list()
        editor._load_item_list()
        editor._load_editor(master_idx)
        if is_sub and sub_tab:
            editor.focus_sub(sub_tab, row_idx, col)
        elif col:
            editor.focus_master_field(col)
        self.raise_(); self.activateWindow()

    # ── Notes (text / table / group pages persisted in config, saved w/ the DB) ─
    # Notes are tracked by a stable id so float-out / close / reopen stay robust.
    def _notes_cfg(self):
        """Mutable note list stored in the current file's config (reserved key).
        Also migrates older notes so each has a stable id and an open flag."""
        notes = self.manager.config.get("__notes__")
        if not isinstance(notes, list):
            notes = []
            self.manager.config["__notes__"] = notes
        for n in notes:
            if isinstance(n, dict):
                n.setdefault("id", uuid.uuid4().hex)
                n.setdefault("open", True)
        return notes

    def _note_by_id(self, nid):
        for n in self._notes_cfg():
            if n.get("id") == nid:
                return n
        return None

    def _live_note_pages(self):
        """Every live NotePage — docked tabs plus floated windows."""
        pages = []
        for i in range(self._tab_widget.count()):
            w = self._tab_widget.widget(i)
            if isinstance(w, NotePage):
                pages.append(w)
        for entry in self._note_floats.values():
            if isinstance(entry.get("page"), NotePage):
                pages.append(entry["page"])
        return pages

    def _flush_notes(self):
        """Copy live NotePage widget state back into its config note (by id)."""
        for page in self._live_note_pages():
            note = self._note_by_id(page._note_id)
            if note is None:
                continue
            data = page.to_data()
            data["id"]   = note["id"]
            data["name"] = note.get("name", "筆記")
            data["open"] = note.get("open", True)
            note.clear(); note.update(data)

    def _rebuild_note_tabs(self):
        """Rebuild docked note tabs from config. Flushes live widgets first so no
        in-progress edits are lost, then recreates open, non-floated notes."""
        self._flush_notes()
        for i in range(self._tab_widget.count() - 1, -1, -1):
            w = self._tab_widget.widget(i)
            if isinstance(w, NotePage):
                self._tab_widget.removeTab(i)
                w.deleteLater()
        if not self.manager.json_path:
            return
        floated = set(self._note_floats.keys())
        for note in self._notes_cfg():
            if not note.get("open", True) or note.get("id") in floated:
                continue
            page = NotePage(note)
            page.changed.connect(self._on_note_changed)
            self._tab_widget.addTab(page, "📝 " + (note.get("name") or "筆記"))

    def _on_note_changed(self):
        self.manager.dirty = True
        self._update_sync()

    def _select_note_tab(self, nid):
        for i in range(self._tab_widget.count() - 1, -1, -1):
            w = self._tab_widget.widget(i)
            if isinstance(w, NotePage) and w._note_id == nid:
                self._tab_widget.setCurrentIndex(i); return

    def _note_menu(self):
        """📝 top-bar button: add a note, or reopen a previously-closed one."""
        if not self.manager.json_path:
            self.show_snackbar("尚未載入任何 JSON 檔案"); return
        m = QMenu(self)
        m.addAction("新增筆記…", self._add_note)
        closed = [n for n in self._notes_cfg() if not n.get("open", True)]
        if closed:
            m.addSeparator()
            sub = m.addMenu("重新開啟已關閉")
            for n in closed:
                sub.addAction(n.get("name") or "筆記",
                              lambda _=False, nid=n["id"]: self._reopen_note(nid))
        m.exec(QCursor.pos())

    def _add_note(self):
        if not self.manager.json_path:
            self.show_snackbar("尚未載入任何 JSON 檔案"); return
        name, ok = QInputDialog.getText(self, "新增筆記", "分頁名稱:")
        if not ok or not name.strip():
            return
        kind, ok2 = QInputDialog.getItem(
            self, "新增筆記", "型態:", ["純文字", "表格", "群組（可放小分頁）"], 0, False)
        if not ok2:
            return
        nid = uuid.uuid4().hex
        if kind == "純文字":
            note = {"id": nid, "name": name.strip(), "type": "text",
                    "content": "", "open": True}
        elif kind == "表格":
            note = {"id": nid, "name": name.strip(), "type": "table",
                    "headers": ["欄1", "欄2"], "content": [["", ""], ["", ""]],
                    "open": True}
        else:
            note = {"id": nid, "name": name.strip(), "type": "group",
                    "children": [], "open": True}
        self._notes_cfg().append(note)
        self._rebuild_note_tabs()
        self.manager.dirty = True
        self._update_sync()
        self._select_note_tab(nid)

    def _note_tab_menu(self, pos):
        bar = self._tab_widget.tabBar()
        idx = bar.tabAt(pos)
        w = self._tab_widget.widget(idx) if idx >= 0 else None
        if not isinstance(w, NotePage):
            return
        nid = w._note_id
        m = QMenu(self)
        m.addAction("浮出視窗", lambda: self._float_note(nid))
        m.addAction("重新命名", lambda: self._rename_note(nid))
        m.addAction("關閉分頁", lambda: self._close_note(nid))
        m.addSeparator()
        m.addAction("刪除筆記", lambda: self._delete_note(nid))
        m.exec(bar.mapToGlobal(pos))

    def _rename_note(self, nid):
        note = self._note_by_id(nid)
        if note is None:
            return
        name, ok = QInputDialog.getText(
            self, "重新命名", "分頁名稱:", text=note.get("name", ""))
        if not (ok and name.strip()):
            return
        note["name"] = name.strip()
        if nid in self._note_floats:
            self._note_floats[nid]["win"].setWindowTitle("筆記 — " + name.strip())
        self._rebuild_note_tabs()
        self.manager.dirty = True
        self._update_sync()

    def _delete_note(self, nid):
        note = self._note_by_id(nid)
        if note is None:
            return
        if QMessageBox.question(
                self, "刪除筆記",
                f"確定刪除筆記「{note.get('name', '筆記')}」？") != QMessageBox.Yes:
            return
        if nid in self._note_floats:
            entry = self._note_floats.pop(nid)
            entry["page"].deleteLater()
            entry["win"].blockSignals(True); entry["win"].deleteLater()
        notes = self._notes_cfg()
        for i, n in enumerate(notes):
            if n.get("id") == nid:
                del notes[i]; break
        self._rebuild_note_tabs()
        self.manager.dirty = True
        self._update_sync()

    def _close_note(self, nid):
        note = self._note_by_id(nid)
        if note is None:
            return
        if nid in self._note_floats:
            self._dock_note(nid)   # dock back first (also flushes)
        note["open"] = False
        self._rebuild_note_tabs()
        self.manager.dirty = True
        self._update_sync()

    def _reopen_note(self, nid):
        note = self._note_by_id(nid)
        if note is None:
            return
        note["open"] = True
        self._rebuild_note_tabs()
        self.manager.dirty = True
        self._update_sync()
        self._select_note_tab(nid)

    def _float_note(self, nid):
        note = self._note_by_id(nid)
        if note is None or nid in self._note_floats:
            return
        self._flush_notes()
        page = NotePage(note)
        page.changed.connect(self._on_note_changed)
        win = QDialog(self)
        win.setWindowTitle("筆記 — " + (note.get("name") or ""))
        win.setStyleSheet(APP_QSS)
        win.resize(660, 540)
        lo = QVBoxLayout(win); lo.setContentsMargins(0, 0, 0, 0); lo.setSpacing(0)
        barw = QWidget(); barw.setStyleSheet("background:transparent;")
        bl = QHBoxLayout(barw); bl.setContentsMargins(8, 6, 8, 0)
        dock = _mk_btn("⤡ 收回分頁", "ghost"); dock.setFixedHeight(26)
        dock.clicked.connect(lambda: self._dock_note(nid))
        bl.addStretch(1); bl.addWidget(dock)
        lo.addWidget(barw); lo.addWidget(page, 1)
        win.finished.connect(lambda *_: self._dock_note(nid))
        self._note_floats[nid] = {"win": win, "page": page}
        self._rebuild_note_tabs()   # drop its docked tab
        win.show()

    def _dock_note(self, nid):
        entry = self._note_floats.get(nid)
        if entry is None:
            return
        self._flush_notes()               # capture edits from the floated widget
        del self._note_floats[nid]        # guard against re-entrancy via finished
        entry["page"].deleteLater()
        entry["win"].blockSignals(True); entry["win"].deleteLater()
        self._rebuild_note_tabs()

    def _close_all_note_floats(self):
        """Tear down floating note windows (e.g. on switching data files)."""
        entries = list(self._note_floats.values())
        self._note_floats.clear()
        for entry in entries:
            try:
                entry["page"].deleteLater()
                entry["win"].blockSignals(True); entry["win"].close()
                entry["win"].deleteLater()
            except Exception:
                pass

    def open_config(self):
        idx = self._tab_widget.currentIndex()
        if idx < 0: return
        tname = self._tab_widget.tabText(idx)
        if tname not in self.manager.tables: return
        self._show_config_dialog(tname)

    def open_health_check(self):
        idx = self._tab_widget.currentIndex()
        if idx < 0:
            return
        tname = self._tab_widget.tabText(idx)
        if tname not in self.manager.tables:
            return
        issues = self.manager.health_check(tname)
        self._show_health_dialog(tname, issues)

    # ── Validation rules ───────────────────────────────────────────────────────

    def open_validation_rules(self):
        idx = self._tab_widget.currentIndex()
        if idx < 0:
            self.show_snackbar("尚未載入任何 JSON 檔案"); return
        tname = self._tab_widget.tabText(idx)
        if tname not in self.manager.tables:
            return
        dlg = ValidationRulesDialog(self, self.manager, tname)
        if dlg.exec() != QDialog.Accepted:
            return
        tcfg = self.manager.config.setdefault(tname, {})
        tcfg["validations"] = dlg.rules
        fb = dlg.bindings()
        if fb:
            tcfg["field_bindings"] = fb
        else:
            tcfg.pop("field_bindings", None)
        self.manager.save_config()
        self.manager.validator.reload()   # 違規標記會依新綁定過濾，需重驗
        if self.manager.validator.last_errors:
            names = "、".join(n for n, _ in self.manager.validator.last_errors)
            QMessageBox.warning(self, "表達式錯誤",
                                f"以下規則的表達式無法解析（將不會生效）：\n{names}")
        ed = self._editors.get(tname)
        if ed is not None:
            ed._load_item_list()
            ed._refresh_sub_tables()          # 重套子表隱欄
            ed._refresh_validation_visuals()  # 母表欄位顯隱＋標記
        self.show_snackbar("✓ 驗證規則／欄位綁定已套用", color=_C["green"])

    def _show_validation_gate(self, items, errs, warns):
        """存檔前的驗證結果對話框。回傳 True = 照存（僅警告時可選）。"""
        dlg = QDialog(self)
        dlg.setWindowTitle("資料驗證未通過")
        dlg.setMinimumWidth(560)
        dlg.resize(680, 480)
        dlg.setStyleSheet(APP_QSS)
        outer = QVBoxLayout(dlg)
        outer.setContentsMargins(0, 0, 0, 0); outer.setSpacing(0)

        hdr = QLabel(
            f"發現 <span style='color:{_C['red']}'>{errs} 個錯誤</span>、"
            f"<span style='color:{_C['yellow']}'>{warns} 個警告</span>"
            + ("&nbsp;&nbsp;——&nbsp;錯誤必須修正後才能儲存" if errs else ""))
        hdr.setStyleSheet(
            f"color:{_C['txt']}; font-size:13px; font-weight:600; "
            f"background:{_C['sidebar']}; padding:12px 16px; "
            f"border-bottom:1px solid {_C['border']};")
        outer.addWidget(hdr)

        lst = QListWidget()
        lst.setStyleSheet(
            f"QListWidget{{background:{_C['panel']}; border:none;}}"
            f"QListWidget::item{{padding:8px 12px; border-bottom:1px solid {_C['border']};}}"
            f"QListWidget::item:hover{{background:{_C['card']};}}")
        for it in items:
            sev_dot = "🔴" if it["severity"] == "error" else "🟡"
            where = f"{it['sheet']}" + (f"（{it['pk_val']}）" if it["pk_val"] else "")
            li = QListWidgetItem(
                f"{sev_dot} {it['rule']['name']}    {where} → {', '.join(it['cols'])}")
            li.setToolTip("點擊跳到該儲存格")
            li.setData(Qt.UserRole, it)
            lst.addItem(li)

        def _jump(li):
            it = li.data(Qt.UserRole)
            if not it:
                return
            dlg.reject()
            col = it["cols"][0] if it["cols"] else None
            self.navigate_to(it["sheet"], it["is_sub"], it["row_idx"], col)

        lst.itemClicked.connect(_jump)
        outer.addWidget(lst, 1)

        bb = QWidget(); bb.setStyleSheet(
            f"background:{_C['sidebar']}; border-top:1px solid {_C['border']};")
        bl = QHBoxLayout(bb); bl.setContentsMargins(16, 10, 16, 10)
        bl.addStretch(1)
        result = {"save": False}
        if not errs:
            b_save = _mk_btn("仍要儲存", "danger"); b_save.setFixedHeight(32)

            def _force():
                result["save"] = True
                dlg.accept()

            b_save.clicked.connect(_force)
            bl.addWidget(b_save)
        b_cancel = _mk_btn("取消（回去修正）", "primary"); b_cancel.setFixedHeight(32)
        b_cancel.clicked.connect(dlg.reject)
        bl.addWidget(b_cancel)
        outer.addWidget(bb)
        dlg.exec()
        return result["save"]

    def _show_health_dialog(self, table_name, issues):
        dlg = QDialog(self)
        dlg.setWindowTitle(f"資料健檢 — {table_name}")
        dlg.setMinimumWidth(480)
        dlg.resize(560, 460)
        dlg.setStyleSheet(APP_QSS)
        outer = QVBoxLayout(dlg)
        outer.setContentsMargins(0, 0, 0, 0); outer.setSpacing(0)

        errs = sum(1 for s, _ in issues if s == "error")
        warns = sum(1 for s, _ in issues if s == "warn")
        hdr = QLabel(
            "✓  未發現問題" if not issues
            else f"發現 {errs} 個錯誤、{warns} 個警告"
        )
        hdr.setStyleSheet(
            f"color:{_C['green'] if not issues else _C['txt']}; font-size:13px; "
            f"font-weight:600; background:{_C['sidebar']}; padding:12px 16px; "
            f"border-bottom:1px solid {_C['border']};"
        )
        outer.addWidget(hdr)

        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        content = QWidget(); content.setStyleSheet(f"background:{_C['panel']};")
        vlo = QVBoxLayout(content)
        vlo.setContentsMargins(14, 12, 14, 12); vlo.setSpacing(8)
        if not issues:
            ok = QLabel("主鍵唯一、子表 FK 都對得到母表，沒有空白主鍵/FK。")
            ok.setStyleSheet(f"color:{_C['txt2']}; font-size:12px; background:transparent;")
            ok.setWordWrap(True)
            vlo.addWidget(ok)
        else:
            for sev, msg in issues:
                color = _C["red"] if sev == "error" else _C["yellow"]
                row = QLabel(f'<span style="color:{color}">●</span>&nbsp;&nbsp;'
                             f'<span style="color:{_C["txt"]}">{msg}</span>')
                row.setWordWrap(True)
                row.setStyleSheet("background:transparent; font-size:12px;")
                vlo.addWidget(row)
        vlo.addStretch(1)
        scroll.setWidget(content)
        outer.addWidget(scroll, 1)

        bb = QWidget(); bb.setStyleSheet(f"background:{_C['sidebar']}; border-top:1px solid {_C['border']};")
        bl = QHBoxLayout(bb); bl.setContentsMargins(16, 10, 16, 10); bl.addStretch(1)
        btn = _mk_btn("關閉", "primary"); btn.setFixedHeight(32); btn.clicked.connect(dlg.accept)
        bl.addWidget(btn)
        outer.addWidget(bb)
        dlg.exec()

    def _show_config_dialog(self, table_name):
        _btn_ss = (
            f"background:{_C['input']}; border:1px solid {_C['border']}; "
            f"color:{_C['txtAcc']}; border-radius:5px; padding:3px 10px; text-align:left;"
        )

        # ── Helper: enum options editor button ────────────────────────────────
        def _make_opts_btn(parent_dlg, cur_opts, col_label, df_source=None, col_name_str=None):
            """Return (button, opts_store) where opts_store[0] is the live list."""
            opts_store = [list(cur_opts)]

            def _label():
                n = len(opts_store[0])
                return f"選項: {n}個  ✎" if n else "設定選項…"

            btn = QPushButton(_label())
            btn.setStyleSheet(_btn_ss)

            def _open():
                od = QDialog(parent_dlg)
                od.setWindowTitle(f"Enum 選項 — {col_label}")
                od.resize(340, 420)
                od.setStyleSheet(APP_QSS)
                ov = QVBoxLayout(od)
                ov.setContentsMargins(16, 16, 16, 16); ov.setSpacing(8)

                hint = QLabel("雙擊選項可編輯；拖曳可排序")
                hint.setStyleSheet(f"color:{_C['txt3']}; font-size:11px; background:transparent;")
                ov.addWidget(hint)

                lw = QListWidget()
                lw.setStyleSheet(
                    f"background:{_C['input']}; border:1px solid {_C['border']}; "
                    f"border-radius:5px; color:{_C['txt']};"
                )
                lw.addItems([str(o) for o in opts_store[0]])
                lw.setDragDropMode(QAbstractItemView.InternalMove)
                lw.setSelectionMode(QAbstractItemView.SingleSelection)
                ov.addWidget(lw, 1)

                inp_row = QWidget(); inp_row.setStyleSheet("background:transparent;")
                il = QHBoxLayout(inp_row); il.setContentsMargins(0, 0, 0, 0); il.setSpacing(6)
                inp = QLineEdit(); inp.setPlaceholderText("輸入新選項名稱")
                add_btn = _mk_btn("+ 新增", "primary"); add_btn.setFixedHeight(30)
                def _add():
                    t = inp.text().strip()
                    if t and not any(lw.item(i).text() == t for i in range(lw.count())):
                        lw.addItem(t); inp.clear()
                add_btn.clicked.connect(_add); inp.returnPressed.connect(_add)
                il.addWidget(inp, 1); il.addWidget(add_btn)
                ov.addWidget(inp_row)

                # Auto-collect button: scan df_source column for unique values
                if df_source is not None and col_name_str and col_name_str in df_source.columns:
                    auto_btn = _mk_btn("⟳ 從資料自動收集", "secondary"); auto_btn.setFixedHeight(30)
                    def _auto_collect():
                        existing = {lw.item(i).text() for i in range(lw.count())}
                        vals = df_source[col_name_str].dropna().astype(str).unique()
                        added = 0
                        for v in sorted(vals):
                            v = v.strip()
                            if v and v not in existing:
                                lw.addItem(v)
                                existing.add(v)
                                added += 1
                        if added == 0:
                            hint.setText("（所有現有值已包含在選項中）")
                        else:
                            hint.setText(f"已新增 {added} 個選項")
                    auto_btn.clicked.connect(_auto_collect)
                    ov.addWidget(auto_btn)

                del_btn = _mk_btn("刪除選取項目", "danger"); del_btn.setFixedHeight(30)
                def _del():
                    for it in lw.selectedItems():
                        lw.takeItem(lw.row(it))
                del_btn.clicked.connect(_del)
                ov.addWidget(del_btn)

                bb2 = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                bb2.accepted.connect(od.accept); bb2.rejected.connect(od.reject)
                ov.addWidget(bb2)

                if od.exec() == QDialog.Accepted:
                    opts_store[0] = [lw.item(ii).text() for ii in range(lw.count())]
                    btn.setText(_label())

            btn.clicked.connect(_open)
            return btn, opts_store

        def _col_row(col, cfg_cols, parent_dlg, df_source=None):
            """Return (row_widget, combo, opts_store, note_edit) for one column."""
            rw  = QFrame(); rw.setObjectName("colCard")
            rw.setStyleSheet(
                f"QFrame#colCard {{ background:{_C['card']}; "
                f"border:1px solid {_C['border']}; border-radius:8px; }}")
            _rwv = QVBoxLayout(rw)
            _rwv.setContentsMargins(10, 8, 10, 8); _rwv.setSpacing(5)
            _hdr = QLabel(str(col))
            _hdr.setStyleSheet(
                f"color:{_C['txtAcc']}; font-weight:600; font-size:12px; background:transparent;")
            _rwv.addWidget(_hdr)
            cb = _NoscrollCombo()
            cb.addItems(["string", "int", "float", "bool", "enum", "text_ref", "array"])
            cur_type = cfg_cols.get(col, {}).get("type", "string")
            cb.setCurrentText(cur_type)

            cur_opts = cfg_cols.get(col, {}).get("options", [])
            opts_btn, opts_store = _make_opts_btn(
                parent_dlg, cur_opts, col,
                df_source=df_source, col_name_str=col
            )
            opts_btn.setVisible(cur_type == "enum")
            cb.currentTextChanged.connect(lambda t, ob=opts_btn: ob.setVisible(t == "enum"))

            note_edit = QTextEdit()
            note_edit.setAcceptRichText(False)
            note_edit.setFixedHeight(50)
            note_edit.setPlaceholderText("欄位備註（可換行；滑鼠停留欄位標題時顯示）")
            note_edit.setPlainText(cfg_cols.get(col, {}).get("note", ""))

            suggest_combo = _NoscrollCombo()
            suggest_combo.setToolTip("依此鄰欄的值過濾建議清單（空白為不啟用）")
            suggest_combo.setMaximumWidth(130)
            suggest_combo.addItem("(無建議)", "")
            if df_source is not None:
                for sc in df_source.columns:
                    if str(sc) != col:
                        suggest_combo.addItem(str(sc), str(sc))
            cur_sg = cfg_cols.get(col, {}).get("suggest_from", "")
            ix = suggest_combo.findData(cur_sg) if cur_sg else 0
            suggest_combo.setCurrentIndex(max(ix, 0))

            def _labeled(label_text, *widgets, stretch_idx=None):
                row = QWidget(); row.setStyleSheet("background:transparent;")
                h = QHBoxLayout(row); h.setContentsMargins(0, 0, 0, 0); h.setSpacing(6)
                _l = QLabel(label_text); _l.setFixedWidth(60)
                _l.setStyleSheet(f"color:{_C['txt3']}; font-size:11px; background:transparent;")
                h.addWidget(_l, 0, Qt.AlignTop)
                for i, wd in enumerate(widgets):
                    h.addWidget(wd, 1 if i == stretch_idx else 0, Qt.AlignTop)
                if stretch_idx is None:
                    h.addStretch(1)
                return row

            _rwv.addWidget(_labeled("型別", cb, opts_btn))
            _rwv.addWidget(_labeled("備註", note_edit, stretch_idx=0))
            _rwv.addWidget(_labeled("建議來源", suggest_combo))

            # per-column external text-ref source (only shown when type == text_ref)
            _tref = cfg_cols.get(col, {}).get("text_ref", {}) or {}
            tr_path = QLineEdit(_tref.get("json_path", ""))
            tr_path.setPlaceholderText("外部文字表路徑（相對/絕對）")
            tr_browse = QPushButton("…"); tr_browse.setFixedWidth(28); tr_browse.setAutoDefault(False)
            tr_browse.setStyleSheet(
                f"background:{_C['card']}; border:1px solid {_C['border']}; "
                f"color:{_C['txt']}; border-radius:5px;")
            def _tr_browse(edit=tr_path):
                base = os.path.dirname(self.manager.json_path) if self.manager.json_path else ""
                p, _ = QFileDialog.getOpenFileName(parent_dlg, "選擇外部文字表", base,
                                                   "JSON (*.json);;所有檔案 (*.*)")
                if p:
                    try: p = os.path.relpath(p, base) if base else p
                    except ValueError: pass
                    edit.setText(p)
            tr_browse.clicked.connect(lambda *_a, _b=_tr_browse: _b())
            tr_key = QLineEdit(_tref.get("key_col", "") or ""); tr_key.setMaximumWidth(120)
            tr_key.setPlaceholderText("key欄(預設TextID)")
            tr_val = QLineEdit(_tref.get("val_col", "") or ""); tr_val.setMaximumWidth(150)
            tr_val.setPlaceholderText("val欄(預設TextContent)")
            tr_row = _labeled("文字表", tr_path, tr_browse, tr_key, tr_val, stretch_idx=0)
            tr_row.setVisible(cur_type == "text_ref")
            cb.currentTextChanged.connect(lambda t, _w=tr_row: _w.setVisible(t == "text_ref"))
            _rwv.addWidget(tr_row)
            return rw, cb, opts_store, note_edit, suggest_combo, (tr_path, tr_key, tr_val)

        # ── Dialog ────────────────────────────────────────────────────────────
        dlg = QDialog(self)
        dlg.setWindowTitle(f"配置 — {table_name}")
        dlg.setMinimumWidth(560)
        dlg.resize(720, 680)
        dlg.setStyleSheet(APP_QSS)

        cfg      = self.manager.config.get(table_name, {})
        df       = self.manager.tables[table_name]
        cols     = list(df.columns)
        cols_cfg = cfg.get("columns", {})

        outer = QVBoxLayout(dlg)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        content = QWidget(); content.setStyleSheet(f"background:{_C['panel']};")
        vlo = QVBoxLayout(content)
        vlo.setContentsMargins(16, 16, 16, 16)
        vlo.setSpacing(12)
        scroll.setWidget(content)
        outer.addWidget(scroll, 1)

        # ── Main table keys ───────────────────────────────────────────────────
        def _sec(text):
            lbl = QLabel(text.upper())
            lbl.setStyleSheet(
                f"color:{_C['txt3']}; font-size:10px; font-weight:600; "
                f"letter-spacing:1px; background:transparent;"
            )
            return lbl

        def _form_row(label_text, widget):
            w = QWidget(); w.setStyleSheet("background:transparent;")
            h = QHBoxLayout(w); h.setContentsMargins(0,0,0,0); h.setSpacing(10)
            lbl = QLabel(label_text)
            lbl.setFixedWidth(160)
            lbl.setStyleSheet(f"color:{_C['txt2']}; background:transparent;")
            h.addWidget(lbl); h.addWidget(widget, 1)
            return w

        vlo.addWidget(_sec("主表設定"))
        pk_var  = _NoscrollCombo(); pk_var.addItems(cols)
        cls_var = _NoscrollCombo(); cls_var.addItems(cols)
        pk_var.setCurrentText(cfg.get("primary_key",         cols[0] if cols else ""))
        cls_var.setCurrentText(cfg.get("classification_key", cols[0] if cols else ""))
        vlo.addWidget(_form_row("Primary Key",         pk_var))
        vlo.addWidget(_form_row("Classification Key",  cls_var))

        # ── Browse helper (used by image folder + text-ref) ──────────────────
        def _browse_row(edit, is_folder=False):
            row = QWidget(); row.setStyleSheet("background:transparent;")
            rl = QHBoxLayout(row); rl.setContentsMargins(0,0,0,0); rl.setSpacing(4)
            btn = QPushButton("…"); btn.setFixedWidth(32)
            btn.setStyleSheet(
                f"background:{_C['card']}; border:1px solid {_C['border']}; "
                f"color:{_C['txt']}; border-radius:5px;"
            )
            def _browse():
                base = os.path.dirname(self.manager.json_path) if self.manager.json_path else ""
                if is_folder:
                    from PySide6.QtWidgets import QFileDialog as _QFD
                    p = _QFD.getExistingDirectory(dlg, "選擇資料夾", base)
                else:
                    from PySide6.QtWidgets import QFileDialog as _QFD
                    p, _ = _QFD.getOpenFileName(dlg, "選擇檔案", base, "JSON (*.json)")
                if p:
                    try: p = os.path.relpath(p, base) if base else p
                    except ValueError: pass
                    edit.setText(p)
            btn.clicked.connect(_browse)
            rl.addWidget(edit, 1); rl.addWidget(btn)
            return row

        # Image base folder (first row)
        img_folder_edit = QLineEdit(cfg.get("image_preview", {}).get("base_folder", ""))
        img_folder_edit.setPlaceholderText("圖片根目錄（相對路徑或絕對路徑）")
        vlo.addWidget(_form_row("Image 資料夾路徑", _browse_row(img_folder_edit, is_folder=True)))

        img_ext_edit = QLineEdit(cfg.get("image_preview", {}).get("ext", ""))
        img_ext_edit.setPlaceholderText("副檔名，例如 .png")
        vlo.addWidget(_form_row("Image 副檔名", img_ext_edit))

        # ── Image path segments builder (second row) ───────────────────────────
        _img_segs_container = QWidget(); _img_segs_container.setStyleSheet("background:transparent;")
        _img_segs_lo = QVBoxLayout(_img_segs_container)
        _img_segs_lo.setContentsMargins(0, 0, 0, 0); _img_segs_lo.setSpacing(2)
        _img_segs_rows = []  # list of (type_cb, val_stack, col_combo, lit_edit, row_w)

        _img_rows_lo = QVBoxLayout()
        _img_segs_lo.addLayout(_img_rows_lo)

        def _add_img_seg_row(seg_type="col", seg_value=""):
            row_w = QWidget(); row_w.setStyleSheet("background:transparent;")
            rl = QHBoxLayout(row_w); rl.setContentsMargins(0, 0, 0, 0); rl.setSpacing(4)
            type_cb = _NoscrollCombo()
            type_cb.addItems(["欄位", "字串"])
            type_cb.setFixedWidth(58)
            col_combo = _NoscrollCombo()
            col_combo.addItems(cols)
            lit_edit = QLineEdit()
            lit_edit.setPlaceholderText("字串值")
            val_stack = QStackedWidget()
            val_stack.setStyleSheet("background:transparent;")
            val_stack.addWidget(col_combo)   # index 0
            val_stack.addWidget(lit_edit)    # index 1
            if seg_type == "lit":
                type_cb.setCurrentIndex(1)
                val_stack.setCurrentIndex(1)
                lit_edit.setText(seg_value)
            else:
                type_cb.setCurrentIndex(0)
                val_stack.setCurrentIndex(0)
                if seg_value in cols:
                    col_combo.setCurrentText(seg_value)
            type_cb.currentIndexChanged.connect(val_stack.setCurrentIndex)
            del_btn = QPushButton("−")
            del_btn.setFixedSize(24, 24)
            del_btn.setStyleSheet(
                f"background:{_C['card']}; border:1px solid {_C['border']}; "
                f"color:{_C['red']}; border-radius:4px; font-weight:700;"
            )
            rl.addWidget(type_cb)
            rl.addWidget(val_stack, 1)
            rl.addWidget(del_btn)
            entry = (type_cb, val_stack, col_combo, lit_edit, row_w)
            _img_segs_rows.append(entry)
            _img_rows_lo.addWidget(row_w)
            def _del_seg(e=entry, w=row_w):
                if e in _img_segs_rows:
                    _img_segs_rows.remove(e)
                w.hide()
                w.deleteLater()
            del_btn.clicked.connect(_del_seg)

        # Load existing segments (backward compat with old "col" key)
        _cur_img_segs = cfg.get("image_preview", {}).get("path_segments", [])
        if not _cur_img_segs:
            _old_img_col = cfg.get("image_preview", {}).get("col", "")
            if _old_img_col:
                _cur_img_segs = [{"type": "col", "col": _old_img_col}]
        for _s in _cur_img_segs:
            if _s.get("type") == "col":
                _add_img_seg_row("col", _s.get("col", ""))
            else:
                _add_img_seg_row("lit", _s.get("value", ""))

        _add_seg_btn = QPushButton("＋ 加段")
        _add_seg_btn.setFixedHeight(26)
        _add_seg_btn.setStyleSheet(
            f"background:{_C['card']}; border:1px solid {_C['border']}; "
            f"color:{_C['txt2']}; border-radius:5px; font-size:11px;"
        )
        _add_seg_btn.clicked.connect(lambda: _add_img_seg_row("col", ""))
        _img_segs_lo.addWidget(_add_seg_btn)

        vlo.addWidget(_form_row("Image 路徑結構", _img_segs_container))

        # （外部文字表已改為「逐欄位」設定：在下方各欄位 type 選 text_ref 後即可填）
        _trs_hint = QLabel("※ 外部文字表改為逐欄位設定：把欄位型別設成 text_ref，下方就會出現該欄的文字表路徑/key/val")
        _trs_hint.setWordWrap(True)
        _trs_hint.setStyleSheet(f"color:{_C['txt3']}; font-size:11px; background:transparent;")
        vlo.addWidget(_trs_hint)

        sep1 = QFrame(); sep1.setFixedHeight(1)
        sep1.setStyleSheet(f"background:{_C['border']}; border:none;")
        vlo.addWidget(sep1)
        vlo.addWidget(_sec("主表欄位類型"))

        main_col_widgets: dict[str, tuple] = {}  # col → (cb, opts_store)
        for col in cols:
            rw, cb, opts_store, note_edit, suggest_combo, tref = _col_row(col, cols_cfg, dlg, df_source=df)
            vlo.addWidget(rw)
            main_col_widgets[col] = (cb, opts_store, note_edit, suggest_combo, tref)

        # ── Sub-tables ────────────────────────────────────────────────────────
        prefix = table_name + "."
        sub_keys = [k for k in self.manager.sub_tables if k.startswith(prefix)]
        sub_widgets: dict[str, dict] = {}  # tab_name → {fk_edit, col_combos}

        if sub_keys:
            sep2 = QFrame(); sep2.setFixedHeight(1)
            sep2.setStyleSheet(f"background:{_C['border']}; border:none;")
            vlo.addSpacing(4); vlo.addWidget(sep2)
            vlo.addWidget(_sec("從表設定"))

            sub_cfg_root = cfg.get("sub_tables", {})

            for full_key in sub_keys:
                tab_name     = full_key[len(prefix):]
                sub_df       = self.manager.sub_tables[full_key]
                sub_cfg      = sub_cfg_root.get(tab_name, {})
                sub_cols_cfg = sub_cfg.get("columns", {})

                shdr = QLabel(f"▸  {tab_name}")
                shdr.setStyleSheet(
                    f"color:{_C['txtAcc']}; font-size:12px; font-weight:600; "
                    f"background:transparent; padding-top:6px;"
                )
                vlo.addWidget(shdr)

                fk_edit = QLineEdit()
                fk_edit.setPlaceholderText("foreign_key 欄位名稱")
                fk_edit.setText(sub_cfg.get("foreign_key", ""))
                vlo.addWidget(_form_row("  Foreign Key", fk_edit))

                st_note_edit = QLineEdit()
                st_note_edit.setPlaceholderText("這張子表的用途說明（滑鼠停在分頁標題時顯示）")
                st_note_edit.setText(sub_cfg.get("note", ""))
                vlo.addWidget(_form_row("  說明 / 備註", st_note_edit))

                col_combos: dict[str, tuple] = {}
                for scol in list(sub_df.columns):
                    rw, cb, opts_store, note_edit, suggest_combo, tref = _col_row(scol, sub_cols_cfg, dlg, df_source=sub_df)
                    vlo.addWidget(rw)
                    col_combos[scol] = (cb, opts_store, note_edit, suggest_combo, tref)

                sub_widgets[tab_name] = {"fk_edit": fk_edit, "note_edit": st_note_edit,
                                         "col_combos": col_combos}

        else:
            # Inform user if no sub-tables detected
            no_sub = QLabel("（此表格在 JSON 中無巢狀陣列資料，故無從表）")
            no_sub.setStyleSheet(f"color:{_C['txt3']}; font-size:11px; background:transparent;")
            vlo.addWidget(no_sub)

        vlo.addStretch(1)

        # ── Buttons ───────────────────────────────────────────────────────────
        bb_w = QWidget()
        bb_w.setStyleSheet(
            f"background:{_C['sidebar']}; border-top:1px solid {_C['border']};"
        )
        bb_lo = QHBoxLayout(bb_w)
        bb_lo.setContentsMargins(16, 10, 16, 10)
        bb_lo.setSpacing(8)
        bb_lo.addStretch(1)
        btn_ok  = _mk_btn("套用", "primary"); btn_ok.setFixedHeight(34)
        btn_can = _mk_btn("取消");             btn_can.setFixedHeight(34)
        btn_ok.clicked.connect(dlg.accept)
        btn_can.clicked.connect(dlg.reject)
        bb_lo.addWidget(btn_can); bb_lo.addWidget(btn_ok)
        outer.addWidget(bb_w)

        if dlg.exec() != QDialog.Accepted:
            return

        # ── Apply ─────────────────────────────────────────────────────────────
        cfg["primary_key"]        = pk_var.currentText()
        cfg["classification_key"] = cls_var.currentText()
        img_folder_val = img_folder_edit.text().strip()
        _new_segs = []
        for (tcb, vstk, ccb, ledit, rw) in _img_segs_rows:
            if tcb.currentIndex() == 0:
                _new_segs.append({"type": "col", "col": ccb.currentText()})
            else:
                _new_segs.append({"type": "lit", "value": ledit.text()})
        img_ext_val = img_ext_edit.text().strip()
        if _new_segs:
            cfg["image_preview"] = {"path_segments": _new_segs}
            if img_folder_val:
                cfg["image_preview"]["base_folder"] = img_folder_val
            if img_ext_val:
                cfg["image_preview"]["ext"] = img_ext_val
        else:
            cfg.pop("image_preview", None)

        cfg.pop("text_ref_source", None)  # legacy table-level key removed (now per-column)

        def _build_col_entry(t, opts_store, note="", suggest_from="", tref=None):
            entry = {"type": t}
            if t == "enum" and opts_store[0]:
                entry["options"] = opts_store[0]
            note = (note or "").strip()
            if note:
                entry["note"] = note
            sg = (suggest_from or "").strip()
            if sg:
                entry["suggest_from"] = sg
            if t == "text_ref" and tref is not None:
                path = tref[0].text().strip()
                if path:
                    entry["text_ref"] = {
                        "json_path": path,
                        "key_col":   tref[1].text().strip() or "TextID",
                        "val_col":   tref[2].text().strip() or "TextContent",
                    }
            return entry

        cfg.setdefault("columns", {})
        for col, (cb, opts_store, note_edit, suggest_combo, tref) in main_col_widgets.items():
            cfg["columns"][col] = _build_col_entry(
                cb.currentText(), opts_store, note_edit.toPlainText(),
                suggest_combo.currentData(), tref
            )

        cfg.setdefault("sub_tables", {})
        for tab_name, data in sub_widgets.items():
            st = cfg["sub_tables"].setdefault(tab_name, {})
            fk = data["fk_edit"].text().strip()
            if fk:
                st["foreign_key"] = fk
            st_note = data["note_edit"].text().strip()
            if st_note:
                st["note"] = st_note
            else:
                st.pop("note", None)
            st.setdefault("columns", {})
            for scol, (scb, opts_store, note_edit, suggest_combo, tref) in data["col_combos"].items():
                st["columns"][scol] = _build_col_entry(
                    scb.currentText(), opts_store, note_edit.toPlainText(),
                    suggest_combo.currentData(), tref
                )

        self.manager.config[table_name] = cfg
        self.manager.save_config()
        self.manager.validator.reload()   # 欄位型別/子表設定變了 → 規則重編譯重驗
        editor = self._editors.get(table_name)
        if editor:
            editor.reload_after_config()
        self.show_snackbar("配置已套用")

    # ── Window close ──────────────────────────────────────────────────────────

    def closeEvent(self, event):
        dirty = [d for d in self._docs if d.manager.dirty]
        if dirty:
            names = "、".join(
                os.path.basename(d.manager.json_path or "未命名") for d in dirty)
            ans = QMessageBox.question(
                self, "未儲存變更", f"有未儲存的變更（{names}），是否全部儲存？",
                QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel
            )
            if ans == QMessageBox.Cancel:
                event.ignore(); return
            if ans == QMessageBox.Save:
                for i, d in enumerate(self._docs):
                    if not d.manager.dirty:
                        continue
                    self._active = i            # so _flush_notes targets this doc
                    self._flush_notes()
                    try:
                        d.manager.save_config()
                        d.manager.save_json()   # synchronous save on shutdown
                    except Exception:
                        pass
        event.accept()


# ── Crash / error logging ─────────────────────────────────────────────────────

def _log_path():
    """log.txt next to the executable (packaged) or the script (dev)."""
    try:
        if getattr(sys, "frozen", False) or "__compiled__" in globals():
            base = os.path.dirname(sys.executable)
        else:
            base = os.path.dirname(os.path.abspath(__file__))
    except Exception:
        base = os.getcwd()
    return os.path.join(base, "log.txt")


def _write_log(header, text):
    try:
        with open(_log_path(), "a", encoding="utf-8") as f:
            ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"\n===== {ts} | {header} =====\n{text}\n")
    except Exception:
        pass


def _excepthook(etype, value, tb):
    _write_log("未捕捉的例外 (Unhandled exception)",
               "".join(traceback.format_exception(etype, value, tb)))
    try:
        sys.__excepthook__(etype, value, tb)
    except Exception:
        pass


if __name__ == "__main__":
    # faulthandler 抓硬性崩潰(segfault)，excepthook 抓 Python 例外，都寫進 log.txt
    try:
        import faulthandler
        _flog = open(_log_path(), "a", encoding="utf-8", buffering=1)
        faulthandler.enable(_flog)
    except Exception:
        pass
    sys.excepthook = _excepthook

    try:
        app = QApplication(sys.argv)
        app.setStyle("Fusion")
        app.setStyleSheet(APP_QSS)
        window = App()
        window.show()
        sys.exit(app.exec())
    except SystemExit:
        raise
    except BaseException:
        _write_log("啟動 / 主迴圈崩潰", traceback.format_exc())
        raise
